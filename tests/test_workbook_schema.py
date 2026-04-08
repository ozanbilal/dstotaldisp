from pathlib import Path

from openpyxl import load_workbook

from disp_core import process_batch_directory


ROOT = Path(__file__).resolve().parents[1]


def _sheetnames(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    return workbook.sheetnames


def _sheet_header(path: Path, sheet_name: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook[sheet_name]
    return [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def test_pair_and_method_outputs_do_not_write_extra_sheets(tmp_path):
    output_dir = tmp_path / "pair_out"
    summary = process_batch_directory(
        ROOT / "_tmp_m23_case_pair_in",
        output_dir,
        {
            "method2Enabled": True,
            "method3Enabled": True,
        },
    )

    output_paths = {result["outputFileName"]: Path(result["writtenPath"]) for result in summary["results"]}

    assert _sheetnames(output_paths["output_total_Results_profile_1_motion_DD1_X_20030501002708_1201_H1.xlsx"]) == [
        "Strain_Relative",
        "Legacy_Methods",
        "Comparison",
        "Depth_Profiles",
        "Profile_BaseCorrected",
        "Direction_X_Time",
        "Direction_Y_Time",
        "Resultant_Time",
        "TBDY_Total_X_Time",
        "TBDY_Total_Y_Time",
        "TBDY_Total_Resultant_Time",
    ]
    assert _sheetnames(output_paths["output_method2_Results_profile_1_motion_DD1_X_20030501002708_1201_H1.xlsx"]) == [
        "Method2_TBDY_X_Time",
        "Method2_Metadata",
    ]
    assert _sheetnames(output_paths["output_method2_Results_profile_1_motion_DD1_Y_20030501002708_1201_H2.xlsx"]) == [
        "Method2_TBDY_Y_Time",
        "Method2_Metadata",
    ]
    assert _sheetnames(output_paths["output_method3_profiles_all.xlsx"]) == [
        "Method3_Profile_X",
        "Method3_Profile_Y",
        "Method3_ApproxTotal_X",
        "Method3_ApproxTotal_Y",
    ]
    assert _sheet_header(
        output_paths["output_method3_profiles_all.xlsx"],
        "Method3_ApproxTotal_X",
    )[0] == "Depth_m"


def test_single_output_does_not_write_profile_verification_sheet(tmp_path):
    output_dir = tmp_path / "single_out"
    summary = process_batch_directory(
        ROOT / "_tmp_single_in",
        output_dir,
        {
            "method2Enabled": False,
            "method3Enabled": False,
        },
    )

    output_path = Path(summary["results"][0]["writtenPath"])
    assert _sheetnames(output_path) == [
        "Single_Direction_Summary",
        "Direction_Time",
        "Strain_Relative_Time",
        "TBDY_Total_Time",
        "InputProxy_Relative_Time",
    ]
