import base64
import io
import math
import os
import re
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl.chart import LineChart, Reference, ScatterChart, Series

try:
    from scipy.signal import bessel as _scipy_bessel
    from scipy.signal import butter as _scipy_butter
    from scipy.signal import cheby1 as _scipy_cheby1
    from scipy.signal import filtfilt as _scipy_filtfilt
    from scipy.signal import lfilter as _scipy_lfilter

    HAS_SCIPY = True
except Exception:  # noqa: BLE001
    _scipy_bessel = None
    _scipy_butter = None
    _scipy_cheby1 = None
    _scipy_filtfilt = None
    _scipy_lfilter = None
    HAS_SCIPY = False


EXCLUDE_PREFIXES = ("output_", "~$")
EXCLUDE_SUFFIXES = ("-manip.xlsx",)
DB_SUFFIXES = (".db", ".db3")
XLSX_SUFFIX = ".xlsx"
DEFAULT_HIGHPASS_CUTOFF_HZ = 0.03
DEFAULT_HIGHPASS_TRANSITION_HZ = 0.02
DEFAULT_FILTER_LOW_HZ = 0.10
DEFAULT_FILTER_HIGH_HZ = 25.0
DEFAULT_FILTER_ORDER = 4
DEFAULT_BASELINE_DEGREE = 4
DEFAULT_BASE_REFERENCE = "input"
DEFAULT_ALT_INTEGRATION_METHOD = "fft_regularized"
DEFAULT_ALT_LOWCUT_HZ = 0.05
ALT_LOWCUT_POLICY = "from_filter_or_default_0p05"
PREVIEW_MAX_POINTS = 160
PREVIEW_MAX_SERIES = 5
SOURCE_VIEW_MAX_POINTS = 420
CALC_PROGRESS_START = 55.0
CALC_PROGRESS_END = 92.0
SUMMARY_CONFIDENCE_LABELS = {
    10: "En guvenilir",
    20: "Yuksek guven",
    30: "Orta guven",
    40: "Yaklasik",
    50: "Dolayli",
    60: "Referans",
}
SUMMARY_SOURCE_SYSTEM = "deepsoil"
SUMMARY_METHOD_METADATA = {
    "db_direct_total": {
        "displayLabel": "DB Direct Total",
        "methodClass": "direct",
        "confidenceRank": 10,
        "sourceRefs": ["VEL_DISP.LAYERn_DISP_TOTAL"],
    },
    "strain_input_total": {
        "displayLabel": "Strain + Input Proxy",
        "methodClass": "computed",
        "confidenceRank": 20,
        "sourceRefs": ["Layer Strain", "Input Motion"],
    },
    "strain_deepest_total": {
        "displayLabel": "Strain + Deepest Layer Proxy",
        "methodClass": "computed",
        "confidenceRank": 30,
        "sourceRefs": ["Layer Strain", "Deepest Layer Acceleration"],
    },
    "profile_offset_total": {
        "displayLabel": "Profile Offset Approximation",
        "methodClass": "approximate",
        "confidenceRank": 40,
        "sourceRefs": ["Profile", "Bottom Offset"],
    },
    "time_history_total": {
        "displayLabel": "Time-History Indirect Total",
        "methodClass": "indirect",
        "confidenceRank": 50,
        "sourceRefs": ["Layer Acceleration"],
    },
    "profile_reference_total": {
        "displayLabel": "Profile Reference",
        "methodClass": "reference",
        "confidenceRank": 60,
        "sourceRefs": ["Profile Maximum Displacement"],
    },
}


def _log(logs: List[Dict[str, str]], level: str, message: str) -> None:
    logs.append({"level": level, "message": message})


def _safe_float_list(values: Any) -> List[float | None]:
    arr = np.asarray(values, dtype=float).reshape(-1)
    out: List[float | None] = []
    for item in arr:
        out.append(float(item) if np.isfinite(item) else None)
    return out


def _series_has_finite_values(values: Any) -> bool:
    arr = np.asarray(values, dtype=float).reshape(-1)
    return bool(arr.size and np.any(np.isfinite(arr)))


def _normalize_source_refs(source_refs: Sequence[str] | None, fallback: Sequence[str] | None = None) -> List[str]:
    refs = source_refs if source_refs is not None else fallback
    if not refs:
        return []
    return [str(item) for item in refs if str(item or "").strip()]


def _build_method_variant(
    variant_key: str,
    depths: Any,
    values: Any,
    *,
    valid: bool,
    reason: str,
    display_label: str | None = None,
    method_class: str | None = None,
    confidence_rank: int | None = None,
    source_refs: Sequence[str] | None = None,
) -> Dict[str, Any]:
    meta = SUMMARY_METHOD_METADATA.get(variant_key, {})
    resolved_rank = int(confidence_rank if confidence_rank is not None else meta.get("confidenceRank", 90))
    resolved_label = str(display_label or meta.get("displayLabel", variant_key))
    resolved_class = str(method_class or meta.get("methodClass", "derived"))
    return {
        "variantKey": str(variant_key),
        "displayLabel": resolved_label,
        "methodClass": resolved_class,
        "depths": _safe_float_list(depths),
        "values": _safe_float_list(values),
        "valid": bool(valid),
        "confidenceRank": resolved_rank,
        "confidenceLabel": SUMMARY_CONFIDENCE_LABELS.get(resolved_rank, "Destekleyici"),
        "reason": str(reason or ""),
        "sourceRefs": _normalize_source_refs(source_refs, meta.get("sourceRefs")),
    }


def _pick_primary_variant_key(variants: Sequence[Mapping[str, Any]]) -> str:
    valid_variants = [
        item
        for item in variants
        if bool(item.get("valid")) and _series_has_finite_values(item.get("values", []))
    ]
    if not valid_variants:
        return ""
    valid_variants.sort(key=lambda item: (int(item.get("confidenceRank", 999)), str(item.get("displayLabel", ""))))
    return str(valid_variants[0].get("variantKey", ""))


def _resultant_profile_from_components(x_values: Any, y_values: Any) -> np.ndarray:
    x_arr = np.asarray(x_values, dtype=float).reshape(-1)
    y_arr = np.asarray(y_values, dtype=float).reshape(-1)
    count = min(int(x_arr.size), int(y_arr.size))
    if count <= 0:
        return np.zeros(0, dtype=float)
    return np.sqrt(x_arr[:count] ** 2 + y_arr[:count] ** 2)


def _matrix_resultant_envelope(x_matrix: Any, y_matrix: Any) -> np.ndarray:
    x_arr = np.asarray(x_matrix, dtype=float)
    y_arr = np.asarray(y_matrix, dtype=float)
    if x_arr.ndim != 2 or y_arr.ndim != 2:
        return np.zeros(0, dtype=float)
    layer_count = min(int(x_arr.shape[0]), int(y_arr.shape[0]))
    step_count = min(int(x_arr.shape[1]), int(y_arr.shape[1]))
    if layer_count <= 0 or step_count <= 0:
        return np.zeros(0, dtype=float)
    resultant = np.sqrt(x_arr[:layer_count, :step_count] ** 2 + y_arr[:layer_count, :step_count] ** 2)
    return np.max(resultant, axis=1)


def _to_number_series(series: pd.Series) -> np.ndarray:
    numeric = pd.to_numeric(series, errors="coerce")
    return numeric.dropna().to_numpy(dtype=float)


def _cumtrapz(y: np.ndarray, x: np.ndarray) -> np.ndarray:
    if y.size == 0:
        return np.array([], dtype=float)
    if y.size == 1:
        return np.array([0.0], dtype=float)
    dx = np.diff(x)
    area = 0.5 * (y[1:] + y[:-1]) * dx
    return np.concatenate(([0.0], np.cumsum(area)))


def _baseline_correct_legacy(acc: np.ndarray, time: np.ndarray) -> np.ndarray:
    if acc.size < 4:
        return acc - np.mean(acc)
    coeff = np.polyfit(time, acc, 3)
    baseline = np.polyval(coeff, time)
    return acc - baseline


def _soft_highpass_fft(
    signal: np.ndarray,
    time: np.ndarray,
    cutoff_hz: float = DEFAULT_HIGHPASS_CUTOFF_HZ,
    transition_hz: float = DEFAULT_HIGHPASS_TRANSITION_HZ,
) -> np.ndarray:
    x = signal.astype(float)
    if x.size < 8:
        return x - np.mean(x)

    dt = float(np.median(np.diff(time.astype(float))))
    if not np.isfinite(dt) or dt <= 0:
        return x - np.mean(x)

    x = x - np.mean(x)
    n = x.size
    freqs = np.fft.rfftfreq(n, d=dt)
    if freqs.size <= 1:
        return x

    nyquist = 0.5 / dt
    cutoff = float(np.clip(cutoff_hz, 0.0, max(0.0, nyquist * 0.999)))
    transition = max(0.0, float(transition_hz))
    stop = max(0.0, cutoff - transition)

    if cutoff <= 0.0:
        return x

    transfer = np.ones_like(freqs)
    transfer[freqs <= stop] = 0.0

    if cutoff > stop:
        mask = (freqs > stop) & (freqs < cutoff)
        xi = (freqs[mask] - stop) / (cutoff - stop)
        transfer[mask] = 0.5 - 0.5 * np.cos(np.pi * xi)

    spectrum = np.fft.rfft(x)
    filtered = np.fft.irfft(spectrum * transfer, n=n)
    return filtered.astype(float)


def _detrend_poly(data: np.ndarray, degree: int = DEFAULT_BASELINE_DEGREE) -> np.ndarray:
    if data is None or data.size == 0:
        return data
    if data.size <= max(1, int(degree)):
        return data - np.mean(data)

    x = np.arange(1, data.size + 1, dtype=float)
    x_mean = np.mean(x)
    x_std = np.std(x)
    if x_std == 0:
        return data - np.mean(data)

    x = (x - x_mean) / x_std
    try:
        coeffs = np.polyfit(x, data, max(1, int(degree)))
        return data - np.polyval(coeffs, x)
    except Exception:  # noqa: BLE001
        return data - np.mean(data)


def _apply_baseline(data: np.ndarray, method: str, degree: int = DEFAULT_BASELINE_DEGREE) -> np.ndarray:
    if data is None or data.size == 0:
        return data

    m = str(method or "poly4").strip().lower()
    if m in {"none", "", "raw"}:
        return data
    if m in {"mean", "dc"}:
        return data - np.mean(data)

    if m.startswith("poly"):
        digits = "".join(ch for ch in m if ch.isdigit())
        if digits:
            try:
                degree = int(digits)
            except ValueError:
                degree = max(1, int(degree))

    return _detrend_poly(data, max(1, int(degree)))


def _build_highpass_transfer(freqs: np.ndarray, cutoff_hz: float, transition_hz: float) -> np.ndarray:
    transfer = np.ones_like(freqs, dtype=float)
    cutoff = max(0.0, float(cutoff_hz))
    transition = max(1e-9, float(transition_hz))

    if cutoff <= 0:
        return transfer

    stop = max(0.0, cutoff - transition)
    transfer[freqs <= stop] = 0.0
    if cutoff > stop:
        mask = (freqs > stop) & (freqs < cutoff)
        xi = (freqs[mask] - stop) / (cutoff - stop)
        transfer[mask] = 0.5 - 0.5 * np.cos(np.pi * xi)
    return transfer


def _build_lowpass_transfer(freqs: np.ndarray, cutoff_hz: float, transition_hz: float) -> np.ndarray:
    transfer = np.ones_like(freqs, dtype=float)
    cutoff = max(0.0, float(cutoff_hz))
    transition = max(1e-9, float(transition_hz))

    if cutoff <= 0:
        transfer[:] = 0.0
        return transfer

    stop = cutoff + transition
    transfer[freqs >= stop] = 0.0
    if stop > cutoff:
        mask = (freqs > cutoff) & (freqs < stop)
        xi = (freqs[mask] - cutoff) / (stop - cutoff)
        transfer[mask] = 0.5 + 0.5 * np.cos(np.pi * xi)
    return transfer


def _fft_filter(
    signal: np.ndarray,
    time: np.ndarray,
    filter_config: str,
    f_low_hz: float,
    f_high_hz: float,
    transition_hz: float,
) -> np.ndarray:
    x = signal.astype(float)
    if x.size < 8:
        return x - np.mean(x)

    dt = float(np.median(np.diff(time.astype(float))))
    if not np.isfinite(dt) or dt <= 0:
        return x - np.mean(x)

    x = x - np.mean(x)
    n = x.size
    freqs = np.fft.rfftfreq(n, d=dt)
    if freqs.size <= 1:
        return x

    nyquist = 0.5 / dt
    transition = max(1e-9, float(transition_hz))
    cfg = str(filter_config or "bandpass").strip().lower()

    low = float(np.clip(max(0.0, f_low_hz), 0.0, max(0.0, nyquist * 0.999)))
    high = float(np.clip(max(0.0, f_high_hz), 0.0, max(0.0, nyquist * 0.999)))

    if cfg in {"low", "lowpass"}:
        transfer = _build_lowpass_transfer(freqs, low, transition)
    elif cfg in {"high", "highpass"}:
        transfer = _build_highpass_transfer(freqs, high, transition)
    elif cfg in {"stop", "bandstop"}:
        if high <= low:
            transfer = np.ones_like(freqs, dtype=float)
        else:
            band = _build_highpass_transfer(freqs, low, transition) * _build_lowpass_transfer(freqs, high, transition)
            transfer = 1.0 - band
    else:  # bandpass
        if high <= low:
            transfer = np.ones_like(freqs, dtype=float)
        else:
            transfer = _build_highpass_transfer(freqs, low, transition) * _build_lowpass_transfer(freqs, high, transition)

    spectrum = np.fft.rfft(x)
    filtered = np.fft.irfft(spectrum * transfer, n=n)
    return filtered.astype(float)


def _time_domain_filter(signal: np.ndarray, time: np.ndarray, cfg: Mapping[str, Any]) -> np.ndarray:
    if not HAS_SCIPY:
        return _fft_filter(
            signal,
            time,
            str(cfg.get("filter_config", "bandpass")),
            float(cfg.get("f_low_hz", DEFAULT_FILTER_LOW_HZ)),
            float(cfg.get("f_high_hz", DEFAULT_FILTER_HIGH_HZ)),
            float(cfg.get("transition_hz", DEFAULT_HIGHPASS_TRANSITION_HZ)),
        )

    dt = float(np.median(np.diff(time.astype(float))))
    if not np.isfinite(dt) or dt <= 0:
        return signal

    fn = 1.0 / (2.0 * dt)
    if not np.isfinite(fn) or fn <= 0:
        return signal

    cfg_name = str(cfg.get("filter_config", "bandpass")).strip().lower()
    filter_type = str(cfg.get("filter_type", "butter")).strip().lower()
    order = max(1, int(cfg.get("filter_order", DEFAULT_FILTER_ORDER)))
    f_low = max(0.0, float(cfg.get("f_low_hz", DEFAULT_FILTER_LOW_HZ)))
    f_high = max(0.0, float(cfg.get("f_high_hz", DEFAULT_FILTER_HIGH_HZ)))
    acausal = bool(cfg.get("acausal", True))

    b = None
    a = None

    if cfg_name in {"low", "lowpass"}:
        wn = min(max(f_low / fn, 1e-6), 0.999)
        if filter_type == "cheby":
            b, a = _scipy_cheby1(order, 0.5, wn, btype="low")
        elif filter_type == "bessel":
            b, a = _scipy_bessel(order, wn, btype="low", norm="phase")
        else:
            b, a = _scipy_butter(order, wn, btype="low")
    elif cfg_name in {"high", "highpass"}:
        wn = min(max(f_high / fn, 1e-6), 0.999)
        if filter_type == "cheby":
            b, a = _scipy_cheby1(order, 0.5, wn, btype="high")
        elif filter_type == "bessel":
            b, a = _scipy_bessel(order, wn, btype="high", norm="phase")
        else:
            b, a = _scipy_butter(order, wn, btype="high")
    elif cfg_name in {"stop", "bandstop"}:
        low = max(f_low / fn, 1e-6)
        high = min(f_high / fn, 0.999)
        if high > low:
            if filter_type == "cheby":
                b, a = _scipy_cheby1(order, 0.5, [low, high], btype="bandstop")
            elif filter_type == "bessel":
                b, a = _scipy_bessel(order, [low, high], btype="bandstop", norm="phase")
            else:
                b, a = _scipy_butter(order, [low, high], btype="bandstop")
    else:  # bandpass
        low = max(f_low / fn, 1e-6)
        high = min(f_high / fn, 0.999)
        if high > low:
            if filter_type == "cheby":
                b, a = _scipy_cheby1(order, 0.5, [low, high], btype="bandpass")
            elif filter_type == "bessel":
                b, a = _scipy_bessel(order, [low, high], btype="bandpass", norm="phase")
            else:
                b, a = _scipy_butter(order, [low, high], btype="band")

    if b is None or a is None:
        return signal

    if not acausal:
        return _scipy_lfilter(b, a, signal)

    try:
        padlen = 3 * (max(len(a), len(b)) - 1)
        if len(signal) <= padlen:
            return _scipy_lfilter(b, a, signal)
        return _scipy_filtfilt(b, a, signal)
    except ValueError:
        return _scipy_lfilter(b, a, signal)


def _normalize_processing_order(value: Any) -> str:
    v = str(value or "").strip().lower()
    if not v:
        return "filter_then_baseline"
    if "baseline" in v and "filter" in v:
        return "baseline_then_filter" if v.find("baseline") < v.find("filter") else "filter_then_baseline"
    if v in {"baseline_then_filter", "baseline-first", "baselinefirst"}:
        return "baseline_then_filter"
    return "filter_then_baseline"


def _normalize_filter_domain(value: Any) -> str:
    v = str(value or "").strip().lower()
    if "time" in v:
        return "time"
    if "freq" in v:
        return "frequency"
    return "frequency"


def _processing_config(options: Mapping[str, Any] | None) -> Dict[str, Any]:
    cfg = options or {}
    has_explicit_processing = any(
        key in cfg
        for key in (
            "processingOrder",
            "baselineMethod",
            "baselineOn",
            "baselineDegree",
            "filterOn",
            "filterDomain",
            "filterConfig",
            "filterType",
            "fLowHz",
            "fHighHz",
            "filterOrder",
            "filterAcausal",
        )
    )
    has_legacy_request = any(
        key in cfg
        for key in (
            "highpassEnabled",
            "highpassCutoffHz",
            "highpassTransitionHz",
        )
    )

    highpass_enabled, highpass_cutoff_hz, highpass_transition_hz = _highpass_config(cfg)

    if not has_explicit_processing:
        if has_legacy_request:
            return {
                "legacy": True,
                "highpass_enabled": bool(highpass_enabled),
                "highpass_cutoff_hz": float(highpass_cutoff_hz),
                "highpass_transition_hz": float(highpass_transition_hz),
            }
        return {
            "legacy": False,
            "processing_order": "filter_then_baseline",
            "baseline_on": False,
            "baseline_method": "poly4",
            "baseline_degree": DEFAULT_BASELINE_DEGREE,
            "filter_on": False,
            "filter_domain": "time",
            "filter_config": "bandpass",
            "filter_type": "butter",
            "f_low_hz": DEFAULT_FILTER_LOW_HZ,
            "f_high_hz": DEFAULT_FILTER_HIGH_HZ,
            "filter_order": DEFAULT_FILTER_ORDER,
            "acausal": True,
            "transition_hz": DEFAULT_HIGHPASS_TRANSITION_HZ,
            "scipy_enabled": bool(HAS_SCIPY),
        }

    f_low_default = _to_float(cfg.get("fLowHz"), DEFAULT_FILTER_LOW_HZ)
    if "fHighHz" in cfg:
        f_high_default = _to_float(cfg.get("fHighHz"), DEFAULT_FILTER_HIGH_HZ)
    else:
        f_high_default = _to_float(cfg.get("highpassCutoffHz"), DEFAULT_FILTER_HIGH_HZ)

    return {
        "legacy": False,
        "processing_order": _normalize_processing_order(cfg.get("processingOrder", "filter_then_baseline")),
        "baseline_on": _to_bool(cfg.get("baselineOn", True), True),
        "baseline_method": str(cfg.get("baselineMethod", "poly4")),
        "baseline_degree": max(1, int(round(_to_float(cfg.get("baselineDegree"), DEFAULT_BASELINE_DEGREE)))),
        "filter_on": _to_bool(cfg.get("filterOn", True), True),
        "filter_domain": _normalize_filter_domain(cfg.get("filterDomain", "frequency")),
        "filter_config": str(cfg.get("filterConfig", "bandpass")),
        "filter_type": str(cfg.get("filterType", "butter")),
        "f_low_hz": max(0.0, f_low_default),
        "f_high_hz": max(0.0, f_high_default),
        "filter_order": max(1, int(round(_to_float(cfg.get("filterOrder"), DEFAULT_FILTER_ORDER)))),
        "acausal": _to_bool(cfg.get("filterAcausal", True), True),
        "transition_hz": max(1e-9, _to_float(cfg.get("highpassTransitionHz"), DEFAULT_HIGHPASS_TRANSITION_HZ)),
        "scipy_enabled": bool(HAS_SCIPY),
    }


def _processing_summary_text(options: Mapping[str, Any] | None) -> str:
    cfg = _processing_config(options)
    base_ref = _normalize_base_reference((options or {}).get("baseReference", DEFAULT_BASE_REFERENCE))
    if cfg.get("legacy", True):
        return (
            "legacy-highpass"
            f" | enabled={'yes' if cfg['highpass_enabled'] else 'no'}"
            f" | cutoff={cfg['highpass_cutoff_hz']:.4f} Hz"
            f" | transition={cfg['highpass_transition_hz']:.4f} Hz"
            f" | base-ref={base_ref}"
        )

    return (
        f"order={cfg['processing_order']} | baseline={'on' if cfg['baseline_on'] else 'off'}"
        f" ({cfg['baseline_method']}) | filter={'on' if cfg['filter_on'] else 'off'}"
        f" [{cfg['filter_domain']}/{cfg['filter_config']}/{cfg['filter_type']}]"
        f" low={cfg['f_low_hz']:.4f}Hz high={cfg['f_high_hz']:.4f}Hz n={cfg['filter_order']}"
        f" | scipy={'yes' if cfg['scipy_enabled'] else 'no'}"
        f" | base-ref={base_ref}"
    )


def _normalize_alt_integration_method(value: Any) -> str:
    method = str(value or DEFAULT_ALT_INTEGRATION_METHOD).strip().lower()
    if method in {"fft_regularized", "fft-regularized", "fft"}:
        return "fft_regularized"
    return DEFAULT_ALT_INTEGRATION_METHOD


def _integration_compare_config(options: Mapping[str, Any] | None) -> Dict[str, Any]:
    cfg = options or {}
    enabled = _to_bool(cfg.get("integrationCompareEnabled", False), False)
    method = _normalize_alt_integration_method(cfg.get("altIntegrationMethod", DEFAULT_ALT_INTEGRATION_METHOD))
    return {
        "enabled": bool(enabled),
        "method": method,
    }


def _preprocess_acc_for_integration(
    t: np.ndarray,
    acc: np.ndarray,
    cfg: Mapping[str, Any],
    *,
    highpass_enabled: bool = True,
    highpass_cutoff_hz: float = DEFAULT_HIGHPASS_CUTOFF_HZ,
    highpass_transition_hz: float = DEFAULT_HIGHPASS_TRANSITION_HZ,
) -> np.ndarray:
    if cfg.get("legacy", True):
        hp_enabled = bool(cfg.get("highpass_enabled", highpass_enabled))
        hp_cutoff = float(cfg.get("highpass_cutoff_hz", highpass_cutoff_hz))
        hp_transition = float(cfg.get("highpass_transition_hz", highpass_transition_hz))

        acc_corr = _baseline_correct_legacy(acc, t)
        if hp_enabled:
            return _soft_highpass_fft(
                acc_corr,
                t,
                cutoff_hz=hp_cutoff,
                transition_hz=hp_transition,
            )
        return acc_corr

    acc_proc = acc.copy()
    if cfg["processing_order"] == "baseline_then_filter":
        if cfg["baseline_on"]:
            acc_proc = _apply_baseline(acc_proc, cfg["baseline_method"], cfg["baseline_degree"])
        if cfg["filter_on"]:
            if cfg["filter_domain"] == "time":
                acc_proc = _time_domain_filter(acc_proc, t, cfg)
            else:
                acc_proc = _fft_filter(
                    acc_proc,
                    t,
                    cfg["filter_config"],
                    cfg["f_low_hz"],
                    cfg["f_high_hz"],
                    cfg["transition_hz"],
                )
    else:
        if cfg["filter_on"]:
            if cfg["filter_domain"] == "time":
                acc_proc = _time_domain_filter(acc_proc, t, cfg)
            else:
                acc_proc = _fft_filter(
                    acc_proc,
                    t,
                    cfg["filter_config"],
                    cfg["f_low_hz"],
                    cfg["f_high_hz"],
                    cfg["transition_hz"],
                )
        if cfg["baseline_on"]:
            acc_proc = _apply_baseline(acc_proc, cfg["baseline_method"], cfg["baseline_degree"])
    return acc_proc


def _integrate_primary_disp(t: np.ndarray, acc_proc: np.ndarray, cfg: Mapping[str, Any]) -> np.ndarray:
    vel = _integrate_primary_velocity(t, acc_proc, cfg)
    disp = _cumtrapz(vel, t)
    if not cfg.get("legacy", True) and cfg.get("baseline_on", False):
        disp = _detrend_poly(disp, degree=1)
    return disp


def _integrate_primary_velocity(t: np.ndarray, acc_proc: np.ndarray, cfg: Mapping[str, Any]) -> np.ndarray:
    vel = _cumtrapz(acc_proc * 9.81, t)
    if not cfg.get("legacy", True) and cfg.get("baseline_on", False):
        vel = vel - np.mean(vel)
    return vel


def _resolve_alt_lowcut_hz(processing_cfg: Mapping[str, Any], options: Mapping[str, Any] | None) -> float:
    _ = options
    if processing_cfg.get("legacy", True):
        if bool(processing_cfg.get("highpass_enabled", True)):
            value = float(processing_cfg.get("highpass_cutoff_hz", DEFAULT_ALT_LOWCUT_HZ))
        else:
            value = DEFAULT_ALT_LOWCUT_HZ
    else:
        if bool(processing_cfg.get("filter_on", False)):
            value = float(processing_cfg.get("f_low_hz", DEFAULT_ALT_LOWCUT_HZ))
        else:
            value = DEFAULT_ALT_LOWCUT_HZ
    return max(1e-4, value)


def _fft_regularized_disp(
    time: np.ndarray,
    acc_proc_g: np.ndarray,
    options: Mapping[str, Any] | None,
    processing_cfg: Mapping[str, Any],
) -> Tuple[np.ndarray, float]:
    t = time.astype(float)
    acc = acc_proc_g.astype(float)

    if t.size == 0:
        return np.array([], dtype=float), _resolve_alt_lowcut_hz(processing_cfg, options)
    if t.size == 1:
        return np.array([0.0], dtype=float), _resolve_alt_lowcut_hz(processing_cfg, options)

    dt = float(np.median(np.diff(t)))
    if not np.isfinite(dt) or dt <= 0:
        return _integrate_primary_disp(t, acc, processing_cfg), _resolve_alt_lowcut_hz(processing_cfg, options)

    n = acc.size
    if n < 8:
        disp = _integrate_primary_disp(t, acc, processing_cfg)
        disp = disp - np.mean(disp)
        disp = _detrend_poly(disp, degree=1)
        return disp, _resolve_alt_lowcut_hz(processing_cfg, options)

    nyquist = 0.5 / dt
    lowcut_raw = _resolve_alt_lowcut_hz(processing_cfg, options)
    lowcut = float(np.clip(lowcut_raw, 1e-4, max(1e-4, nyquist * 0.999)))

    if processing_cfg.get("legacy", True):
        transition = float(processing_cfg.get("highpass_transition_hz", DEFAULT_HIGHPASS_TRANSITION_HZ))
    else:
        transition = float(processing_cfg.get("transition_hz", DEFAULT_HIGHPASS_TRANSITION_HZ))
    transition = max(1e-9, transition)

    acc_ms2 = acc * 9.81
    spectrum = np.fft.rfft(acc_ms2)
    freqs = np.fft.rfftfreq(n, d=dt)
    omega = 2.0 * np.pi * freqs

    hp = _build_highpass_transfer(freqs, lowcut, transition)
    disp_spec = np.zeros_like(spectrum, dtype=np.complex128)
    nz = omega > 0.0
    disp_spec[nz] = -spectrum[nz] * hp[nz] / (omega[nz] ** 2)
    disp = np.fft.irfft(disp_spec, n=n).astype(float)

    # Keep small post-correction to suppress residual DC/linear drift.
    disp = disp - np.mean(disp)
    disp = _detrend_poly(disp, degree=1)
    return disp, lowcut


def _acc_to_disp_dual(
    time: np.ndarray,
    acc_g: np.ndarray,
    *,
    options: Mapping[str, Any] | None = None,
    highpass_enabled: bool = True,
    highpass_cutoff_hz: float = DEFAULT_HIGHPASS_CUTOFF_HZ,
    highpass_transition_hz: float = DEFAULT_HIGHPASS_TRANSITION_HZ,
) -> Dict[str, Any]:
    t = time.astype(float)
    acc = acc_g.astype(float)
    processing_cfg = _processing_config(options)
    compare_cfg = _integration_compare_config(options)

    acc_proc = _preprocess_acc_for_integration(
        t,
        acc,
        processing_cfg,
        highpass_enabled=highpass_enabled,
        highpass_cutoff_hz=highpass_cutoff_hz,
        highpass_transition_hz=highpass_transition_hz,
    )
    primary = _integrate_primary_disp(t, acc_proc, processing_cfg)
    primary_velocity = _integrate_primary_velocity(t, acc_proc, processing_cfg)

    meta: Dict[str, Any] = {
        "integrationPrimary": "cumtrapz",
        "integrationCompareEnabled": bool(compare_cfg["enabled"]),
    }
    alt: np.ndarray | None = None
    if compare_cfg["enabled"] and compare_cfg["method"] == "fft_regularized":
        alt, lowcut = _fft_regularized_disp(t, acc_proc, options, processing_cfg)
        meta.update(
            {
                "altIntegrationMethod": "fft_regularized",
                "altLowCutHz": float(lowcut),
                "altLowCutPolicy": ALT_LOWCUT_POLICY,
            }
        )

    return {
        "primary": primary,
        "primary_velocity": primary_velocity,
        "alt": alt,
        "acc_processed_g": acc_proc,
        "processing_cfg": processing_cfg,
        "meta": meta,
    }


def _acc_to_disp(
    time: np.ndarray,
    acc_g: np.ndarray,
    *,
    options: Mapping[str, Any] | None = None,
    highpass_enabled: bool = True,
    highpass_cutoff_hz: float = DEFAULT_HIGHPASS_CUTOFF_HZ,
    highpass_transition_hz: float = DEFAULT_HIGHPASS_TRANSITION_HZ,
) -> np.ndarray:
    return _acc_to_disp_dual(
        time,
        acc_g,
        options=options,
        highpass_enabled=highpass_enabled,
        highpass_cutoff_hz=highpass_cutoff_hz,
        highpass_transition_hz=highpass_transition_hz,
    )["primary"]


def _normalize_options(options: Any) -> Dict[str, Any]:
    if options is None:
        return {}
    if hasattr(options, "to_py"):
        options = options.to_py()
    if isinstance(options, dict):
        return dict(options)
    return {}


def _report_batch_progress(
    options: Mapping[str, Any] | None,
    completed_steps: int,
    total_steps: int,
    message: str,
) -> None:
    if not isinstance(options, Mapping):
        return

    callback = options.get("_progress_callback")
    if callback is None:
        return

    safe_total = max(1, int(total_steps))
    safe_completed = min(max(int(completed_steps), 0), safe_total)
    ratio = safe_completed / safe_total
    progress = CALC_PROGRESS_START + ratio * (CALC_PROGRESS_END - CALC_PROGRESS_START)
    try:
        callback(str(message), "run", float(progress), False)
    except Exception:  # noqa: BLE001
        return


def _build_web_result_payload(result: Mapping[str, Any]) -> Dict[str, Any]:
    output_bytes = result.get("outputBytes", b"")
    if output_bytes is None:
        output_bytes = b""
    viewer_group, viewer_kind, viewer_order = _viewer_result_group(result)
    return {
        "pairKey": result.get("pairKey", ""),
        "xFileName": result.get("xFileName", ""),
        "yFileName": result.get("yFileName", ""),
        "outputFileName": result.get("outputFileName", ""),
        "outputBytesB64": base64.b64encode(bytes(output_bytes)).decode("ascii"),
        "previewCharts": result.get("previewCharts", []),
        "metrics": result.get("metrics", {}),
        "viewerGroup": viewer_group,
        "viewerKind": viewer_kind,
        "viewerGroupOrder": viewer_order,
    }


def _viewer_result_group(result: Mapping[str, Any]) -> tuple[str, str, int]:
    metrics = result.get("metrics", {}) if isinstance(result, Mapping) else {}
    mode = str(metrics.get("mode", "") or "").strip().lower()
    axis = str(metrics.get("axis", "") or "").strip().upper()

    if mode.startswith("db_"):
        if mode == "db_pair":
            return "DB Direct", "DB Pair", 40
        if mode == "db_single":
            return "DB Direct", "DB Single", 40
        if mode == "db_method2_single":
            return "DB Direct", "DB Method-2", 41
        if mode in {"db_method3", "db_method3_aggregate"}:
            return "DB Direct", "DB Method-3", 42
        return "DB Direct", "DB Output", 40

    if mode == "method2_single":
        return "Method-2", f"Method-2 {axis}" if axis in {"X", "Y"} else "Method-2", 20

    if mode in {"method3", "method3_aggregate"}:
        return "Method-3 Aggregate", "Method-3", 30

    if mode == "pair":
        return "Primary Outputs", "Pair", 10

    if mode == "single":
        return "Primary Outputs", "Single", 11

    return "Other Outputs", mode or "Output", 90


def _source_slug(value: Any) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", str(value or "").strip().lower())
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text[:80] or "item"


def _resolve_column_aliases(columns: Sequence[Any], aliases: Sequence[str]) -> str | None:
    normalized: Dict[str, str] = {}
    for column in columns:
        if column is None:
            continue
        normalized[str(column).strip().lower()] = str(column)

    cleaned_aliases = [str(alias or "").strip().lower() for alias in aliases if str(alias or "").strip()]
    for alias in cleaned_aliases:
        if alias in normalized:
            return normalized[alias]

    for alias in cleaned_aliases:
        for key, original in normalized.items():
            if alias in key:
                return original
    return None


def _read_numeric_pair_from_df(
    frame: pd.DataFrame,
    x_aliases: Sequence[str],
    y_aliases: Sequence[str],
) -> Tuple[np.ndarray, np.ndarray]:
    def _extract_subset(candidate: pd.DataFrame) -> pd.DataFrame:
        x_column = _resolve_column_aliases(candidate.columns, x_aliases)
        y_column = _resolve_column_aliases(candidate.columns, y_aliases)
        if x_column is None or y_column is None:
            raise ValueError(f"Missing columns for aliases: {x_aliases} / {y_aliases}")

        subset = candidate[[x_column, y_column]].copy()
        subset.columns = ["x", "y"]
        subset["x"] = pd.to_numeric(subset["x"], errors="coerce")
        subset["y"] = pd.to_numeric(subset["y"], errors="coerce")
        subset = subset.dropna(subset=["x", "y"])
        if subset.empty:
            raise ValueError(f"No numeric rows for aliases: {x_aliases} / {y_aliases}")
        return subset

    try:
        subset = _extract_subset(frame)
    except Exception:
        if frame.empty:
            raise
        promoted = frame.copy()
        promoted.columns = [str(value).strip() for value in promoted.iloc[0].tolist()]
        promoted = promoted.iloc[1:].reset_index(drop=True)
        subset = _extract_subset(promoted)

    return subset["x"].to_numpy(dtype=float), subset["y"].to_numpy(dtype=float)


def _profile_numeric_pair(xl: pd.ExcelFile, aliases: Sequence[str]) -> Tuple[np.ndarray, np.ndarray]:
    if "Profile" not in xl.sheet_names:
        raise ValueError("Missing 'Profile' sheet.")

    profile = _parse_sheet_cached(xl, "Profile")
    data = profile.iloc[1:].reset_index(drop=True)
    if data.empty:
        raise ValueError("Profile sheet has no numeric content.")

    target_column = _resolve_column_aliases(data.columns, aliases)
    if target_column is None:
        raise ValueError(f"Profile group not found for aliases: {aliases}")

    depth_idx = data.columns.get_loc(target_column)
    value_idx = min(depth_idx + 1, len(data.columns) - 1)
    depths = _to_number_series(data.iloc[:, depth_idx])
    values = _to_number_series(data.iloc[:, value_idx])
    count = min(depths.size, values.size)
    if count <= 0:
        raise ValueError(f"Profile group contains no numeric rows for aliases: {aliases}")
    return depths[:count], values[:count]


def _read_input_motion_curves(xl: pd.ExcelFile) -> Dict[str, Tuple[np.ndarray, np.ndarray]]:
    if "Input Motion" not in xl.sheet_names:
        return {}

    motion = _parse_sheet_cached(xl, "Input Motion")
    curves: Dict[str, Tuple[np.ndarray, np.ndarray]] = {}
    specs = {
        "acceleration": (("Time (s)", "Time", "Time (sec)"), ("Acceleration (g)", "Acceleration", "Accel (g)")),
        "psa": (
            ("Period (sec)", "Period", "Period (s)"),
            ("PSA (g)", "PSA", "5% Damped Spectral", "5% Damped Spectrum", "Damped Spectral"),
        ),
        "fourier": (("Frequency (Hz)", "Frequency"), ("Fourier Amplitude", "Fourier")),
    }
    for key, (x_aliases, y_aliases) in specs.items():
        try:
            curves[key] = _read_numeric_pair_from_df(motion, x_aliases, y_aliases)
        except Exception:
            continue
    return curves


def _read_layer_curve(
    xl: pd.ExcelFile,
    layer_name: str,
    x_aliases: Sequence[str],
    y_aliases: Sequence[str],
) -> Tuple[np.ndarray, np.ndarray]:
    frame = _parse_sheet_cached(xl, layer_name)
    return _read_numeric_pair_from_df(frame, x_aliases, y_aliases)


def _source_series(
    name: str,
    x: np.ndarray | Sequence[float],
    y: np.ndarray | Sequence[float],
    *,
    series_key: str | None = None,
    max_points: int = SOURCE_VIEW_MAX_POINTS,
) -> Dict[str, Any] | None:
    points = _preview_points(x, y, max_points=max_points)
    if not points:
        return None
    return {
        "seriesKey": str(series_key or _source_slug(name)),
        "name": str(name),
        "points": points,
    }


def _source_chart(
    chart_key: str,
    chart_label: str,
    sheet_name: str,
    chart_type: str,
    x_label: str,
    y_label: str,
    *,
    series: Sequence[Dict[str, Any] | None] | None = None,
    layer_views: Sequence[Dict[str, Any]] | None = None,
    invert_y: bool = False,
) -> Dict[str, Any] | None:
    clean_series = [item for item in (series or []) if isinstance(item, dict) and item.get("points")]
    clean_views: List[Dict[str, Any]] = []
    for item in layer_views or []:
        if not isinstance(item, dict):
            continue
        item_series = [series_item for series_item in item.get("series", []) if isinstance(series_item, dict) and series_item.get("points")]
        if not item_series:
            continue
        clean_views.append(
            {
                "layerIndex": int(item.get("layerIndex", 0)),
                "layerLabel": str(item.get("layerLabel", "")),
                "depth": float(item.get("depth", 0.0)),
                "series": item_series,
            }
        )

    if not clean_series and not clean_views:
        return None

    out: Dict[str, Any] = {
        "chartKey": str(chart_key),
        "chartLabel": str(chart_label),
        "sheetName": str(sheet_name),
        "chartType": str(chart_type),
        "xLabel": str(x_label),
        "yLabel": str(y_label),
        "invertY": bool(invert_y),
    }
    if clean_series:
        out["series"] = clean_series
    if clean_views:
        out["layerViews"] = clean_views
    return out


def _source_family(
    family_key: str,
    family_label: str,
    chart_type: str,
    charts: Sequence[Dict[str, Any] | None],
    *,
    supports_overlay: bool = False,
    supports_layer_selection: bool = False,
    layers: Sequence[Dict[str, Any]] | None = None,
) -> Dict[str, Any] | None:
    clean_charts = [item for item in charts if isinstance(item, dict)]
    if not clean_charts:
        return None

    default_visible: List[str] = []
    first_chart = clean_charts[0]
    if "series" in first_chart:
        default_visible = [str(item.get("seriesKey", "")) for item in first_chart.get("series", []) if item.get("seriesKey")]
    elif "layerViews" in first_chart and first_chart["layerViews"]:
        default_visible = [
            str(item.get("seriesKey", ""))
            for item in first_chart["layerViews"][0].get("series", [])
            if item.get("seriesKey")
        ]

    return {
        "familyKey": str(family_key),
        "familyLabel": str(family_label),
        "chartType": str(chart_type),
        "supportsOverlay": bool(supports_overlay),
        "supportsLayerSelection": bool(supports_layer_selection),
        "defaultVisibleSeries": default_visible,
        "layers": list(layers or []),
        "charts": clean_charts,
    }


def _source_layers(layer_names: Sequence[str], depths: Sequence[float]) -> List[Dict[str, Any]]:
    depth_arr = np.asarray(depths, dtype=float)
    layers: List[Dict[str, Any]] = []
    for idx, layer_name in enumerate(layer_names):
        depth = float(depth_arr[idx]) if idx < depth_arr.size and np.isfinite(depth_arr[idx]) else float(idx + 1)
        layers.append(
            {
                "layerIndex": int(idx),
                "layerLabel": str(layer_name),
                "depth": round(depth, 6),
            }
        )
    return layers


def _source_entry(
    source_id: str,
    source_label: str,
    source_kind: str,
    axis: str,
    pair_key: str,
    families: Sequence[Dict[str, Any] | None],
    *,
    artifact_pair_keys: Sequence[str] | None = None,
) -> Dict[str, Any]:
    clean_families = [item for item in families if isinstance(item, dict)]
    return {
        "sourceId": str(source_id),
        "sourceLabel": str(source_label),
        "sourceSystem": SUMMARY_SOURCE_SYSTEM,
        "sourceKind": str(source_kind),
        "axis": str(axis or ""),
        "pairKey": str(pair_key or ""),
        "artifactPairKeys": list(artifact_pair_keys or []),
        "families": clean_families,
    }


def _build_summary_entry(
    summary_id: str,
    summary_label: str,
    summary_kind: str,
    axis: str,
    pair_key: str,
    variants: Sequence[Mapping[str, Any]],
    *,
    input_kind: str,
    available_layer_count: int,
    profile_layer_count: int,
    warnings: Sequence[str] | None = None,
    detail_source_ids: Sequence[str] | None = None,
    artifact_pair_keys: Sequence[str] | None = None,
) -> Dict[str, Any]:
    clean_variants = [dict(item) for item in variants if isinstance(item, Mapping)]
    primary_variant_key = _pick_primary_variant_key(clean_variants)
    valid_variant_count = sum(
        1 for item in clean_variants if bool(item.get("valid")) and _series_has_finite_values(item.get("values", []))
    )
    profile_count = max(0, int(profile_layer_count))
    available_count = max(0, int(available_layer_count))
    limited_data = profile_count > 0 and available_count < profile_count
    warning_list = [str(item) for item in (warnings or []) if str(item or "").strip()]
    return {
        "summaryId": str(summary_id),
        "summaryLabel": str(summary_label),
        "sourceSystem": SUMMARY_SOURCE_SYSTEM,
        "summaryKind": str(summary_kind),
        "axis": str(axis or ""),
        "pairKey": str(pair_key or ""),
        "inputKind": str(input_kind or ""),
        "preferredVariantKey": primary_variant_key,
        "validVariantCount": int(valid_variant_count),
        "detailSourceIds": list(detail_source_ids or []),
        "artifactPairKeys": list(artifact_pair_keys or []),
        "warnings": warning_list,
        "coverage": {
            "availableLayerCount": available_count,
            "profileLayerCount": profile_count,
            "limitedData": bool(limited_data),
            "label": "Sinirli veri" if limited_data else "Tam veri",
        },
        "variants": clean_variants,
    }


def _profile_chart_from_single(
    xl: pd.ExcelFile,
    chart_key: str,
    chart_label: str,
    aliases: Sequence[str],
    value_label: str,
) -> Dict[str, Any] | None:
    try:
        depths, values = _profile_numeric_pair(xl, aliases)
    except Exception:
        return None
    return _source_chart(
        chart_key,
        chart_label,
        "Profile",
        "depth",
        value_label,
        "Depth (m)",
        series=[_source_series(value_label, values, depths, series_key=_source_slug(value_label))],
        invert_y=True,
    )


def _pair_depth_series_chart(
    chart_key: str,
    chart_label: str,
    sheet_name: str,
    depths: np.ndarray,
    series_specs: Sequence[Tuple[str, np.ndarray | Sequence[float]]],
    *,
    x_label: str,
) -> Dict[str, Any] | None:
    series: List[Dict[str, Any] | None] = []
    depth_arr = np.asarray(depths, dtype=float)
    for label, values in series_specs:
        value_arr = np.asarray(values, dtype=float)
        count = min(depth_arr.size, value_arr.size)
        if count <= 0:
            continue
        series.append(_source_series(label, value_arr[:count], depth_arr[:count], series_key=_source_slug(label)))

    return _source_chart(
        chart_key,
        chart_label,
        sheet_name,
        "depth",
        x_label,
        "Depth (m)",
        series=series,
        invert_y=True,
    )


def _single_layer_views(
    layers: Sequence[Dict[str, Any]],
    time: np.ndarray,
    matrix: np.ndarray,
    series_name: str,
    series_key: str,
) -> List[Dict[str, Any]]:
    layer_views: List[Dict[str, Any]] = []
    matrix_arr = np.asarray(matrix, dtype=float)
    for layer in layers:
        idx = int(layer["layerIndex"])
        if idx >= matrix_arr.shape[0]:
            continue
        series = _source_series(series_name, time, matrix_arr[idx], series_key=series_key)
        if series is None:
            continue
        layer_views.append(
            {
                "layerIndex": idx,
                "layerLabel": layer["layerLabel"],
                "depth": layer["depth"],
                "series": [series],
            }
        )
    return layer_views


def _paired_layer_views(
    layers: Sequence[Dict[str, Any]],
    x_time: np.ndarray,
    x_matrix: np.ndarray,
    x_name: str,
    x_series_key: str,
    y_time: np.ndarray,
    y_matrix: np.ndarray,
    y_name: str,
    y_series_key: str,
    *,
    resultant_name: str | None = None,
    resultant_series_key: str | None = None,
) -> List[Dict[str, Any]]:
    layer_views: List[Dict[str, Any]] = []
    x_arr = np.asarray(x_matrix, dtype=float)
    y_arr = np.asarray(y_matrix, dtype=float)
    for layer in layers:
        idx = int(layer["layerIndex"])
        if idx >= x_arr.shape[0] or idx >= y_arr.shape[0]:
            continue

        series: List[Dict[str, Any] | None] = [
            _source_series(x_name, x_time, x_arr[idx], series_key=x_series_key),
            _source_series(y_name, y_time, y_arr[idx], series_key=y_series_key),
        ]
        if resultant_name and resultant_series_key:
            common_time, x_interp, y_interp = _align_two_series(x_time, x_arr[idx], y_time, y_arr[idx])
            series.append(
                _source_series(
                    resultant_name,
                    common_time,
                    np.sqrt(x_interp**2 + y_interp**2),
                    series_key=resultant_series_key,
                )
            )

        clean_series = [item for item in series if isinstance(item, dict)]
        if not clean_series:
            continue
        layer_views.append(
            {
                "layerIndex": idx,
                "layerLabel": layer["layerLabel"],
                "depth": layer["depth"],
                "series": clean_series,
            }
        )
    return layer_views


def _single_input_motion_family(xl: pd.ExcelFile, axis_label: str) -> Dict[str, Any] | None:
    curves = _read_input_motion_curves(xl)
    axis = axis_label.upper() if axis_label in {"X", "Y"} else ""
    series_label = f"Input {axis}".strip() or "Input"
    charts = [
        _source_chart(
            "input-acceleration",
            "Input Acceleration",
            "Input Motion",
            "time",
            "Time (s)",
            "Acceleration (g)",
            series=[_source_series(series_label, *curves["acceleration"], series_key=f"input-{_source_slug(axis or 'single')}-acc")],
        )
        if "acceleration" in curves
        else None,
        _source_chart(
            "input-psa",
            "Input PSA Spectrum",
            "Input Motion",
            "spectrum",
            "Period (s)",
            "PSA (g)",
            series=[_source_series(series_label, *curves["psa"], series_key=f"input-{_source_slug(axis or 'single')}-psa")],
        )
        if "psa" in curves
        else None,
        _source_chart(
            "input-fourier",
            "Input Fourier Amplitude",
            "Input Motion",
            "fourier",
            "Frequency (Hz)",
            "Fourier Amplitude",
            series=[_source_series(series_label, *curves["fourier"], series_key=f"input-{_source_slug(axis or 'single')}-fourier")],
        )
        if "fourier" in curves
        else None,
    ]
    return _source_family("input-motion", "Input Motion", "time", charts)


def _pair_input_motion_family(x_xl: pd.ExcelFile, y_xl: pd.ExcelFile) -> Dict[str, Any] | None:
    x_curves = _read_input_motion_curves(x_xl)
    y_curves = _read_input_motion_curves(y_xl)
    charts = [
        _source_chart(
            "input-acceleration",
            "Input Acceleration",
            "Input Motion",
            "time",
            "Time (s)",
            "Acceleration (g)",
            series=[
                _source_series("Input X", *x_curves["acceleration"], series_key="input-x-acc"),
                _source_series("Input Y", *y_curves["acceleration"], series_key="input-y-acc"),
            ],
        )
        if "acceleration" in x_curves and "acceleration" in y_curves
        else None,
        _source_chart(
            "input-psa",
            "Input PSA Spectrum",
            "Input Motion",
            "spectrum",
            "Period (s)",
            "PSA (g)",
            series=[
                _source_series("Input X", *x_curves["psa"], series_key="input-x-psa"),
                _source_series("Input Y", *y_curves["psa"], series_key="input-y-psa"),
            ],
        )
        if "psa" in x_curves and "psa" in y_curves
        else None,
        _source_chart(
            "input-fourier",
            "Input Fourier Amplitude",
            "Input Motion",
            "fourier",
            "Frequency (Hz)",
            "Fourier Amplitude",
            series=[
                _source_series("Input X", *x_curves["fourier"], series_key="input-x-fourier"),
                _source_series("Input Y", *y_curves["fourier"], series_key="input-y-fourier"),
            ],
        )
        if "fourier" in x_curves and "fourier" in y_curves
        else None,
    ]
    return _source_family("input-motion", "Input Motion", "time", charts, supports_overlay=True)


def _single_profile_family(xl: pd.ExcelFile) -> Dict[str, Any] | None:
    charts = [
        _profile_chart_from_single(
            xl,
            "profile-max-displacement",
            "Profile Max Displacement",
            ("Maximum Displacement", "Max Displacement"),
            "Displacement (m)",
        ),
        _profile_chart_from_single(
            xl,
            "profile-max-strain",
            "Profile Max Strain",
            ("Max. Strain", "Maximum Strain", "Max Strain"),
            "Strain (%)",
        ),
        _profile_chart_from_single(
            xl,
            "profile-max-stress-ratio",
            "Profile Max Stress Ratio",
            ("Max. Stress Ratio", "Max Stress Ratio", "Stress Ratio"),
            "Stress Ratio",
        ),
        _profile_chart_from_single(
            xl,
            "profile-pga",
            "Profile PGA",
            ("PGA (g)", "PGA", "Peak Ground Acceleration", "Peak Acceleration", "Max. Acceleration"),
            "PGA (g)",
        ),
        _profile_chart_from_single(
            xl,
            "profile-effective-stress",
            "Profile Effective Stress",
            (
                "Effective Stress",
                "Effective Vertical Stress",
                "Effective Vert. Stress",
                "Eff. Vert. Stress",
            ),
            "Effective Stress (kPa)",
        ),
    ]
    return _source_family("profile", "Profile", "depth", charts)


def _pair_profile_family(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    profile_sheet_df: pd.DataFrame,
) -> Dict[str, Any] | None:
    depths = pd.to_numeric(profile_sheet_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    charts: List[Dict[str, Any] | None] = [
        _pair_depth_series_chart(
            "profile-max-displacement",
            "Profile Max Displacement",
            "Profile",
            depths,
            [
                ("Profile X", profile_sheet_df.get("Profile_X_raw_max_m", [])),
                ("Profile Y", profile_sheet_df.get("Profile_Y_raw_max_m", [])),
                ("Profile RSS", profile_sheet_df.get("Profile_RSS_raw_max_m", [])),
            ],
            x_label="Displacement (m)",
        )
    ]

    for chart_key, chart_label, aliases, x_label in (
        ("profile-max-strain", "Profile Max Strain", ("Max. Strain", "Maximum Strain", "Max Strain"), "Strain (%)"),
        (
            "profile-max-stress-ratio",
            "Profile Max Stress Ratio",
            ("Max. Stress Ratio", "Max Stress Ratio", "Stress Ratio"),
            "Stress Ratio",
        ),
        (
            "profile-pga",
            "Profile PGA",
            ("PGA (g)", "PGA", "Peak Ground Acceleration", "Peak Acceleration", "Max. Acceleration"),
            "PGA (g)",
        ),
        (
            "profile-effective-stress",
            "Profile Effective Stress",
            (
                "Effective Stress",
                "Effective Vertical Stress",
                "Effective Vert. Stress",
                "Eff. Vert. Stress",
            ),
            "Effective Stress (kPa)",
        ),
    ):
        try:
            x_depths, x_values = _profile_numeric_pair(x_xl, aliases)
            y_depths, y_values = _profile_numeric_pair(y_xl, aliases)
            count = min(x_depths.size, y_depths.size, x_values.size, y_values.size)
            if count > 0:
                charts.append(
                    _pair_depth_series_chart(
                        chart_key,
                        chart_label,
                        "Profile",
                        x_depths[:count],
                        [("Profile X", x_values[:count]), ("Profile Y", y_values[:count])],
                        x_label=x_label,
                    )
                )
        except Exception:
            continue

    return _source_family("profile", "Profile", "depth", charts, supports_overlay=True)


def _single_layer_family(
    xl: pd.ExcelFile,
    axis_label: str,
    direction_bundle: Mapping[str, Any],
    strain_bundle: Mapping[str, Any],
) -> Dict[str, Any] | None:
    def _single_layer_curve_views(
        x_aliases: Sequence[str],
        y_aliases: Sequence[str],
        series_name: str,
        series_key: str,
    ) -> List[Dict[str, Any]]:
        views: List[Dict[str, Any]] = []
        for layer in layers:
            layer_name = str(layer["layerLabel"])
            try:
                x_values, y_values = _read_layer_curve(xl, layer_name, x_aliases, y_aliases)
            except Exception:
                continue
            series = _source_series(series_name, x_values, y_values, series_key=series_key)
            if series is None:
                continue
            views.append(
                {
                    "layerIndex": int(layer["layerIndex"]),
                    "layerLabel": layer_name,
                    "depth": float(layer["depth"]),
                    "series": [series],
                }
            )
        return views

    layers = _source_layers(direction_bundle.get("layer_names", []), direction_bundle.get("depths", []))
    charts = [
        _source_chart(
            "layer-acceleration",
            "Layer Acceleration",
            "Layer",
            "time",
            "Time (s)",
            "Acceleration (g)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(direction_bundle.get("time", []), dtype=float),
                np.asarray(direction_bundle.get("acc_matrix", np.zeros((0, 0))), dtype=float),
                f"Layer {axis_label or 'Input'} Acceleration",
                f"layer-{_source_slug(axis_label or 'single')}-acc",
            ),
        ),
        _source_chart(
            "layer-strain",
            "Layer Strain",
            "Layer",
            "time",
            "Time (s)",
            "Strain (%)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(strain_bundle.get("time", []), dtype=float),
                np.asarray(strain_bundle.get("strain_pct_matrix", np.zeros((0, 0))), dtype=float),
                f"Layer {axis_label or 'Input'} Strain",
                f"layer-{_source_slug(axis_label or 'single')}-strain",
            ),
        ),
        _source_chart(
            "layer-velocity",
            "Layer Velocity",
            "Layer",
            "time",
            "Time (s)",
            "Velocity (m/s)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(direction_bundle.get("time", []), dtype=float),
                np.asarray(direction_bundle.get("vel_matrix", np.zeros((0, 0))), dtype=float),
                f"Layer {axis_label or 'Input'} Velocity",
                f"layer-{_source_slug(axis_label or 'single')}-vel",
            ),
        ),
        _source_chart(
            "layer-displacement",
            "Layer Displacement",
            "Layer",
            "time",
            "Time (s)",
            "Displacement (m)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(direction_bundle.get("time", []), dtype=float),
                np.asarray(direction_bundle.get("disp_matrix", np.zeros((0, 0))), dtype=float),
                f"Layer {axis_label or 'Input'} Displacement",
                f"layer-{_source_slug(axis_label or 'single')}-disp",
            ),
        ),
        _source_chart(
            "layer-tbdy-total",
            "Layer TBDY Total",
            "Layer",
            "time",
            "Time (s)",
            "TBDY Total (m)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(strain_bundle.get("time", []), dtype=float),
                np.asarray(strain_bundle.get("u_tbdy_total", np.zeros((0, 0))), dtype=float),
                f"Layer {axis_label or 'Input'} TBDY Total",
                f"layer-{_source_slug(axis_label or 'single')}-tbdy",
            ),
        ),
        _source_chart(
            "layer-stress-ratio",
            "Layer Stress Ratio",
            "Layer",
            "time",
            "Time (s)",
            "Stress Ratio",
            layer_views=_single_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Shear Stress Ratio", "Stress Ratio", "Shear/Eff. Vert."),
                f"Layer {axis_label or 'Input'} Stress Ratio",
                f"layer-{_source_slug(axis_label or 'single')}-stress-ratio",
            ),
        ),
        _source_chart(
            "layer-shear-stress",
            "Layer Shear Stress",
            "Layer",
            "time",
            "Time (s)",
            "Shear Stress (kPa)",
            layer_views=_single_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Shear Stress (kPa)", "Shear Stress", "Stress (kPa)"),
                f"Layer {axis_label or 'Input'} Shear Stress",
                f"layer-{_source_slug(axis_label or 'single')}-shear-stress",
            ),
        ),
        _source_chart(
            "layer-arias-intensity",
            "Layer Arias Intensity",
            "Layer",
            "time",
            "Time (s)",
            "Arias Intensity",
            layer_views=_single_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Arias Intensity", "Arias"),
                f"Layer {axis_label or 'Input'} Arias",
                f"layer-{_source_slug(axis_label or 'single')}-arias",
            ),
        ),
        _source_chart(
            "layer-housner-intensity",
            "Layer Housner Intensity",
            "Layer",
            "time",
            "Time (s)",
            "Housner Intensity",
            layer_views=_single_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Housner Intensity", "Housner"),
                f"Layer {axis_label or 'Input'} Housner",
                f"layer-{_source_slug(axis_label or 'single')}-housner",
            ),
        ),
        _source_chart(
            "layer-output-spectrum",
            "Layer Output Spectrum",
            "Layer",
            "spectrum",
            "Period (s)",
            "5% Damped Spectral",
            layer_views=_single_layer_curve_views(
                ("Period (sec)", "Period", "Period (s)"),
                ("5% Damped Spectral", "5% Damped Spectrum", "PSA (g)", "PSA"),
                f"Layer {axis_label or 'Input'} Spectrum",
                f"layer-{_source_slug(axis_label or 'single')}-spectrum",
            ),
        ),
        _source_chart(
            "layer-output-fourier",
            "Layer Output Fourier",
            "Layer",
            "fourier",
            "Frequency (Hz)",
            "Fourier Amplitude",
            layer_views=_single_layer_curve_views(
                ("Frequency (Hz)", "Frequency"),
                ("Fourier Amplitude", "Fourier"),
                f"Layer {axis_label or 'Input'} Fourier",
                f"layer-{_source_slug(axis_label or 'single')}-fourier",
            ),
        ),
        _source_chart(
            "layer-output-fourier-ratio",
            "Layer Output Fourier Ratio",
            "Layer",
            "fourier",
            "Frequency (Hz)",
            "Fourier Amplitude Ratio",
            layer_views=_single_layer_curve_views(
                ("Frequency (Hz)", "Frequency"),
                ("Fourier Amplitude Ratio", "Fourier Ratio", "Amplitude Ratio"),
                f"Layer {axis_label or 'Input'} Fourier Ratio",
                f"layer-{_source_slug(axis_label or 'single')}-fourier-ratio",
            ),
        ),
    ]

    spectra_views_psa: List[Dict[str, Any]] = []
    spectra_views_fourier: List[Dict[str, Any]] = []
    input_curves = _read_input_motion_curves(xl)
    for layer in layers:
        idx = int(layer["layerIndex"])
        layer_name = layer["layerLabel"]
        try:
            series = []
            if "psa" in input_curves:
                series.append(_source_series("Input PSA", *input_curves["psa"], series_key="input-psa"))
            period, psa = _read_layer_curve(
                xl,
                layer_name,
                ("Period (sec)", "Period", "Period (s)"),
                ("5% Damped Spectral", "5% Damped Spectrum", "PSA (g)", "PSA"),
            )
            series.append(_source_series("Layer PSA", period, psa, series_key="layer-psa"))
            clean_series = [item for item in series if isinstance(item, dict)]
            if clean_series:
                spectra_views_psa.append(
                    {"layerIndex": idx, "layerLabel": layer_name, "depth": layer["depth"], "series": clean_series}
                )
        except Exception:
            pass

        try:
            series = []
            if "fourier" in input_curves:
                series.append(_source_series("Input Fourier", *input_curves["fourier"], series_key="input-fourier"))
            freq, amp = _read_layer_curve(xl, layer_name, ("Frequency (Hz)", "Frequency"), ("Fourier Amplitude", "Fourier"))
            series.append(_source_series("Layer Fourier", freq, amp, series_key="layer-fourier"))
            clean_series = [item for item in series if isinstance(item, dict)]
            if clean_series:
                spectra_views_fourier.append(
                    {"layerIndex": idx, "layerLabel": layer_name, "depth": layer["depth"], "series": clean_series}
                )
        except Exception:
            pass

    charts.extend(
        [
            _source_chart(
                "layer-psa-compare",
                "Input vs Output PSA",
                "Layer",
                "spectrum",
                "Period (s)",
                "PSA (g)",
                layer_views=spectra_views_psa,
            ),
            _source_chart(
                "layer-fourier-compare",
                "Input vs Output Fourier",
                "Layer",
                "fourier",
                "Frequency (Hz)",
                "Fourier Amplitude",
                layer_views=spectra_views_fourier,
            ),
        ]
    )

    return _source_family("layer-series", "Layer Series", "time", charts, supports_layer_selection=True)


def _pair_layer_family(
    x_direction_bundle: Mapping[str, Any],
    y_direction_bundle: Mapping[str, Any],
    strain_bundle: Mapping[str, Any],
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
) -> Dict[str, Any] | None:
    def _paired_layer_curve_views(
        x_aliases: Sequence[str],
        y_aliases: Sequence[str],
        x_series_name: str,
        x_series_key: str,
        y_series_name: str,
        y_series_key: str,
        *,
        resultant_name: str | None = None,
        resultant_series_key: str | None = None,
    ) -> List[Dict[str, Any]]:
        views: List[Dict[str, Any]] = []
        for layer in layers:
            layer_name = str(layer["layerLabel"])
            try:
                x_curve_x, x_curve_y = _read_layer_curve(x_xl, layer_name, x_aliases, y_aliases)
                y_curve_x, y_curve_y = _read_layer_curve(y_xl, layer_name, x_aliases, y_aliases)
            except Exception:
                continue

            series: List[Dict[str, Any] | None] = [
                _source_series(x_series_name, x_curve_x, x_curve_y, series_key=x_series_key),
                _source_series(y_series_name, y_curve_x, y_curve_y, series_key=y_series_key),
            ]
            if resultant_name and resultant_series_key:
                common_x, x_interp, y_interp = _align_two_series(x_curve_x, x_curve_y, y_curve_x, y_curve_y)
                series.append(
                    _source_series(
                        resultant_name,
                        common_x,
                        np.sqrt(x_interp**2 + y_interp**2),
                        series_key=resultant_series_key,
                    )
                )

            clean_series = [item for item in series if isinstance(item, dict)]
            if not clean_series:
                continue
            views.append(
                {
                    "layerIndex": int(layer["layerIndex"]),
                    "layerLabel": layer_name,
                    "depth": float(layer["depth"]),
                    "series": clean_series,
                }
            )
        return views

    layers = _source_layers(x_direction_bundle.get("layer_names", []), x_direction_bundle.get("depths", []))
    charts = [
        _source_chart(
            "layer-acceleration",
            "Layer Acceleration",
            "Layer",
            "time",
            "Time (s)",
            "Acceleration (g)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(x_direction_bundle.get("time", []), dtype=float),
                np.asarray(x_direction_bundle.get("acc_matrix", np.zeros((0, 0))), dtype=float),
                "Layer X Acceleration",
                "layer-x-acc",
                np.asarray(y_direction_bundle.get("time", []), dtype=float),
                np.asarray(y_direction_bundle.get("acc_matrix", np.zeros((0, 0))), dtype=float),
                "Layer Y Acceleration",
                "layer-y-acc",
            ),
        ),
        _source_chart(
            "layer-strain",
            "Layer Strain",
            "Layer",
            "time",
            "Time (s)",
            "Strain (%)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(strain_bundle.get("time", []), dtype=float),
                np.asarray(strain_bundle.get("strain_x_pct_matrix", np.zeros((0, 0))), dtype=float),
                "Layer X Strain",
                "layer-x-strain",
                np.asarray(strain_bundle.get("time", []), dtype=float),
                np.asarray(strain_bundle.get("strain_y_pct_matrix", np.zeros((0, 0))), dtype=float),
                "Layer Y Strain",
                "layer-y-strain",
            ),
        ),
        _source_chart(
            "layer-velocity",
            "Layer Velocity",
            "Layer",
            "time",
            "Time (s)",
            "Velocity (m/s)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(x_direction_bundle.get("time", []), dtype=float),
                np.asarray(x_direction_bundle.get("vel_matrix", np.zeros((0, 0))), dtype=float),
                "Layer X Velocity",
                "layer-x-vel",
                np.asarray(y_direction_bundle.get("time", []), dtype=float),
                np.asarray(y_direction_bundle.get("vel_matrix", np.zeros((0, 0))), dtype=float),
                "Layer Y Velocity",
                "layer-y-vel",
            ),
        ),
        _source_chart(
            "layer-displacement",
            "Layer Displacement",
            "Layer",
            "time",
            "Time (s)",
            "Displacement (m)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(x_direction_bundle.get("time", []), dtype=float),
                np.asarray(x_direction_bundle.get("disp_matrix", np.zeros((0, 0))), dtype=float),
                "Layer X Displacement",
                "layer-x-disp",
                np.asarray(y_direction_bundle.get("time", []), dtype=float),
                np.asarray(y_direction_bundle.get("disp_matrix", np.zeros((0, 0))), dtype=float),
                "Layer Y Displacement",
                "layer-y-disp",
                resultant_name="Layer Resultant",
                resultant_series_key="layer-resultant-disp",
            ),
        ),
        _source_chart(
            "layer-tbdy-total",
            "Layer TBDY Total",
            "Layer",
            "time",
            "Time (s)",
            "TBDY Total (m)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(strain_bundle.get("time", []), dtype=float),
                np.asarray(strain_bundle.get("u_tbdy_total_x", np.zeros((0, 0))), dtype=float),
                "Layer X TBDY",
                "layer-x-tbdy",
                np.asarray(strain_bundle.get("time", []), dtype=float),
                np.asarray(strain_bundle.get("u_tbdy_total_y", np.zeros((0, 0))), dtype=float),
                "Layer Y TBDY",
                "layer-y-tbdy",
                resultant_name="Layer Resultant TBDY",
                resultant_series_key="layer-resultant-tbdy",
            ),
        ),
        _source_chart(
            "layer-stress-ratio",
            "Layer Stress Ratio",
            "Layer",
            "time",
            "Time (s)",
            "Stress Ratio",
            layer_views=_paired_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Shear Stress Ratio", "Stress Ratio", "Shear/Eff. Vert."),
                "Layer X Stress Ratio",
                "layer-x-stress-ratio",
                "Layer Y Stress Ratio",
                "layer-y-stress-ratio",
            ),
        ),
        _source_chart(
            "layer-shear-stress",
            "Layer Shear Stress",
            "Layer",
            "time",
            "Time (s)",
            "Shear Stress (kPa)",
            layer_views=_paired_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Shear Stress (kPa)", "Shear Stress", "Stress (kPa)"),
                "Layer X Shear Stress",
                "layer-x-shear-stress",
                "Layer Y Shear Stress",
                "layer-y-shear-stress",
            ),
        ),
        _source_chart(
            "layer-arias-intensity",
            "Layer Arias Intensity",
            "Layer",
            "time",
            "Time (s)",
            "Arias Intensity",
            layer_views=_paired_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Arias Intensity", "Arias"),
                "Layer X Arias",
                "layer-x-arias",
                "Layer Y Arias",
                "layer-y-arias",
            ),
        ),
        _source_chart(
            "layer-housner-intensity",
            "Layer Housner Intensity",
            "Layer",
            "time",
            "Time (s)",
            "Housner Intensity",
            layer_views=_paired_layer_curve_views(
                ("Time (s)", "Time", "Time (sec)"),
                ("Housner Intensity", "Housner"),
                "Layer X Housner",
                "layer-x-housner",
                "Layer Y Housner",
                "layer-y-housner",
            ),
        ),
        _source_chart(
            "layer-output-spectrum",
            "Layer Output Spectrum",
            "Layer",
            "spectrum",
            "Period (s)",
            "5% Damped Spectral",
            layer_views=_paired_layer_curve_views(
                ("Period (sec)", "Period", "Period (s)"),
                ("5% Damped Spectral", "5% Damped Spectrum", "PSA (g)", "PSA"),
                "Layer X Spectrum",
                "layer-x-spectrum",
                "Layer Y Spectrum",
                "layer-y-spectrum",
            ),
        ),
        _source_chart(
            "layer-output-fourier",
            "Layer Output Fourier",
            "Layer",
            "fourier",
            "Frequency (Hz)",
            "Fourier Amplitude",
            layer_views=_paired_layer_curve_views(
                ("Frequency (Hz)", "Frequency"),
                ("Fourier Amplitude", "Fourier"),
                "Layer X Fourier",
                "layer-x-fourier",
                "Layer Y Fourier",
                "layer-y-fourier",
            ),
        ),
        _source_chart(
            "layer-output-fourier-ratio",
            "Layer Output Fourier Ratio",
            "Layer",
            "fourier",
            "Frequency (Hz)",
            "Fourier Amplitude Ratio",
            layer_views=_paired_layer_curve_views(
                ("Frequency (Hz)", "Frequency"),
                ("Fourier Amplitude Ratio", "Fourier Ratio", "Amplitude Ratio"),
                "Layer X Fourier Ratio",
                "layer-x-fourier-ratio",
                "Layer Y Fourier Ratio",
                "layer-y-fourier-ratio",
            ),
        ),
    ]
    return _source_family("layer-series", "Layer Series", "time", charts, supports_layer_selection=True)


def _single_derived_family(
    file_name: str,
    summary_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame,
    input_motion_max_abs: float,
) -> Dict[str, Any] | None:
    approx_df = _build_input_motion_added_profile_df(profile_sheet_df, input_motion_max_abs, Path(file_name).stem)
    depths_summary = pd.to_numeric(summary_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    charts = [
        _pair_depth_series_chart(
            "derived-total-profile",
            "Total Profile",
            "Derived",
            depths_summary,
            [
                ("Profile Max Displacement", profile_sheet_df.get("Profile_raw_max_m", [])),
                ("Computed Total (TBDY)", summary_df.get("TBDY_total_max_m", [])),
                ("Approx Total (Ubase + Urel)", approx_df.iloc[:, 1] if approx_df.shape[1] > 1 else []),
            ],
            x_label="Displacement (m)",
        ),
    ]
    return _source_family("derived-profiles", "Derived Profiles", "depth", charts, supports_overlay=True)


def _pair_derived_family(
    comparison_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame,
    x_input_added_df: pd.DataFrame,
    y_input_added_df: pd.DataFrame,
) -> Dict[str, Any] | None:
    depths_profile = pd.to_numeric(profile_sheet_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    depths_compare = pd.to_numeric(comparison_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    charts = [
        _pair_depth_series_chart(
            "derived-x-profile",
            "Total X Profile",
            "Derived",
            depths_compare,
            [
                ("Profile X Max Displacement", profile_sheet_df.get("Profile_X_raw_max_m", [])),
                ("Computed X Total (TBDY)", comparison_df.get("X_tbdy_total_max_m", [])),
                ("Approx Total X (Ubase + Urel)", x_input_added_df.iloc[:, 1] if x_input_added_df.shape[1] > 1 else []),
            ],
            x_label="Displacement (m)",
        ),
        _pair_depth_series_chart(
            "derived-y-profile",
            "Total Y Profile",
            "Derived",
            depths_compare,
            [
                ("Profile Y Max Displacement", profile_sheet_df.get("Profile_Y_raw_max_m", [])),
                ("Computed Y Total (TBDY)", comparison_df.get("Y_tbdy_total_max_m", [])),
                ("Approx Total Y (Ubase + Urel)", y_input_added_df.iloc[:, 1] if y_input_added_df.shape[1] > 1 else []),
            ],
            x_label="Displacement (m)",
        ),
        _pair_depth_series_chart(
            "derived-resultant-profile",
            "Total Resultant Profile",
            "Derived",
            depths_profile,
            [
                ("Profile RSS Max Displacement", profile_sheet_df.get("Profile_RSS_raw_max_m", [])),
                ("Computed RSS Total (TBDY)", comparison_df.get("Total_tbdy_total_max_m", [])),
                ("TimeHist RSS Total", comparison_df.get("TimeHist_Resultant_total_m", [])),
            ],
            x_label="Displacement (m)",
        ),
    ]
    return _source_family("derived-profiles", "Derived Profiles", "depth", charts, supports_overlay=True)


def _db_source_family_single(bundle: Mapping[str, Any]) -> Dict[str, Any] | None:
    layers = _source_layers(bundle.get("layer_numbers", []), bundle.get("depths", []))
    charts = [
        _pair_depth_series_chart(
            "db-total-profile",
            "DB Total Profile",
            "DB",
            np.asarray(bundle["summary_df"]["Depth_m"], dtype=float),
            [("DB Total", bundle["summary_df"]["DB_Total_maxabs_m"])],
            x_label="Displacement (m)",
        ),
        _source_chart(
            "db-total-time",
            "DB Total Time",
            "DB",
            "time",
            "Time (s)",
            "Displacement (m)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(bundle.get("time", []), dtype=float),
                np.asarray(bundle.get("disp_matrix", np.zeros((0, 0))), dtype=float),
                "DB Total",
                "db-total",
            ),
        ),
        _pair_depth_series_chart(
            "db-relative-profile",
            "DB Relative Profile",
            "DB",
            np.asarray(bundle["summary_df"]["Depth_m"], dtype=float),
            [("DB Relative", bundle["summary_df"]["DB_Relative_maxabs_m"])],
            x_label="Displacement (m)",
        ),
        _source_chart(
            "db-relative-time",
            "DB Relative Time",
            "DB",
            "time",
            "Time (s)",
            "Displacement (m)",
            layer_views=_single_layer_views(
                layers,
                np.asarray(bundle.get("time", []), dtype=float),
                np.asarray(bundle.get("relative_matrix", np.zeros((0, 0))), dtype=float),
                "DB Relative",
                "db-relative",
            ),
        ),
    ]
    return _source_family("db-motion", "DB Motion", "time", charts, supports_layer_selection=True)


def _db_source_family_pair(summary_df: pd.DataFrame, x_bundle: Mapping[str, Any], y_bundle: Mapping[str, Any]) -> Dict[str, Any] | None:
    layers = _source_layers(x_bundle.get("layer_numbers", []), x_bundle.get("depths", []))
    charts = [
        _pair_depth_series_chart(
            "db-total-profile",
            "DB Total Resultant Profile",
            "DB",
            np.asarray(summary_df["Depth_m"], dtype=float),
            [
                ("DB Total X", summary_df.get("X_total_maxabs_m", [])),
                ("DB Total Y", summary_df.get("Y_total_maxabs_m", [])),
                ("DB Total RSS", summary_df.get("Total_resultant_maxabs_m", [])),
            ],
            x_label="Displacement (m)",
        ),
        _source_chart(
            "db-total-time",
            "DB Total Time",
            "DB",
            "time",
            "Time (s)",
            "Displacement (m)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(x_bundle.get("time", []), dtype=float),
                np.asarray(x_bundle.get("disp_matrix", np.zeros((0, 0))), dtype=float),
                "DB X Total",
                "db-x-total",
                np.asarray(y_bundle.get("time", []), dtype=float),
                np.asarray(y_bundle.get("disp_matrix", np.zeros((0, 0))), dtype=float),
                "DB Y Total",
                "db-y-total",
                resultant_name="DB Resultant",
                resultant_series_key="db-total-resultant",
            ),
        ),
        _pair_depth_series_chart(
            "db-relative-profile",
            "DB Relative Resultant Profile",
            "DB",
            np.asarray(summary_df["Depth_m"], dtype=float),
            [
                ("DB Relative X", summary_df.get("X_relative_maxabs_m", [])),
                ("DB Relative Y", summary_df.get("Y_relative_maxabs_m", [])),
                ("DB Relative RSS", summary_df.get("Relative_resultant_maxabs_m", [])),
            ],
            x_label="Displacement (m)",
        ),
        _source_chart(
            "db-relative-time",
            "DB Relative Time",
            "DB",
            "time",
            "Time (s)",
            "Displacement (m)",
            layer_views=_paired_layer_views(
                layers,
                np.asarray(x_bundle.get("time", []), dtype=float),
                np.asarray(x_bundle.get("relative_matrix", np.zeros((0, 0))), dtype=float),
                "DB X Relative",
                "db-x-relative",
                np.asarray(y_bundle.get("time", []), dtype=float),
                np.asarray(y_bundle.get("relative_matrix", np.zeros((0, 0))), dtype=float),
                "DB Y Relative",
                "db-y-relative",
                resultant_name="DB Relative Resultant",
                resultant_series_key="db-relative-resultant",
            ),
        ),
    ]
    return _source_family("db-motion", "DB Motion", "time", charts, supports_layer_selection=True, supports_overlay=True)


def _build_single_source_catalog_entry(
    xl: pd.ExcelFile,
    file_name: str,
    axis_label: str,
    summary_df: pd.DataFrame,
    direction_bundle: Mapping[str, Any],
    strain_bundle: Mapping[str, Any],
    profile_sheet_df: pd.DataFrame,
    *,
    artifact_pair_keys: Sequence[str] | None = None,
) -> Dict[str, Any]:
    input_proxy = np.asarray(strain_bundle.get("u_input_proxy", np.array([])), dtype=float)
    input_motion_max_abs = float(np.max(np.abs(input_proxy))) if input_proxy.size else float("nan")
    source_label = Path(file_name).stem
    families = [
        _single_input_motion_family(xl, axis_label),
        _single_profile_family(xl),
        _single_layer_family(xl, axis_label, direction_bundle, strain_bundle),
        _single_derived_family(file_name, summary_df, profile_sheet_df, input_motion_max_abs),
    ]
    return _source_entry(
        f"source-single-{_source_slug(file_name)}",
        source_label,
        "single",
        axis_label,
        f"SINGLE|{source_label}",
        families,
        artifact_pair_keys=list(artifact_pair_keys or [f"SINGLE|{source_label}", f"METHOD2|{source_label}"]),
    )


def _build_pair_member_summary_df(comparison_df: pd.DataFrame, axis_label: str) -> pd.DataFrame:
    axis_key = str(axis_label or "").upper()
    base_col = "X_base_rel_max_m" if axis_key == "X" else "Y_base_rel_max_m"
    tbdy_col = "X_tbdy_total_max_m" if axis_key == "X" else "Y_tbdy_total_max_m"
    out = pd.DataFrame(
        {
            "Layer_Index": comparison_df.get(
                "Layer_Index",
                pd.Series(np.arange(1, len(comparison_df) + 1, dtype=int)),
            ),
            "Depth_m": comparison_df.get("Depth_m", pd.Series(dtype=float)),
        }
    )
    if base_col in comparison_df.columns:
        out["Base_rel_max_m"] = comparison_df[base_col]
    if tbdy_col in comparison_df.columns:
        out["TBDY_total_max_m"] = comparison_df[tbdy_col]
    return out


def _build_pair_member_strain_bundle(strain_bundle: Mapping[str, Any], axis_label: str) -> Dict[str, Any]:
    axis_key = str(axis_label or "").upper()
    suffix = "x" if axis_key == "X" else "y"
    out: Dict[str, Any] = {
        "time": strain_bundle.get("time", np.array([])),
        "u_input_proxy": strain_bundle.get(f"u_input_proxy_{suffix}", np.array([])),
        "u_tbdy_total": strain_bundle.get(f"u_tbdy_total_{suffix}", np.zeros((0, 0))),
        "strain_pct_matrix": strain_bundle.get(f"strain_{suffix}_pct_matrix", np.zeros((0, 0))),
    }
    alt_total = strain_bundle.get(f"u_tbdy_total_{suffix}_alt")
    if alt_total is not None:
        out["u_tbdy_total_alt"] = alt_total
    return out


def _build_pair_source_catalog_entries(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    x_name: str,
    y_name: str,
    comparison_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame,
    x_direction_bundle: Mapping[str, Any],
    y_direction_bundle: Mapping[str, Any],
    strain_bundle: Mapping[str, Any],
) -> List[Dict[str, Any]]:
    x_profile_view = _build_single_profile_sheet_df(x_xl)
    y_profile_view = _build_single_profile_sheet_df(y_xl)
    x_input_proxy = np.asarray(strain_bundle.get("u_input_proxy_x", np.array([])), dtype=float)
    y_input_proxy = np.asarray(strain_bundle.get("u_input_proxy_y", np.array([])), dtype=float)
    x_input_added_df = _build_input_motion_added_profile_df(
        x_profile_view,
        float(np.max(np.abs(x_input_proxy))) if x_input_proxy.size else float("nan"),
        Path(x_name).stem,
    )
    y_input_added_df = _build_input_motion_added_profile_df(
        y_profile_view,
        float(np.max(np.abs(y_input_proxy))) if y_input_proxy.size else float("nan"),
        Path(y_name).stem,
    )
    pair_key = _build_pair_key(x_name, y_name)
    source_label = f"{Path(x_name).stem} | {Path(y_name).stem}"
    x_source = _build_single_source_catalog_entry(
        x_xl,
        x_name,
        "X",
        _build_pair_member_summary_df(comparison_df, "X"),
        x_direction_bundle,
        _build_pair_member_strain_bundle(strain_bundle, "X"),
        x_profile_view,
        artifact_pair_keys=[f"METHOD2|{Path(x_name).stem}"],
    )
    y_source = _build_single_source_catalog_entry(
        y_xl,
        y_name,
        "Y",
        _build_pair_member_summary_df(comparison_df, "Y"),
        y_direction_bundle,
        _build_pair_member_strain_bundle(strain_bundle, "Y"),
        y_profile_view,
        artifact_pair_keys=[f"METHOD2|{Path(y_name).stem}"],
    )
    pair_source = _source_entry(
        f"source-pair-{_source_slug(pair_key)}",
        source_label,
        "pair",
        "XY",
        pair_key,
        [
            _pair_input_motion_family(x_xl, y_xl),
            _pair_profile_family(x_xl, y_xl, profile_sheet_df),
            _pair_layer_family(x_direction_bundle, y_direction_bundle, strain_bundle, x_xl, y_xl),
            _pair_derived_family(comparison_df, profile_sheet_df, x_input_added_df, y_input_added_df),
        ],
        artifact_pair_keys=[pair_key],
    )
    return [entry for entry in (x_source, y_source, pair_source) if isinstance(entry, dict) and entry.get("families")]


def _build_db_single_source_catalog_entry(file_name: str, bundle: Mapping[str, Any]) -> Dict[str, Any]:
    source_label = Path(file_name).stem
    pair_key = f"DB_SINGLE|{bundle.get('recordLabel', source_label)}"
    return _source_entry(
        f"source-db-single-{_source_slug(file_name)}",
        source_label,
        "db_single",
        str(bundle.get("axis", "")),
        pair_key,
        [_db_source_family_single(bundle)],
        artifact_pair_keys=[pair_key, f"DB_METHOD2|{source_label}", "DB_METHOD3|ALL"],
    )


def _build_db_pair_source_catalog_entry(
    x_name: str,
    y_name: str,
    summary_df: pd.DataFrame,
    x_bundle: Mapping[str, Any],
    y_bundle: Mapping[str, Any],
) -> Dict[str, Any]:
    pair_key = f"DB|{_build_pair_key(x_name, y_name)}"
    return _source_entry(
        f"source-db-pair-{_source_slug(pair_key)}",
        f"{Path(x_name).stem} | {Path(y_name).stem}",
        "db_pair",
        "XY",
        pair_key,
        [_db_source_family_pair(summary_df, x_bundle, y_bundle)],
        artifact_pair_keys=[pair_key, f"DB_METHOD2|{Path(x_name).stem}", f"DB_METHOD2|{Path(y_name).stem}", "DB_METHOD3|ALL"],
    )


def _build_method3_source_catalog_entry(
    profile_x_df: pd.DataFrame,
    profile_y_df: pd.DataFrame,
    input_added_profile_x_df: pd.DataFrame,
    input_added_profile_y_df: pd.DataFrame,
) -> Dict[str, Any] | None:
    charts = [
        _pair_depth_series_chart(
            "method3-x-profiles",
            "Method-3 X Profiles",
            "Method3_Profile_X",
            pd.to_numeric(profile_x_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float),
            [(str(column), profile_x_df[column]) for column in profile_x_df.columns if column != "Depth_m"],
            x_label="Displacement (m)",
        ),
        _pair_depth_series_chart(
            "method3-y-profiles",
            "Method-3 Y Profiles",
            "Method3_Profile_Y",
            pd.to_numeric(profile_y_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float),
            [(str(column), profile_y_df[column]) for column in profile_y_df.columns if column != "Depth_m"],
            x_label="Displacement (m)",
        ),
        _pair_depth_series_chart(
            "method3-x-approx-total",
            "Method-3 X Approx Total",
            "Method3_ApproxTotal_X",
            pd.to_numeric(input_added_profile_x_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float),
            [(str(column), input_added_profile_x_df[column]) for column in input_added_profile_x_df.columns if column != "Depth_m"],
            x_label="Displacement (m)",
        ),
        _pair_depth_series_chart(
            "method3-y-approx-total",
            "Method-3 Y Approx Total",
            "Method3_ApproxTotal_Y",
            pd.to_numeric(input_added_profile_y_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float),
            [(str(column), input_added_profile_y_df[column]) for column in input_added_profile_y_df.columns if column != "Depth_m"],
            x_label="Displacement (m)",
        ),
    ]
    family = _source_family("method3-aggregate", "Method-3 Aggregate", "depth", charts, supports_overlay=True)
    if family is None:
        return None
    return _source_entry(
        "source-method3-all",
        "Method-3 Aggregate",
        "method3_aggregate",
        "XY",
        "METHOD3|ALL",
        [family],
        artifact_pair_keys=["METHOD3|ALL"],
    )


def _profile_layer_count_from_frame(frame: pd.DataFrame | None) -> int:
    if frame is None or frame.empty or "Depth_m" not in frame.columns:
        return 0
    depths = pd.to_numeric(frame["Depth_m"], errors="coerce").to_numpy(dtype=float)
    return int(np.count_nonzero(np.isfinite(depths)))


def _limited_layer_warning(available_count: int, profile_count: int) -> str:
    return (
        f"Layer sheet kapsami sinirli ({available_count}/{profile_count}); "
        "yalniz yaklasik/profile tabanli toplam profiller guvenle gosterildi."
    )


def _pair_total_profile_from_base_refs(
    strain_bundle: Mapping[str, Any],
    base_ref_x: Any,
    base_ref_y: Any,
) -> np.ndarray:
    u_rel_x = np.asarray(strain_bundle.get("u_rel_base_x", np.zeros((0, 0))), dtype=float)
    u_rel_y = np.asarray(strain_bundle.get("u_rel_base_y", np.zeros((0, 0))), dtype=float)
    base_x = np.asarray(base_ref_x, dtype=float).reshape(-1)
    base_y = np.asarray(base_ref_y, dtype=float).reshape(-1)
    if u_rel_x.ndim != 2 or u_rel_y.ndim != 2 or base_x.size == 0 or base_y.size == 0:
        return np.zeros(0, dtype=float)
    step_count = min(int(u_rel_x.shape[1]), int(u_rel_y.shape[1]), int(base_x.size), int(base_y.size))
    layer_count = min(int(u_rel_x.shape[0]), int(u_rel_y.shape[0]))
    if step_count <= 0 or layer_count <= 0:
        return np.zeros(0, dtype=float)
    total_x = u_rel_x[:layer_count, :step_count] + base_x[:step_count][None, :]
    total_y = u_rel_y[:layer_count, :step_count] + base_y[:step_count][None, :]
    return _matrix_resultant_envelope(total_x, total_y)


def _single_total_profile_from_base_ref(strain_bundle: Mapping[str, Any], base_ref: Any) -> np.ndarray:
    u_rel = np.asarray(strain_bundle.get("u_rel_base", np.zeros((0, 0))), dtype=float)
    base = np.asarray(base_ref, dtype=float).reshape(-1)
    if u_rel.ndim != 2 or base.size == 0:
        return np.zeros(0, dtype=float)
    step_count = min(int(u_rel.shape[1]), int(base.size))
    layer_count = int(u_rel.shape[0])
    if step_count <= 0 or layer_count <= 0:
        return np.zeros(0, dtype=float)
    total = u_rel[:layer_count, :step_count] + base[:step_count][None, :]
    return np.max(np.abs(total), axis=1)


def _compute_pair_deepest_base_refs(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    strain_bundle: Mapping[str, Any],
    options: Mapping[str, Any] | None = None,
) -> Tuple[np.ndarray | None, np.ndarray | None, str]:
    layer_names = list(strain_bundle.get("layer_names", []))
    time = np.asarray(strain_bundle.get("time", np.array([])), dtype=float)
    if not layer_names or time.size == 0:
        return None, None, "Deepest layer proxy icin gerekli strain zaman tabani bulunamadi."
    deepest_layer = str(layer_names[-1])
    try:
        t_deep_x, a_deep_x = _read_layer_column(x_xl, deepest_layer, "Acceleration (g)")
        t_deep_y, a_deep_y = _read_layer_column(y_xl, deepest_layer, "Acceleration (g)")
        deep_x = _acc_to_disp_dual(t_deep_x, a_deep_x, options=options)["primary"]
        deep_y = _acc_to_disp_dual(t_deep_y, a_deep_y, options=options)["primary"]
    except Exception as exc:  # noqa: BLE001
        return None, None, f"Deepest layer proxy okunamadi: {exc}"
    return (
        np.interp(time, t_deep_x, np.asarray(deep_x, dtype=float)),
        np.interp(time, t_deep_y, np.asarray(deep_y, dtype=float)),
        "Strain bazli goreceli deplasman deepest layer proxy ile toplandi.",
    )


def _compute_single_deepest_base_ref(
    xl: pd.ExcelFile,
    strain_bundle: Mapping[str, Any],
    options: Mapping[str, Any] | None = None,
) -> Tuple[np.ndarray | None, str]:
    layer_names = list(strain_bundle.get("layer_names", []))
    time = np.asarray(strain_bundle.get("time", np.array([])), dtype=float)
    if not layer_names or time.size == 0:
        return None, "Deepest layer proxy icin gerekli strain zaman tabani bulunamadi."
    deepest_layer = str(layer_names[-1])
    try:
        t_deep, a_deep = _read_layer_column(xl, deepest_layer, "Acceleration (g)")
        deep = _acc_to_disp_dual(t_deep, a_deep, options=options)["primary"]
    except Exception as exc:  # noqa: BLE001
        return None, f"Deepest layer proxy okunamadi: {exc}"
    return np.interp(time, t_deep, np.asarray(deep, dtype=float)), "Strain bazli goreceli deplasman deepest layer proxy ile toplandi."


def _build_pair_summary_entry(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    x_name: str,
    y_name: str,
    comparison_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame,
    strain_bundle: Mapping[str, Any],
    x_direction_bundle: Mapping[str, Any],
    y_direction_bundle: Mapping[str, Any],
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    depths = pd.to_numeric(comparison_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    available_count = int(depths.size)
    profile_count = _profile_layer_count_from_frame(profile_sheet_df)
    has_full_layer_coverage = profile_count <= 0 or available_count >= profile_count
    warnings: List[str] = []
    if not has_full_layer_coverage:
        warnings.append(_limited_layer_warning(available_count, profile_count))

    input_values = _pair_total_profile_from_base_refs(
        strain_bundle,
        strain_bundle.get("u_input_proxy_x", np.array([])),
        strain_bundle.get("u_input_proxy_y", np.array([])),
    )
    deepest_base_x: np.ndarray | None = None
    deepest_base_y: np.ndarray | None = None
    deepest_reason = "Strain bazli goreceli deplasman deepest layer proxy ile toplandi."
    if str(strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE)) == "deepest_layer":
        deepest_base_x = np.asarray(strain_bundle.get("u_base_ref_x", np.array([])), dtype=float)
        deepest_base_y = np.asarray(strain_bundle.get("u_base_ref_y", np.array([])), dtype=float)
    else:
        deepest_base_x, deepest_base_y, deepest_reason = _compute_pair_deepest_base_refs(
            x_xl,
            y_xl,
            strain_bundle,
            options=options,
        )
    deepest_values = (
        _pair_total_profile_from_base_refs(strain_bundle, deepest_base_x, deepest_base_y)
        if deepest_base_x is not None and deepest_base_y is not None
        else np.zeros(0, dtype=float)
    )
    approximate_values = pd.to_numeric(
        comparison_df.get("Total_profile_offset_total_est_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)
    indirect_values = pd.to_numeric(
        comparison_df.get("TimeHist_Resultant_total_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)
    reference_values = pd.to_numeric(
        comparison_df.get("Profile_RSS_total_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)

    variants = [
        _build_method_variant(
            "strain_input_total",
            depths,
            input_values,
            valid=has_full_layer_coverage and _series_has_finite_values(input_values),
            reason=(
                "Strain bazli goreceli deplasman Input Motion proxy ile toplandi."
                if has_full_layer_coverage and _series_has_finite_values(input_values)
                else "Tum layer sheet'leri bulunmadigi icin strain + input proxy profili gizlendi."
            ),
        ),
        _build_method_variant(
            "strain_deepest_total",
            depths,
            deepest_values,
            valid=has_full_layer_coverage and _series_has_finite_values(deepest_values),
            reason=(
                deepest_reason
                if has_full_layer_coverage and _series_has_finite_values(deepest_values)
                else (
                    "Tum layer sheet'leri bulunmadigi icin strain + deepest layer proxy profili gizlendi."
                    if not has_full_layer_coverage
                    else deepest_reason
                )
            ),
        ),
        _build_method_variant(
            "profile_offset_total",
            depths,
            approximate_values,
            valid=_series_has_finite_values(approximate_values),
            reason=(
                "Profile taban ofseti ve base-relative zarfi uzerinden yaklasik toplam profil uretildi."
                if _series_has_finite_values(approximate_values)
                else "Profile ofset tabanli yaklasik toplam profil uretilemedi."
            ),
        ),
        _build_method_variant(
            "time_history_total",
            depths,
            indirect_values,
            valid=has_full_layer_coverage and _series_has_finite_values(indirect_values),
            reason=(
                "Layer ivmeleri integre edilerek dolayli maksimum toplam profil uretildi."
                if has_full_layer_coverage and _series_has_finite_values(indirect_values)
                else "Tum layer ivme kayitlari olmadigi icin dolayli zaman-gecmis profili gizlendi."
            ),
        ),
        _build_method_variant(
            "profile_reference_total",
            depths,
            reference_values,
            valid=_series_has_finite_values(reference_values),
            reason=(
                "DEEPSOIL Profile sheet maksimum deplasman referansi."
                if _series_has_finite_values(reference_values)
                else "Profile referans profili bulunamadi."
            ),
        ),
    ]

    pair_key = _build_pair_key(x_name, y_name)
    return _build_summary_entry(
        f"summary-pair-{_source_slug(pair_key)}",
        f"{_record_label_from_name(x_name)} | {_record_label_from_name(y_name)}",
        "pair",
        "XY",
        pair_key,
        variants,
        input_kind="xlsx",
        available_layer_count=available_count,
        profile_layer_count=profile_count,
        warnings=warnings,
        detail_source_ids=[
            f"source-pair-{_source_slug(pair_key)}",
            f"source-single-{_source_slug(x_name)}",
            f"source-single-{_source_slug(y_name)}",
        ],
        artifact_pair_keys=[pair_key, f"METHOD2|{Path(x_name).stem}", f"METHOD2|{Path(y_name).stem}"],
    )


def _build_single_summary_entry(
    xl: pd.ExcelFile,
    file_name: str,
    axis_label: str,
    summary_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame,
    strain_bundle: Mapping[str, Any],
    direction_bundle: Mapping[str, Any],
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    depths = pd.to_numeric(summary_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    available_count = int(depths.size)
    profile_count = _profile_layer_count_from_frame(profile_sheet_df)
    has_full_layer_coverage = profile_count <= 0 or available_count >= profile_count
    warnings: List[str] = []
    if not has_full_layer_coverage:
        warnings.append(_limited_layer_warning(available_count, profile_count))

    input_values = _single_total_profile_from_base_ref(strain_bundle, strain_bundle.get("u_input_proxy", np.array([])))
    deepest_base, deepest_reason = (
        (np.asarray(strain_bundle.get("u_base_ref", np.array([])), dtype=float), "Strain bazli goreceli deplasman deepest layer proxy ile toplandi.")
        if str(strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE)) == "deepest_layer"
        else _compute_single_deepest_base_ref(xl, strain_bundle, options=options)
    )
    deepest_values = _single_total_profile_from_base_ref(strain_bundle, deepest_base) if deepest_base is not None else np.zeros(0, dtype=float)
    approximate_values = pd.to_numeric(
        summary_df.get("Profile_offset_total_est_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)
    indirect_values = pd.to_numeric(
        summary_df.get("TimeHist_maxabs_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)
    reference_values = pd.to_numeric(
        summary_df.get("Profile_max_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)

    variants = [
        _build_method_variant(
            "strain_input_total",
            depths,
            input_values,
            valid=has_full_layer_coverage and _series_has_finite_values(input_values),
            reason=(
                "Strain bazli goreceli deplasman Input Motion proxy ile toplandi."
                if has_full_layer_coverage and _series_has_finite_values(input_values)
                else "Tum layer sheet'leri bulunmadigi icin strain + input proxy profili gizlendi."
            ),
            display_label=f"Strain + Input Proxy ({axis_label})",
        ),
        _build_method_variant(
            "strain_deepest_total",
            depths,
            deepest_values,
            valid=has_full_layer_coverage and _series_has_finite_values(deepest_values),
            reason=(
                deepest_reason
                if has_full_layer_coverage and _series_has_finite_values(deepest_values)
                else (
                    "Tum layer sheet'leri bulunmadigi icin strain + deepest layer proxy profili gizlendi."
                    if not has_full_layer_coverage
                    else deepest_reason
                )
            ),
            display_label=f"Strain + Deepest Layer Proxy ({axis_label})",
        ),
        _build_method_variant(
            "profile_offset_total",
            depths,
            approximate_values,
            valid=_series_has_finite_values(approximate_values),
            reason=(
                "Profile taban ofseti ve base-relative zarfi uzerinden yaklasik toplam profil uretildi."
                if _series_has_finite_values(approximate_values)
                else "Profile ofset tabanli yaklasik toplam profil uretilemedi."
            ),
            display_label=f"Profile Offset Approximation ({axis_label})",
        ),
        _build_method_variant(
            "time_history_total",
            depths,
            indirect_values,
            valid=has_full_layer_coverage and _series_has_finite_values(indirect_values),
            reason=(
                "Layer ivmeleri integre edilerek dolayli maksimum toplam profil uretildi."
                if has_full_layer_coverage and _series_has_finite_values(indirect_values)
                else "Tum layer ivme kayitlari olmadigi icin dolayli zaman-gecmis profili gizlendi."
            ),
            display_label=f"Time-History Indirect ({axis_label})",
        ),
        _build_method_variant(
            "profile_reference_total",
            depths,
            reference_values,
            valid=_series_has_finite_values(reference_values),
            reason=(
                "DEEPSOIL Profile sheet maksimum deplasman referansi."
                if _series_has_finite_values(reference_values)
                else "Profile referans profili bulunamadi."
            ),
            display_label=f"Profile Reference ({axis_label})",
        ),
    ]

    record_label = _record_label_from_name(file_name)
    pair_key = f"SINGLE|{Path(file_name).stem}"
    return _build_summary_entry(
        f"summary-single-{_source_slug(file_name)}",
        record_label,
        "single",
        axis_label,
        pair_key,
        variants,
        input_kind="xlsx",
        available_layer_count=available_count,
        profile_layer_count=profile_count,
        warnings=warnings,
        detail_source_ids=[f"source-single-{_source_slug(file_name)}"],
        artifact_pair_keys=[pair_key, f"METHOD2|{Path(file_name).stem}"],
    )


def _build_db_pair_summary_entry(
    x_name: str,
    y_name: str,
    summary_df: pd.DataFrame,
) -> Dict[str, Any]:
    depths = pd.to_numeric(summary_df.get("Depth_m"), errors="coerce").to_numpy(dtype=float)
    direct_values = pd.to_numeric(
        summary_df.get("Total_resultant_maxabs_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)
    pair_key = f"DB|{_build_pair_key(x_name, y_name)}"
    variants = [
        _build_method_variant(
            "db_direct_total",
            depths,
            direct_values,
            valid=_series_has_finite_values(direct_values),
            reason=(
                "DEEPSOIL veritabanindaki toplam deplasman kolonlari dogrudan okundu."
                if _series_has_finite_values(direct_values)
                else "DB direct toplam deplasman profili bulunamadi."
            ),
        )
    ]
    return _build_summary_entry(
        f"summary-db-pair-{_source_slug(pair_key)}",
        f"{_record_label_from_name(x_name)} | {_record_label_from_name(y_name)}",
        "db_pair",
        "XY",
        pair_key,
        variants,
        input_kind="db",
        available_layer_count=int(depths.size),
        profile_layer_count=int(depths.size),
        detail_source_ids=[f"source-db-pair-{_source_slug(pair_key)}"],
        artifact_pair_keys=[pair_key, f"DB_METHOD2|{Path(x_name).stem}", f"DB_METHOD2|{Path(y_name).stem}"],
    )


def _build_db_single_summary_entry(file_name: str, bundle: Mapping[str, Any]) -> Dict[str, Any]:
    summary_df = bundle.get("summary_df", pd.DataFrame())
    if not isinstance(summary_df, pd.DataFrame):
        summary_df = pd.DataFrame()
    depths = pd.to_numeric(summary_df.get("Depth_m", pd.Series(dtype=float)), errors="coerce").to_numpy(dtype=float)
    direct_values = pd.to_numeric(
        summary_df.get("DB_Total_maxabs_m", pd.Series(dtype=float)),
        errors="coerce",
    ).to_numpy(dtype=float)
    record_label = str(bundle.get("recordLabel") or _record_label_from_name(file_name))
    pair_key = f"DB_SINGLE|{record_label}"
    variants = [
        _build_method_variant(
            "db_direct_total",
            depths,
            direct_values,
            valid=_series_has_finite_values(direct_values),
            reason=(
                "DEEPSOIL veritabanindaki toplam deplasman kolonlari dogrudan okundu."
                if _series_has_finite_values(direct_values)
                else "DB direct toplam deplasman profili bulunamadi."
            ),
            display_label=f"DB Direct Total ({bundle.get('axis', '') or 'Single'})",
        )
    ]
    return _build_summary_entry(
        f"summary-db-single-{_source_slug(file_name)}",
        record_label,
        "db_single",
        str(bundle.get("axis", "")),
        pair_key,
        variants,
        input_kind="db",
        available_layer_count=int(depths.size),
        profile_layer_count=int(depths.size),
        detail_source_ids=[f"source-db-single-{_source_slug(file_name)}"],
        artifact_pair_keys=[pair_key, f"DB_METHOD2|{record_label}"],
    )


def _to_float(value: Any, default: float) -> float:
    try:
        out = float(value)
    except (TypeError, ValueError):
        return default
    return out if np.isfinite(out) else default


def _to_bool(value: Any, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"1", "true", "yes", "on"}:
            return True
        if v in {"0", "false", "no", "off"}:
            return False
    return default


def _normalize_base_reference(value: Any) -> str:
    ref = str(value or DEFAULT_BASE_REFERENCE).strip().lower()
    if ref in {"deepest", "deep", "deepest_layer", "deepest-layer", "rock", "bedrock"}:
        return "deepest_layer"
    return "input"


def _include_resultant_profiles(options: Mapping[str, Any] | None) -> bool:
    cfg = options or {}
    return _to_bool(cfg.get("includeResultantProfiles", True), True)


def _use_db3_directly(options: Mapping[str, Any] | None) -> bool:
    cfg = options or {}
    return _to_bool(cfg.get("useDb3Directly", False), False)


def _primary_outputs_enabled(options: Mapping[str, Any] | None) -> bool:
    cfg = options or {}
    return not _to_bool(cfg.get("skipPrimaryOutputs", False), False)


def _method23_outputs_enabled(options: Mapping[str, Any] | None) -> bool:
    cfg = options or {}
    return not _to_bool(cfg.get("skipMethod23Outputs", False), False)


def _highpass_config(options: Mapping[str, Any] | None) -> Tuple[bool, float, float]:
    cfg = options or {}
    enabled = _to_bool(cfg.get("highpassEnabled", True), True)
    cutoff = max(0.0, _to_float(cfg.get("highpassCutoffHz"), DEFAULT_HIGHPASS_CUTOFF_HZ))
    transition = max(0.0, _to_float(cfg.get("highpassTransitionHz"), DEFAULT_HIGHPASS_TRANSITION_HZ))
    return enabled, cutoff, transition


def _layer_sort_key(name: str) -> Tuple[int, str]:
    match = re.search(r"(\d+)$", name)
    if match:
        return (int(match.group(1)), name)
    return (10**9, name)


def _list_layer_sheets(xl: pd.ExcelFile) -> List[str]:
    sheets = [name for name in xl.sheet_names if name.startswith("Layer")]
    return sorted(sheets, key=_layer_sort_key)


def _ensure_common_layers(x_xl: pd.ExcelFile, y_xl: pd.ExcelFile) -> List[str]:
    x_layers = _list_layer_sheets(x_xl)
    y_layers = _list_layer_sheets(y_xl)
    if not x_layers or not y_layers:
        raise ValueError("Missing Layer sheets in one or both files.")
    if x_layers != y_layers:
        raise ValueError("Layer sheet sets are not identical between X and Y files.")
    return x_layers


def _parse_sheet_cached(xl: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    cache = getattr(xl, "_deepsoil_sheet_cache", None)
    if cache is None:
        cache = {}
        setattr(xl, "_deepsoil_sheet_cache", cache)

    cached = cache.get(sheet_name)
    if cached is None:
        cached = xl.parse(sheet_name)
        cache[sheet_name] = cached
    return cached


def _resolve_layer_column_name(columns: Sequence[Any], requested: str) -> str | None:
    requested_key = str(requested or "").strip().lower()
    aliases = {
        "time (s)": ("time (s)", "time", "time (sec)", "time (seconds)"),
        "strain (%)": ("strain (%)", "shear strain (%)", "shear strain"),
    }.get(requested_key, (requested_key,))

    normalized = {}
    for column in columns:
        if column is None:
            continue
        normalized[str(column).strip().lower()] = str(column)

    for alias in aliases:
        candidate = normalized.get(str(alias).strip().lower())
        if candidate is not None:
            return candidate
    return None


def _read_layer_column(xl: pd.ExcelFile, layer_name: str, value_column: str) -> Tuple[np.ndarray, np.ndarray]:
    df = _parse_sheet_cached(xl, layer_name)
    time_column = _resolve_layer_column_name(df.columns, "Time (s)")
    resolved_value_column = _resolve_layer_column_name(df.columns, value_column)
    if time_column is None or resolved_value_column is None:
        raise ValueError(f"Sheet '{layer_name}' is missing 'Time (s)' or '{value_column}'.")

    data = df[[time_column, resolved_value_column]].copy()
    data.columns = ["Time (s)", value_column]
    data["Time (s)"] = pd.to_numeric(data["Time (s)"], errors="coerce")
    data[value_column] = pd.to_numeric(data[value_column], errors="coerce")
    data = data.dropna(subset=["Time (s)", value_column]).sort_values("Time (s)")

    if data.empty:
        raise ValueError(f"Sheet '{layer_name}' has no numeric rows for '{value_column}'.")

    return data["Time (s)"].to_numpy(dtype=float), data[value_column].to_numpy(dtype=float)


def _align_two_series(
    time_x: np.ndarray,
    value_x: np.ndarray,
    time_y: np.ndarray,
    value_y: np.ndarray,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    if time_x.size == time_y.size and np.allclose(time_x, time_y):
        return time_x, value_x, value_y

    start = max(float(time_x[0]), float(time_y[0]))
    end = min(float(time_x[-1]), float(time_y[-1]))
    if end <= start:
        raise ValueError("X and Y time windows do not overlap.")

    dx = np.median(np.diff(time_x)) if time_x.size > 1 else 0.01
    dy = np.median(np.diff(time_y)) if time_y.size > 1 else 0.01
    dt = min(dx, dy)
    if not np.isfinite(dt) or dt <= 0:
        dt = 0.01

    count = int(math.floor((end - start) / dt)) + 1
    common_time = start + np.arange(count, dtype=float) * dt

    x_interp = np.interp(common_time, time_x, value_x)
    y_interp = np.interp(common_time, time_y, value_y)
    return common_time, x_interp, y_interp


def parse_profile_thickness(xl: pd.ExcelFile) -> Tuple[np.ndarray, np.ndarray]:
    if "Profile" not in xl.sheet_names:
        raise ValueError("Missing 'Profile' sheet.")

    profile = _parse_sheet_cached(xl, "Profile")
    data = profile.iloc[1:].reset_index(drop=True)
    if data.empty:
        raise ValueError("Profile sheet has no numeric content.")

    mid_col = "Effective Stress" if "Effective Stress" in data.columns else data.columns[0]
    depth_col = "Maximum Displacement" if "Maximum Displacement" in data.columns else data.columns[6]

    mid_depths = _to_number_series(data[mid_col])
    out_depths = _to_number_series(data[depth_col])

    n = min(mid_depths.size, out_depths.size)
    if n == 0:
        raise ValueError("Unable to parse profile depths from Profile sheet.")

    mid_depths = mid_depths[:n]
    out_depths = out_depths[:n]

    thickness = np.zeros(n, dtype=float)
    cumulative = 0.0
    for i, depth_mid in enumerate(mid_depths):
        if i == 0:
            h = 2.0 * depth_mid
        else:
            h = 2.0 * (depth_mid - cumulative)
        if h <= 0:
            raise ValueError("Non-positive layer thickness detected while parsing profile.")
        thickness[i] = h
        cumulative += h

    predicted_depths = np.concatenate(([0.0], np.cumsum(thickness)))[:-1]
    if predicted_depths.size == out_depths.size and np.max(np.abs(predicted_depths - out_depths)) > 1e-3:
        raise ValueError("Profile depth and inferred thickness are inconsistent.")

    return out_depths, thickness


def _parse_profile_displacement_max(xl: pd.ExcelFile) -> Tuple[np.ndarray, np.ndarray]:
    if "Profile" not in xl.sheet_names:
        raise ValueError("Missing 'Profile' sheet.")

    profile = _parse_sheet_cached(xl, "Profile")
    data = profile.iloc[1:].reset_index(drop=True)
    if data.empty:
        raise ValueError("Profile sheet has no numeric content.")

    depth_col = "Maximum Displacement" if "Maximum Displacement" in data.columns else data.columns[6]
    depth_idx = data.columns.get_loc(depth_col)
    disp_idx = min(depth_idx + 1, len(data.columns) - 1)

    depths = _to_number_series(data.iloc[:, depth_idx])
    max_disp = _to_number_series(data.iloc[:, disp_idx])

    n = min(depths.size, max_disp.size)
    if n == 0:
        raise ValueError("Unable to parse Profile maximum displacement columns.")

    return depths[:n], max_disp[:n]


def _profile_bottom_max_disp(xl: pd.ExcelFile, target_depth: float) -> float:
    depths, max_disp = _parse_profile_displacement_max(xl)
    if depths.size == 0 or max_disp.size == 0:
        return float("nan")
    idx = int(np.argmin(np.abs(depths - float(target_depth))))
    return float(max_disp[idx])


def _read_input_motion(xl: pd.ExcelFile) -> Tuple[np.ndarray, np.ndarray]:
    if "Input Motion" not in xl.sheet_names:
        raise ValueError("Missing 'Input Motion' sheet.")

    motion = _parse_sheet_cached(xl, "Input Motion")
    if motion.shape[1] < 2:
        raise ValueError("Input Motion sheet has fewer than two columns.")

    subset = motion.iloc[:, :2].copy()
    subset.columns = ["Time (s)", "Acceleration (g)"]
    subset["Time (s)"] = pd.to_numeric(subset["Time (s)"], errors="coerce")
    subset["Acceleration (g)"] = pd.to_numeric(subset["Acceleration (g)"], errors="coerce")
    subset = subset.dropna(subset=["Time (s)", "Acceleration (g)"]).sort_values("Time (s)")

    if subset.empty:
        raise ValueError("Input Motion sheet does not contain numeric time/acceleration rows.")

    return (
        subset["Time (s)"].to_numpy(dtype=float),
        subset["Acceleration (g)"].to_numpy(dtype=float),
    )


def _compute_strain_bundle(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    layer_names = _ensure_common_layers(x_xl, y_xl)

    x_depths, thickness = parse_profile_thickness(x_xl)
    y_depths, _ = parse_profile_thickness(y_xl)
    if x_depths.size != y_depths.size or not np.allclose(x_depths, y_depths):
        raise ValueError("Profile depths are inconsistent between X and Y files.")

    n_layers = min(len(layer_names), x_depths.size, thickness.size)
    layer_names = layer_names[:n_layers]
    depths = x_depths[:n_layers]
    thickness = thickness[:n_layers]

    layer_payload: List[Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]] = []
    t_start = -np.inf
    t_end = np.inf
    dt_min = np.inf

    for layer_name in layer_names:
        tx, strain_x_pct = _read_layer_column(x_xl, layer_name, "Strain (%)")
        ty, strain_y_pct = _read_layer_column(y_xl, layer_name, "Strain (%)")
        gamma_x = strain_x_pct / 100.0
        gamma_y = strain_y_pct / 100.0

        t_start = max(t_start, tx[0], ty[0])
        t_end = min(t_end, tx[-1], ty[-1])
        if tx.size > 1:
            dt_min = min(dt_min, float(np.median(np.diff(tx))))
        if ty.size > 1:
            dt_min = min(dt_min, float(np.median(np.diff(ty))))

        layer_payload.append((tx, gamma_x, ty, gamma_y))

    if t_end <= t_start:
        raise ValueError("No overlapping time window found across layer strain records.")

    if not np.isfinite(dt_min) or dt_min <= 0:
        dt_min = 0.01

    sample_count = int(math.floor((t_end - t_start) / dt_min)) + 1
    time = t_start + np.arange(sample_count, dtype=float) * dt_min

    gamma_x_matrix = np.zeros((n_layers, sample_count), dtype=float)
    gamma_y_matrix = np.zeros((n_layers, sample_count), dtype=float)

    for i, (tx, gx, ty, gy) in enumerate(layer_payload):
        gamma_x_matrix[i, :] = np.interp(time, tx, gx)
        gamma_y_matrix[i, :] = np.interp(time, ty, gy)

    du_x = gamma_x_matrix * thickness[:, None]
    du_y = gamma_y_matrix * thickness[:, None]

    u_rel_base_x = np.flip(np.cumsum(np.flip(du_x, axis=0), axis=0), axis=0)
    u_rel_base_y = np.flip(np.cumsum(np.flip(du_y, axis=0), axis=0), axis=0)

    t_input_x, a_input_x = _read_input_motion(x_xl)
    t_input_y, a_input_y = _read_input_motion(y_xl)

    a_input_x_i = np.interp(time, t_input_x, a_input_x)
    a_input_y_i = np.interp(time, t_input_y, a_input_y)

    compare_cfg = _integration_compare_config(options)
    integration_meta: Dict[str, Any] = {
        "integrationPrimary": "cumtrapz",
        "integrationCompareEnabled": bool(compare_cfg["enabled"]),
    }

    input_x_dual = _acc_to_disp_dual(
        time,
        a_input_x_i,
        options=options,
    )
    input_y_dual = _acc_to_disp_dual(
        time,
        a_input_y_i,
        options=options,
    )
    u_input_proxy_x = input_x_dual["primary"]
    u_input_proxy_y = input_y_dual["primary"]
    u_input_proxy_x_alt = input_x_dual.get("alt")
    u_input_proxy_y_alt = input_y_dual.get("alt")

    for candidate in (input_x_dual.get("meta"), input_y_dual.get("meta")):
        if isinstance(candidate, dict) and candidate.get("integrationCompareEnabled", False):
            integration_meta.update(candidate)

    base_reference = _normalize_base_reference((options or {}).get("baseReference", DEFAULT_BASE_REFERENCE))
    u_base_ref_x_alt: np.ndarray | None = None
    u_base_ref_y_alt: np.ndarray | None = None
    if base_reference == "deepest_layer":
        deepest_layer = layer_names[-1]
        t_deep_x, a_deep_x = _read_layer_column(x_xl, deepest_layer, "Acceleration (g)")
        t_deep_y, a_deep_y = _read_layer_column(y_xl, deepest_layer, "Acceleration (g)")
        deep_x_dual = _acc_to_disp_dual(t_deep_x, a_deep_x, options=options)
        deep_y_dual = _acc_to_disp_dual(t_deep_y, a_deep_y, options=options)
        u_deep_x = deep_x_dual["primary"]
        u_deep_y = deep_y_dual["primary"]
        u_deep_x_alt = deep_x_dual.get("alt")
        u_deep_y_alt = deep_y_dual.get("alt")
        u_base_ref_x = np.interp(time, t_deep_x, u_deep_x)
        u_base_ref_y = np.interp(time, t_deep_y, u_deep_y)
        if u_deep_x_alt is not None:
            u_base_ref_x_alt = np.interp(time, t_deep_x, u_deep_x_alt)
        if u_deep_y_alt is not None:
            u_base_ref_y_alt = np.interp(time, t_deep_y, u_deep_y_alt)
        for candidate in (deep_x_dual.get("meta"), deep_y_dual.get("meta")):
            if isinstance(candidate, dict) and candidate.get("integrationCompareEnabled", False):
                integration_meta.update(candidate)
    else:
        u_base_ref_x = u_input_proxy_x
        u_base_ref_y = u_input_proxy_y
        if u_input_proxy_x_alt is not None:
            u_base_ref_x_alt = u_input_proxy_x_alt.copy()
        if u_input_proxy_y_alt is not None:
            u_base_ref_y_alt = u_input_proxy_y_alt.copy()

    u_rel_input_x = u_rel_base_x - u_input_proxy_x[None, :]
    u_rel_input_y = u_rel_base_y - u_input_proxy_y[None, :]
    u_tbdy_total_x = u_rel_base_x + u_base_ref_x[None, :]
    u_tbdy_total_y = u_rel_base_y + u_base_ref_y[None, :]

    u_rel_input_x_alt: np.ndarray | None = None
    u_rel_input_y_alt: np.ndarray | None = None
    u_tbdy_total_x_alt: np.ndarray | None = None
    u_tbdy_total_y_alt: np.ndarray | None = None
    if u_input_proxy_x_alt is not None and u_input_proxy_y_alt is not None:
        u_rel_input_x_alt = u_rel_base_x - u_input_proxy_x_alt[None, :]
        u_rel_input_y_alt = u_rel_base_y - u_input_proxy_y_alt[None, :]
    if u_base_ref_x_alt is not None and u_base_ref_y_alt is not None:
        u_tbdy_total_x_alt = u_rel_base_x + u_base_ref_x_alt[None, :]
        u_tbdy_total_y_alt = u_rel_base_y + u_base_ref_y_alt[None, :]

    x_base = np.max(np.abs(u_rel_base_x), axis=1)
    y_base = np.max(np.abs(u_rel_base_y), axis=1)
    x_base_pos = np.max(u_rel_base_x, axis=1)
    x_base_neg = np.min(u_rel_base_x, axis=1)
    y_base_pos = np.max(u_rel_base_y, axis=1)
    y_base_neg = np.min(u_rel_base_y, axis=1)
    total_base = np.max(np.sqrt(u_rel_base_x**2 + u_rel_base_y**2), axis=1)

    x_tbdy_total = np.max(np.abs(u_tbdy_total_x), axis=1)
    y_tbdy_total = np.max(np.abs(u_tbdy_total_y), axis=1)
    x_tbdy_total_pos = np.max(u_tbdy_total_x, axis=1)
    x_tbdy_total_neg = np.min(u_tbdy_total_x, axis=1)
    y_tbdy_total_pos = np.max(u_tbdy_total_y, axis=1)
    y_tbdy_total_neg = np.min(u_tbdy_total_y, axis=1)
    total_tbdy_total = np.max(np.sqrt(u_tbdy_total_x**2 + u_tbdy_total_y**2), axis=1)

    x_input = np.max(np.abs(u_rel_input_x), axis=1)
    y_input = np.max(np.abs(u_rel_input_y), axis=1)
    total_input = np.max(np.sqrt(u_rel_input_x**2 + u_rel_input_y**2), axis=1)

    x_profile_offset_total = np.full(n_layers, np.nan, dtype=float)
    y_profile_offset_total = np.full(n_layers, np.nan, dtype=float)
    total_profile_offset_total = np.full(n_layers, np.nan, dtype=float)
    try:
        x_profile_bottom = _profile_bottom_max_disp(x_xl, depths[-1])
        y_profile_bottom = _profile_bottom_max_disp(y_xl, depths[-1])
        x_profile_offset_total = (x_base - x_base[-1]) + x_profile_bottom
        y_profile_offset_total = (y_base - y_base[-1]) + y_profile_bottom
        total_profile_offset_total = np.sqrt(x_profile_offset_total**2 + y_profile_offset_total**2)
    except Exception:
        pass

    summary_data: Dict[str, Any] = {
        "Layer_Index": np.arange(1, n_layers + 1, dtype=int),
        "Depth_m": depths,
        "Thickness_m": thickness,
        "X_base_rel_max_m": x_base,
        "Y_base_rel_max_m": y_base,
        "X_base_rel_pos_max_m": x_base_pos,
        "X_base_rel_neg_min_m": x_base_neg,
        "Y_base_rel_pos_max_m": y_base_pos,
        "Y_base_rel_neg_min_m": y_base_neg,
        "Total_base_rel_max_m": total_base,
        "X_tbdy_total_max_m": x_tbdy_total,
        "Y_tbdy_total_max_m": y_tbdy_total,
        "X_tbdy_total_pos_max_m": x_tbdy_total_pos,
        "X_tbdy_total_neg_min_m": x_tbdy_total_neg,
        "Y_tbdy_total_pos_max_m": y_tbdy_total_pos,
        "Y_tbdy_total_neg_min_m": y_tbdy_total_neg,
        "Total_tbdy_total_max_m": total_tbdy_total,
        "X_profile_offset_total_est_m": x_profile_offset_total,
        "Y_profile_offset_total_est_m": y_profile_offset_total,
        "Total_profile_offset_total_est_m": total_profile_offset_total,
        "X_input_proxy_rel_max_m": x_input,
        "Y_input_proxy_rel_max_m": y_input,
        "Total_input_proxy_rel_max_m": total_input,
        "Base_Reference": [base_reference] * n_layers,
    }

    if u_tbdy_total_x_alt is not None and u_tbdy_total_y_alt is not None:
        x_tbdy_total_alt = np.max(np.abs(u_tbdy_total_x_alt), axis=1)
        y_tbdy_total_alt = np.max(np.abs(u_tbdy_total_y_alt), axis=1)
        x_tbdy_total_alt_pos = np.max(u_tbdy_total_x_alt, axis=1)
        x_tbdy_total_alt_neg = np.min(u_tbdy_total_x_alt, axis=1)
        y_tbdy_total_alt_pos = np.max(u_tbdy_total_y_alt, axis=1)
        y_tbdy_total_alt_neg = np.min(u_tbdy_total_y_alt, axis=1)
        total_tbdy_total_alt = np.max(np.sqrt(u_tbdy_total_x_alt**2 + u_tbdy_total_y_alt**2), axis=1)
        summary_data["X_tbdy_total_alt_max_m"] = x_tbdy_total_alt
        summary_data["Y_tbdy_total_alt_max_m"] = y_tbdy_total_alt
        summary_data["X_tbdy_total_alt_pos_max_m"] = x_tbdy_total_alt_pos
        summary_data["X_tbdy_total_alt_neg_min_m"] = x_tbdy_total_alt_neg
        summary_data["Y_tbdy_total_alt_pos_max_m"] = y_tbdy_total_alt_pos
        summary_data["Y_tbdy_total_alt_neg_min_m"] = y_tbdy_total_alt_neg
        summary_data["Total_tbdy_total_alt_max_m"] = total_tbdy_total_alt
        summary_data["Delta_Total_tbdy_alt_minus_primary_m"] = total_tbdy_total_alt - total_tbdy_total
        with np.errstate(divide="ignore", invalid="ignore"):
            summary_data["Ratio_Total_tbdy_alt_to_primary"] = np.where(
                total_tbdy_total != 0,
                total_tbdy_total_alt / total_tbdy_total,
                np.nan,
            )

    summary_df = pd.DataFrame(summary_data)

    return {
        "layer_names": layer_names,
        "depths": depths,
        "thickness": thickness,
        "time": time,
        "strain_x_pct_matrix": gamma_x_matrix * 100.0,
        "strain_y_pct_matrix": gamma_y_matrix * 100.0,
        "u_rel_base_x": u_rel_base_x,
        "u_rel_base_y": u_rel_base_y,
        "u_input_proxy_x": u_input_proxy_x,
        "u_input_proxy_y": u_input_proxy_y,
        "u_input_proxy_x_alt": u_input_proxy_x_alt,
        "u_input_proxy_y_alt": u_input_proxy_y_alt,
        "u_base_ref_x": u_base_ref_x,
        "u_base_ref_y": u_base_ref_y,
        "u_base_ref_x_alt": u_base_ref_x_alt,
        "u_base_ref_y_alt": u_base_ref_y_alt,
        "base_reference": base_reference,
        "u_rel_input_x": u_rel_input_x,
        "u_rel_input_y": u_rel_input_y,
        "u_rel_input_x_alt": u_rel_input_x_alt,
        "u_rel_input_y_alt": u_rel_input_y_alt,
        "u_tbdy_total_x": u_tbdy_total_x,
        "u_tbdy_total_y": u_tbdy_total_y,
        "u_tbdy_total_x_alt": u_tbdy_total_x_alt,
        "u_tbdy_total_y_alt": u_tbdy_total_y_alt,
        "integration_meta": integration_meta,
        "summary_df": summary_df,
    }


def compute_strain_relative(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    options: Mapping[str, Any] | None = None,
) -> pd.DataFrame:
    bundle = _compute_strain_bundle(x_xl, y_xl, options)
    return bundle["summary_df"].copy()


def _build_layer_time_df(
    time: np.ndarray,
    depths: np.ndarray,
    matrix: np.ndarray,
    value_suffix: str,
) -> pd.DataFrame:
    n_layers = min(int(matrix.shape[0]), int(depths.size))
    data: Dict[str, np.ndarray] = {"Time_s": time}
    for i in range(n_layers):
        depth = float(depths[i])
        data[f"L{i + 1:02d}_z{depth:.3f}m_{value_suffix}"] = matrix[i]
    return pd.DataFrame(data)


def _method2_sheet_name(axis_label: str) -> str:
    axis = axis_label.upper()
    if axis == "X":
        return "Method2_TBDY_X_Time"
    if axis == "Y":
        return "Method2_TBDY_Y_Time"
    return "Method2_TBDY_Time"


def _method2_alt_sheet_name(axis_label: str) -> str:
    axis = axis_label.upper()
    if axis == "X":
        return "Method2_TBDY_X_Time_ALT"
    if axis == "Y":
        return "Method2_TBDY_Y_Time_ALT"
    return "Method2_TBDY_Time_ALT"


def _method2_delta_sheet_name(axis_label: str) -> str:
    axis = axis_label.upper()
    if axis == "X":
        return "Method2_TBDY_X_Delta"
    if axis == "Y":
        return "Method2_TBDY_Y_Delta"
    return "Method2_TBDY_Delta"


def _build_method2_workbook(
    time_df: pd.DataFrame,
    meta_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame | None = None,
    alt_time_df: pd.DataFrame | None = None,
    delta_time_df: pd.DataFrame | None = None,
) -> bytes:
    axis_label = "UNKNOWN"
    if "Axis" in meta_df.columns and not meta_df.empty:
        axis_label = str(meta_df["Axis"].iloc[0]).upper()
    sheet_name = _method2_sheet_name(axis_label)
    alt_sheet = _method2_alt_sheet_name(axis_label)
    delta_sheet = _method2_delta_sheet_name(axis_label)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        time_df.to_excel(writer, sheet_name=sheet_name, index=False)
        if alt_time_df is not None and not alt_time_df.empty:
            alt_time_df.to_excel(writer, sheet_name=alt_sheet, index=False)
        if delta_time_df is not None and not delta_time_df.empty:
            delta_time_df.to_excel(writer, sheet_name=delta_sheet, index=False)
        meta_df.to_excel(writer, sheet_name="Method2_Metadata", index=False)

        if sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            _add_all_layers_chart(
                ws,
                ws.max_row - 1,
                ws.max_column - 1,
                f"Method-2 TBDY {axis_label}: All Layers Time-Displacement",
            )
        if alt_sheet in writer.sheets:
            ws_alt = writer.sheets[alt_sheet]
            _add_all_layers_chart(
                ws_alt,
                ws_alt.max_row - 1,
                ws_alt.max_column - 1,
                f"Method-2 TBDY {axis_label} ALT: FFT-Regularized",
            )
        if delta_sheet in writer.sheets:
            ws_delta = writer.sheets[delta_sheet]
            _add_all_layers_chart(
                ws_delta,
                ws_delta.max_row - 1,
                ws_delta.max_column - 1,
                f"Method-2 TBDY {axis_label} Delta: ALT - Primary",
            )

    return buffer.getvalue()


def _build_method2_extract_from_bundle(
    file_name: str,
    axis_label: str,
    strain_bundle: Mapping[str, Any],
    profile_sheet_view_df: pd.DataFrame | None,
    input_motion_max_abs: float,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    stem = Path(file_name).stem
    normalized_options = _normalize_options(options)
    processing_cfg = _processing_config(normalized_options)
    integration_meta = strain_bundle.get("integration_meta", {})

    time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_tbdy_total"],
        f"tbdy_total_{axis_label.lower()}_m",
    )
    max_abs = np.max(np.abs(strain_bundle["u_tbdy_total"]), axis=1)
    profile_df = pd.DataFrame(
        {
            "Depth_m": strain_bundle["depths"][: len(max_abs)],
            f"{stem}_maxabs_m": max_abs,
        }
    )
    profile_relative_df = pd.DataFrame(columns=["Depth_m"])
    if profile_sheet_view_df is not None and not profile_sheet_view_df.empty:
        profile_relative_df = pd.DataFrame(
            {
                "Depth_m": profile_sheet_view_df["Depth_m"].to_numpy(dtype=float),
                f"{stem}_profile_rel_m": profile_sheet_view_df["Profile_relative_m"].to_numpy(dtype=float),
            }
        )
    input_added_profile_df = _build_input_motion_added_profile_df(profile_sheet_view_df, input_motion_max_abs, stem)

    alt_time_df = None
    delta_time_df = None
    profile_alt_df = pd.DataFrame(columns=["Depth_m"])
    max_abs_alt = None
    if strain_bundle.get("u_tbdy_total_alt") is not None:
        u_tbdy_total_alt = strain_bundle["u_tbdy_total_alt"]
        alt_time_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            u_tbdy_total_alt,
            f"tbdy_total_{axis_label.lower()}_alt_m",
        )
        delta_time_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            u_tbdy_total_alt - strain_bundle["u_tbdy_total"],
            f"tbdy_total_{axis_label.lower()}_delta_m",
        )
        max_abs_alt = np.max(np.abs(u_tbdy_total_alt), axis=1)
        profile_alt_df = pd.DataFrame(
            {
                "Depth_m": strain_bundle["depths"][: len(max_abs_alt)],
                f"{stem}_maxabs_alt_m": max_abs_alt,
            }
        )

    meta_df = pd.DataFrame(
        [
            {
                "Source_File": file_name,
                "Axis": axis_label,
                "Layer_Count": int(strain_bundle["u_tbdy_total"].shape[0]),
                "Base_Reference": str(strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE)),
                "Processing_Mode": "legacy-highpass" if processing_cfg.get("legacy", True) else "custom",
                "Processing_Summary": _processing_summary_text(normalized_options),
                "Baseline_On": bool(processing_cfg.get("baseline_on", True)),
                "Baseline_Method": str(processing_cfg.get("baseline_method", "poly4")),
                "Filter_On": bool(
                    processing_cfg.get(
                        "filter_on",
                        processing_cfg.get("highpass_enabled", True),
                    )
                ),
                "Filter_Domain": str(processing_cfg.get("filter_domain", "frequency")),
                "Filter_Config": str(processing_cfg.get("filter_config", "highpass")),
                "Filter_Type": str(processing_cfg.get("filter_type", "fft")),
                "F_Low_Hz": float(processing_cfg.get("f_low_hz", np.nan)),
                "F_High_Hz": float(
                    processing_cfg.get(
                        "f_high_hz",
                        processing_cfg.get("highpass_cutoff_hz", np.nan),
                    )
                ),
                "Filter_Order": int(processing_cfg.get("filter_order", DEFAULT_FILTER_ORDER)),
                "Highpass_Cutoff_Hz": float(processing_cfg.get("highpass_cutoff_hz", np.nan)),
                "Highpass_Transition_Hz": float(processing_cfg.get("highpass_transition_hz", np.nan)),
                "Integration_Primary": str(integration_meta.get("integrationPrimary", "cumtrapz")),
                "Integration_Compare_On": bool(integration_meta.get("integrationCompareEnabled", False)),
                "Alt_Integration_Method": integration_meta.get("altIntegrationMethod"),
                "Alt_LowCut_Hz": integration_meta.get("altLowCutHz"),
            }
        ]
    )

    output_bytes = _build_method2_workbook(
        time_df,
        meta_df,
        profile_sheet_df=profile_sheet_view_df,
        alt_time_df=alt_time_df,
        delta_time_df=delta_time_df,
    )
    sheet_name = _method2_sheet_name(axis_label)
    time_sheets = [sheet_name]
    if alt_time_df is not None and not alt_time_df.empty:
        time_sheets.append(_method2_alt_sheet_name(axis_label))
    if delta_time_df is not None and not delta_time_df.empty:
        time_sheets.append(_method2_delta_sheet_name(axis_label))

    return {
        "skipped": False,
        "axis": axis_label,
        "profile_df": profile_df,
        "relative_profile_df": profile_relative_df,
        "input_added_profile_df": input_added_profile_df,
        "profile_alt_df": profile_alt_df,
        "result": {
            "pairKey": f"METHOD2|{stem}",
            "xFileName": file_name if axis_label == "X" else "",
            "yFileName": file_name if axis_label == "Y" else "",
            "outputFileName": f"output_method2_{stem}.xlsx",
            "outputBytes": output_bytes,
            "previewCharts": _method2_preview_charts(axis_label, strain_bundle),
            "metrics": {
                "mode": "method2_single",
                "axis": axis_label,
                "baseReference": str(strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE)),
                "integrationPrimary": str(integration_meta.get("integrationPrimary", "cumtrapz")),
                "integrationCompareEnabled": bool(integration_meta.get("integrationCompareEnabled", False)),
                "altIntegrationMethod": integration_meta.get("altIntegrationMethod"),
                "altLowCutHz": integration_meta.get("altLowCutHz"),
                "layerCount": int(strain_bundle["u_tbdy_total"].shape[0]),
                "timeSeriesSheets": len(time_sheets),
                "timeSheets": time_sheets,
                "surfaceTBDYTotal_m": float(max_abs[0]) if max_abs.size else float("nan"),
                "surfaceTBDYTotalAlt_m": (
                    float(max_abs_alt[0]) if max_abs_alt is not None and max_abs_alt.size else float("nan")
                ),
                "inputMotionDispMax_m": input_motion_max_abs if np.isfinite(input_motion_max_abs) else float("nan"),
            },
        },
    }


def _merge_profile_frames(profile_frames: Sequence[pd.DataFrame]) -> pd.DataFrame:
    clean_frames: List[pd.DataFrame] = []
    for frame in profile_frames:
        if frame is None or frame.empty:
            continue
        f = frame.copy()
        if "Depth_m" not in f.columns or f.shape[1] < 2:
            continue
        f["Depth_m"] = pd.to_numeric(f["Depth_m"], errors="coerce")
        f = f.dropna(subset=["Depth_m"])
        f = f.groupby("Depth_m", as_index=False).max()
        clean_frames.append(f)

    if not clean_frames:
        return pd.DataFrame(columns=["Depth_m"])

    merged = clean_frames[0]
    for frame in clean_frames[1:]:
        merged = merged.merge(frame, on="Depth_m", how="outer")
    return merged.sort_values("Depth_m").reset_index(drop=True)


def _build_method3_delta_df(primary_df: pd.DataFrame, alt_df: pd.DataFrame) -> pd.DataFrame:
    if primary_df is None or primary_df.empty or alt_df is None or alt_df.empty:
        return pd.DataFrame(columns=["Depth_m"])

    merged = primary_df.merge(alt_df, on="Depth_m", how="outer")
    out = pd.DataFrame({"Depth_m": merged["Depth_m"]})

    primary_map: Dict[str, str] = {}
    for col in primary_df.columns:
        if col == "Depth_m":
            continue
        stem = col.removesuffix("_maxabs_m")
        primary_map[stem] = col

    alt_map: Dict[str, str] = {}
    for col in alt_df.columns:
        if col == "Depth_m":
            continue
        stem = col.removesuffix("_maxabs_alt_m")
        alt_map[stem] = col

    common = sorted(set(primary_map.keys()) & set(alt_map.keys()))
    for stem in common:
        p_col = primary_map[stem]
        a_col = alt_map[stem]
        out[f"{stem}_delta_m"] = pd.to_numeric(merged[a_col], errors="coerce") - pd.to_numeric(
            merged[p_col], errors="coerce"
        )

    return out.sort_values("Depth_m").reset_index(drop=True)


def _build_method3_aggregate_workbook(
    profile_x_df: pd.DataFrame,
    profile_y_df: pd.DataFrame,
    profile_x_alt_df: pd.DataFrame | None = None,
    profile_y_alt_df: pd.DataFrame | None = None,
    profile_x_delta_df: pd.DataFrame | None = None,
    profile_y_delta_df: pd.DataFrame | None = None,
    relative_profile_x_df: pd.DataFrame | None = None,
    relative_profile_y_df: pd.DataFrame | None = None,
    input_added_profile_x_df: pd.DataFrame | None = None,
    input_added_profile_y_df: pd.DataFrame | None = None,
) -> bytes:
    x_df = profile_x_df if profile_x_df is not None and not profile_x_df.empty else pd.DataFrame(columns=["Depth_m"])
    y_df = profile_y_df if profile_y_df is not None and not profile_y_df.empty else pd.DataFrame(columns=["Depth_m"])
    x_alt_df = (
        profile_x_alt_df
        if profile_x_alt_df is not None and not profile_x_alt_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    y_alt_df = (
        profile_y_alt_df
        if profile_y_alt_df is not None and not profile_y_alt_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    x_delta_df = (
        profile_x_delta_df
        if profile_x_delta_df is not None and not profile_x_delta_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    y_delta_df = (
        profile_y_delta_df
        if profile_y_delta_df is not None and not profile_y_delta_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    rel_profile_x_df = (
        relative_profile_x_df
        if relative_profile_x_df is not None and not relative_profile_x_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    rel_profile_y_df = (
        relative_profile_y_df
        if relative_profile_y_df is not None and not relative_profile_y_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    input_added_x_df = (
        input_added_profile_x_df
        if input_added_profile_x_df is not None and not input_added_profile_x_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )
    input_added_y_df = (
        input_added_profile_y_df
        if input_added_profile_y_df is not None and not input_added_profile_y_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        x_df.to_excel(writer, sheet_name="Method3_Profile_X", index=False)
        y_df.to_excel(writer, sheet_name="Method3_Profile_Y", index=False)
        if not input_added_x_df.empty:
            input_added_x_df.to_excel(writer, sheet_name="Method3_ApproxTotal_X", index=False)
        if not input_added_y_df.empty:
            input_added_y_df.to_excel(writer, sheet_name="Method3_ApproxTotal_Y", index=False)
        if not x_alt_df.empty:
            x_alt_df.to_excel(writer, sheet_name="Method3_Profile_X_ALT", index=False)
        if not y_alt_df.empty:
            y_alt_df.to_excel(writer, sheet_name="Method3_Profile_Y_ALT", index=False)
        if not x_delta_df.empty:
            x_delta_df.to_excel(writer, sheet_name="Method3_Delta_X", index=False)
        if not y_delta_df.empty:
            y_delta_df.to_excel(writer, sheet_name="Method3_Delta_Y", index=False)

        if "Method3_Profile_X" in writer.sheets:
            ws_x = writer.sheets["Method3_Profile_X"]
            _add_depth_profile_chart(ws_x, len(x_df), depth_col=1, series_start_col=2)
        if "Method3_Profile_Y" in writer.sheets:
            ws_y = writer.sheets["Method3_Profile_Y"]
            _add_depth_profile_chart(ws_y, len(y_df), depth_col=1, series_start_col=2)
        if "Method3_ApproxTotal_X" in writer.sheets:
            ws_x_approx = writer.sheets["Method3_ApproxTotal_X"]
            _add_depth_profile_chart(
                ws_x_approx,
                len(input_added_x_df),
                depth_col=1,
                series_start_col=2,
                title="Method-3 X Approx Total (Ubase + Urel)",
            )
        if "Method3_ApproxTotal_Y" in writer.sheets:
            ws_y_approx = writer.sheets["Method3_ApproxTotal_Y"]
            _add_depth_profile_chart(
                ws_y_approx,
                len(input_added_y_df),
                depth_col=1,
                series_start_col=2,
                title="Method-3 Y Approx Total (Ubase + Urel)",
            )
        if "Method3_Profile_X_ALT" in writer.sheets:
            ws_x_alt = writer.sheets["Method3_Profile_X_ALT"]
            _add_depth_profile_chart(ws_x_alt, len(x_alt_df), depth_col=1, series_start_col=2)
        if "Method3_Profile_Y_ALT" in writer.sheets:
            ws_y_alt = writer.sheets["Method3_Profile_Y_ALT"]
            _add_depth_profile_chart(ws_y_alt, len(y_alt_df), depth_col=1, series_start_col=2)
        if "Method3_Delta_X" in writer.sheets:
            ws_x_delta = writer.sheets["Method3_Delta_X"]
            _add_depth_profile_chart(ws_x_delta, len(x_delta_df), depth_col=1, series_start_col=2)
        if "Method3_Delta_Y" in writer.sheets:
            ws_y_delta = writer.sheets["Method3_Delta_Y"]
            _add_depth_profile_chart(ws_y_delta, len(y_delta_df), depth_col=1, series_start_col=2)

    return buffer.getvalue()


def _compute_single_strain_bundle(
    xl: pd.ExcelFile,
    axis_label: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    layer_names = _list_layer_sheets(xl)
    if not layer_names:
        raise ValueError(f"Missing Layer sheets in {axis_label} file.")

    depths, thickness = parse_profile_thickness(xl)
    n_layers = min(len(layer_names), depths.size, thickness.size)
    layer_names = layer_names[:n_layers]
    depths = depths[:n_layers]
    thickness = thickness[:n_layers]

    payload: List[Tuple[np.ndarray, np.ndarray]] = []
    t_start = -np.inf
    t_end = np.inf
    dt_min = np.inf

    for layer_name in layer_names:
        t, strain_pct = _read_layer_column(xl, layer_name, "Strain (%)")
        gamma = strain_pct / 100.0
        payload.append((t, gamma))

        t_start = max(t_start, t[0])
        t_end = min(t_end, t[-1])
        if t.size > 1:
            dt_min = min(dt_min, float(np.median(np.diff(t))))

    if t_end <= t_start:
        raise ValueError(f"No overlapping time window found across strain records in {axis_label} file.")

    if not np.isfinite(dt_min) or dt_min <= 0:
        dt_min = 0.01

    sample_count = int(math.floor((t_end - t_start) / dt_min)) + 1
    time = t_start + np.arange(sample_count, dtype=float) * dt_min

    gamma_matrix = np.zeros((n_layers, sample_count), dtype=float)
    for i, (t, gamma) in enumerate(payload):
        gamma_matrix[i, :] = np.interp(time, t, gamma)

    du = gamma_matrix * thickness[:, None]
    u_rel_base = np.flip(np.cumsum(np.flip(du, axis=0), axis=0), axis=0)

    t_input, a_input = _read_input_motion(xl)
    a_input_i = np.interp(time, t_input, a_input)
    compare_cfg = _integration_compare_config(options)
    integration_meta: Dict[str, Any] = {
        "integrationPrimary": "cumtrapz",
        "integrationCompareEnabled": bool(compare_cfg["enabled"]),
    }

    input_dual = _acc_to_disp_dual(
        time,
        a_input_i,
        options=options,
    )
    u_input_proxy = input_dual["primary"]
    u_input_proxy_alt = input_dual.get("alt")
    if isinstance(input_dual.get("meta"), dict):
        integration_meta.update(input_dual["meta"])

    base_reference = _normalize_base_reference((options or {}).get("baseReference", DEFAULT_BASE_REFERENCE))
    u_base_ref_alt: np.ndarray | None = None
    if base_reference == "deepest_layer":
        deepest_layer = layer_names[-1]
        t_deep, a_deep = _read_layer_column(xl, deepest_layer, "Acceleration (g)")
        deep_dual = _acc_to_disp_dual(t_deep, a_deep, options=options)
        u_deep = deep_dual["primary"]
        u_deep_alt = deep_dual.get("alt")
        u_base_ref = np.interp(time, t_deep, u_deep)
        if u_deep_alt is not None:
            u_base_ref_alt = np.interp(time, t_deep, u_deep_alt)
        if isinstance(deep_dual.get("meta"), dict):
            integration_meta.update(deep_dual["meta"])
    else:
        u_base_ref = u_input_proxy
        if u_input_proxy_alt is not None:
            u_base_ref_alt = u_input_proxy_alt.copy()

    u_rel_input = u_rel_base - u_input_proxy[None, :]
    u_tbdy_total = u_rel_base + u_base_ref[None, :]
    u_rel_input_alt: np.ndarray | None = None
    u_tbdy_total_alt: np.ndarray | None = None
    if u_input_proxy_alt is not None:
        u_rel_input_alt = u_rel_base - u_input_proxy_alt[None, :]
    if u_base_ref_alt is not None:
        u_tbdy_total_alt = u_rel_base + u_base_ref_alt[None, :]

    base_rel_max = np.max(np.abs(u_rel_base), axis=1)
    tbdy_total_max = np.max(np.abs(u_tbdy_total), axis=1)
    input_proxy_rel_max = np.max(np.abs(u_rel_input), axis=1)

    profile_offset_total_est = np.full(n_layers, np.nan, dtype=float)
    try:
        profile_bottom = _profile_bottom_max_disp(xl, depths[-1])
        profile_offset_total_est = (base_rel_max - base_rel_max[-1]) + profile_bottom
    except Exception:
        pass

    summary_data: Dict[str, Any] = {
        "Layer_Index": np.arange(1, n_layers + 1, dtype=int),
        "Depth_m": depths,
        "Thickness_m": thickness,
        "Axis": axis_label,
        "Base_rel_max_m": base_rel_max,
        "TBDY_total_max_m": tbdy_total_max,
        "Profile_offset_total_est_m": profile_offset_total_est,
        "Input_proxy_rel_max_m": input_proxy_rel_max,
        "Base_Reference": [base_reference] * n_layers,
    }
    if u_tbdy_total_alt is not None:
        tbdy_total_alt_max = np.max(np.abs(u_tbdy_total_alt), axis=1)
        summary_data["TBDY_total_alt_max_m"] = tbdy_total_alt_max
        summary_data["Delta_TBDY_alt_minus_primary_m"] = tbdy_total_alt_max - tbdy_total_max
        with np.errstate(divide="ignore", invalid="ignore"):
            summary_data["Ratio_TBDY_alt_to_primary"] = np.where(
                tbdy_total_max != 0,
                tbdy_total_alt_max / tbdy_total_max,
                np.nan,
            )

    summary_df = pd.DataFrame(summary_data)

    return {
        "layer_names": layer_names,
        "depths": depths,
        "thickness": thickness,
        "time": time,
        "strain_pct_matrix": gamma_matrix * 100.0,
        "u_input_proxy": u_input_proxy,
        "u_input_proxy_alt": u_input_proxy_alt,
        "u_rel_base": u_rel_base,
        "u_rel_input": u_rel_input,
        "u_rel_input_alt": u_rel_input_alt,
        "u_base_ref": u_base_ref,
        "u_base_ref_alt": u_base_ref_alt,
        "base_reference": base_reference,
        "u_tbdy_total": u_tbdy_total,
        "u_tbdy_total_alt": u_tbdy_total_alt,
        "integration_meta": integration_meta,
        "summary_df": summary_df,
    }


def _compute_single_direction_disp_bundle(
    xl: pd.ExcelFile,
    axis_label: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    layer_names = _list_layer_sheets(xl)
    if not layer_names:
        raise ValueError(f"Missing Layer sheets in {axis_label} file.")

    depths, _ = parse_profile_thickness(xl)
    n_layers = min(len(layer_names), depths.size)
    layer_names = layer_names[:n_layers]
    depths = depths[:n_layers]

    payload: List[Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray | None]] = []
    t_start = -np.inf
    t_end = np.inf
    dt_min = np.inf
    integration_meta: Dict[str, Any] = {
        "integrationPrimary": "cumtrapz",
        "integrationCompareEnabled": bool(_integration_compare_config(options).get("enabled", False)),
    }
    for layer_name in layer_names:
        t, a = _read_layer_column(xl, layer_name, "Acceleration (g)")
        dual = _acc_to_disp_dual(
            t,
            a,
            options=options,
        )
        v = np.asarray(dual.get("primary_velocity", np.array([])), dtype=float)
        d = dual["primary"]
        d_alt = dual.get("alt")
        payload.append((t, a, v, d, d_alt))
        if isinstance(dual.get("meta"), dict) and dual["meta"].get("integrationCompareEnabled", False):
            integration_meta.update(dual["meta"])

        t_start = max(t_start, t[0])
        t_end = min(t_end, t[-1])
        if t.size > 1:
            dt_min = min(dt_min, float(np.median(np.diff(t))))

    if t_end <= t_start:
        common_time = payload[0][0]
    else:
        if not np.isfinite(dt_min) or dt_min <= 0:
            dt_min = 0.01
        sample_count = int(math.floor((t_end - t_start) / dt_min)) + 1
        common_time = t_start + np.arange(sample_count, dtype=float) * dt_min

    acc_matrix = np.zeros((n_layers, common_time.size), dtype=float)
    vel_matrix = np.zeros((n_layers, common_time.size), dtype=float)
    disp_matrix = np.zeros((n_layers, common_time.size), dtype=float)
    disp_matrix_alt: np.ndarray | None = None
    if any(item[4] is not None for item in payload):
        disp_matrix_alt = np.zeros((n_layers, common_time.size), dtype=float)

    for i, (t, a, v, d, d_alt) in enumerate(payload):
        acc_matrix[i, :] = np.interp(common_time, t, a)
        vel_matrix[i, :] = np.interp(common_time, t, v)
        disp_matrix[i, :] = np.interp(common_time, t, d)
        if disp_matrix_alt is not None and d_alt is not None:
            disp_matrix_alt[i, :] = np.interp(common_time, t, d_alt)

    table_df = _build_layer_time_df(common_time, depths, disp_matrix, "disp_m")
    table_alt_df = None
    delta_df = None
    if disp_matrix_alt is not None:
        table_alt_df = _build_layer_time_df(common_time, depths, disp_matrix_alt, "disp_alt_m")
        delta_df = _build_layer_time_df(common_time, depths, disp_matrix_alt - disp_matrix, "disp_delta_m")

    return {
        "axis": axis_label,
        "layer_names": layer_names,
        "depths": depths,
        "time": common_time,
        "acc_matrix": acc_matrix,
        "vel_matrix": vel_matrix,
        "disp_matrix": disp_matrix,
        "disp_matrix_alt": disp_matrix_alt,
        "table_df": table_df,
        "table_alt_df": table_alt_df,
        "delta_df": delta_df,
        "integration_meta": integration_meta,
    }


def _build_resultant_time_df(
    x_bundle: Mapping[str, Any],
    y_bundle: Mapping[str, Any],
    *,
    matrix_key: str = "disp_matrix",
    value_suffix: str = "resultant_m",
) -> pd.DataFrame:
    n_layers = min(
        int(x_bundle[matrix_key].shape[0]),
        int(y_bundle[matrix_key].shape[0]),
        int(x_bundle["depths"].size),
        int(y_bundle["depths"].size),
    )
    if n_layers <= 0:
        return pd.DataFrame({"Time_s": []})

    t = _get_common_time_for_layer(x_bundle["time"], y_bundle["time"])
    data: Dict[str, np.ndarray] = {"Time_s": t}

    for i in range(n_layers):
        x_i = np.interp(t, x_bundle["time"], x_bundle[matrix_key][i])
        y_i = np.interp(t, y_bundle["time"], y_bundle[matrix_key][i])
        total = np.sqrt(x_i**2 + y_i**2)
        depth = float(x_bundle["depths"][i])
        data[f"L{i + 1:02d}_z{depth:.3f}m_{value_suffix}"] = total

    return pd.DataFrame(data)


def _compute_legacy_bundle(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    layer_names = _ensure_common_layers(x_xl, y_xl)

    depth_x, profile_x = _parse_profile_displacement_max(x_xl)
    depth_y, profile_y = _parse_profile_displacement_max(y_xl)

    if depth_x.size != depth_y.size or not np.allclose(depth_x, depth_y):
        raise ValueError("Profile maximum displacement depths mismatch between X and Y files.")

    n_layers = min(len(layer_names), depth_x.size, profile_x.size, profile_y.size)
    layer_names = layer_names[:n_layers]
    depths = depth_x[:n_layers]
    profile_x = np.abs(profile_x[:n_layers])
    profile_y = np.abs(profile_y[:n_layers])

    time_hist_x = np.zeros(n_layers, dtype=float)
    time_hist_y = np.zeros(n_layers, dtype=float)
    time_hist_resultant = np.zeros(n_layers, dtype=float)
    time_hist_x_pos = np.zeros(n_layers, dtype=float)
    time_hist_x_neg = np.zeros(n_layers, dtype=float)
    time_hist_y_pos = np.zeros(n_layers, dtype=float)
    time_hist_y_neg = np.zeros(n_layers, dtype=float)
    time_hist_x_alt = np.full(n_layers, np.nan, dtype=float)
    time_hist_y_alt = np.full(n_layers, np.nan, dtype=float)
    time_hist_resultant_alt = np.full(n_layers, np.nan, dtype=float)
    time_hist_x_alt_pos = np.full(n_layers, np.nan, dtype=float)
    time_hist_x_alt_neg = np.full(n_layers, np.nan, dtype=float)
    time_hist_y_alt_pos = np.full(n_layers, np.nan, dtype=float)
    time_hist_y_alt_neg = np.full(n_layers, np.nan, dtype=float)
    integration_meta: Dict[str, Any] = {
        "integrationPrimary": "cumtrapz",
        "integrationCompareEnabled": bool(_integration_compare_config(options).get("enabled", False)),
    }

    for i, layer_name in enumerate(layer_names):
        tx, ax = _read_layer_column(x_xl, layer_name, "Acceleration (g)")
        ty, ay = _read_layer_column(y_xl, layer_name, "Acceleration (g)")
        t, ax_i, ay_i = _align_two_series(tx, ax, ty, ay)

        dx_dual = _acc_to_disp_dual(
            t,
            ax_i,
            options=options,
        )
        dy_dual = _acc_to_disp_dual(
            t,
            ay_i,
            options=options,
        )
        dx = dx_dual["primary"]
        dy = dy_dual["primary"]
        dx_alt = dx_dual.get("alt")
        dy_alt = dy_dual.get("alt")
        total = np.sqrt(dx**2 + dy**2)

        time_hist_x[i] = float(np.max(np.abs(dx)))
        time_hist_y[i] = float(np.max(np.abs(dy)))
        time_hist_resultant[i] = float(np.max(total))
        time_hist_x_pos[i] = float(np.max(dx))
        time_hist_x_neg[i] = float(np.min(dx))
        time_hist_y_pos[i] = float(np.max(dy))
        time_hist_y_neg[i] = float(np.min(dy))

        if dx_alt is not None:
            time_hist_x_alt[i] = float(np.max(np.abs(dx_alt)))
            time_hist_x_alt_pos[i] = float(np.max(dx_alt))
            time_hist_x_alt_neg[i] = float(np.min(dx_alt))
        if dy_alt is not None:
            time_hist_y_alt[i] = float(np.max(np.abs(dy_alt)))
            time_hist_y_alt_pos[i] = float(np.max(dy_alt))
            time_hist_y_alt_neg[i] = float(np.min(dy_alt))
        if dx_alt is not None and dy_alt is not None:
            total_alt = np.sqrt(dx_alt**2 + dy_alt**2)
            time_hist_resultant_alt[i] = float(np.max(total_alt))

        for candidate in (dx_dual.get("meta"), dy_dual.get("meta")):
            if isinstance(candidate, dict) and candidate.get("integrationCompareEnabled", False):
                integration_meta.update(candidate)

    profile_rss = np.sqrt(profile_x**2 + profile_y**2)

    summary_data: Dict[str, Any] = {
        "Layer_Index": np.arange(1, n_layers + 1, dtype=int),
        "Depth_m": depths,
        "Profile_X_max_m": profile_x,
        "Profile_Y_max_m": profile_y,
        "Profile_RSS_total_m": profile_rss,
        "Direction_X_maxabs_m": time_hist_x,
        "Direction_Y_maxabs_m": time_hist_y,
        "Direction_X_pos_max_m": time_hist_x_pos,
        "Direction_X_neg_min_m": time_hist_x_neg,
        "Direction_Y_pos_max_m": time_hist_y_pos,
        "Direction_Y_neg_min_m": time_hist_y_neg,
        "TimeHist_X_maxabs_m": time_hist_x,
        "TimeHist_Y_maxabs_m": time_hist_y,
        "TimeHist_X_pos_max_m": time_hist_x_pos,
        "TimeHist_X_neg_min_m": time_hist_x_neg,
        "TimeHist_Y_pos_max_m": time_hist_y_pos,
        "TimeHist_Y_neg_min_m": time_hist_y_neg,
        "TimeHist_Resultant_total_m": time_hist_resultant,
    }
    if np.any(np.isfinite(time_hist_x_alt)) or np.any(np.isfinite(time_hist_y_alt)):
        summary_data["Direction_X_maxabs_alt_m"] = time_hist_x_alt
        summary_data["Direction_Y_maxabs_alt_m"] = time_hist_y_alt
        summary_data["Direction_X_alt_pos_max_m"] = time_hist_x_alt_pos
        summary_data["Direction_X_alt_neg_min_m"] = time_hist_x_alt_neg
        summary_data["Direction_Y_alt_pos_max_m"] = time_hist_y_alt_pos
        summary_data["Direction_Y_alt_neg_min_m"] = time_hist_y_alt_neg
        summary_data["TimeHist_X_maxabs_alt_m"] = time_hist_x_alt
        summary_data["TimeHist_Y_maxabs_alt_m"] = time_hist_y_alt
        summary_data["TimeHist_X_alt_pos_max_m"] = time_hist_x_alt_pos
        summary_data["TimeHist_X_alt_neg_min_m"] = time_hist_x_alt_neg
        summary_data["TimeHist_Y_alt_pos_max_m"] = time_hist_y_alt_pos
        summary_data["TimeHist_Y_alt_neg_min_m"] = time_hist_y_alt_neg
    if np.any(np.isfinite(time_hist_resultant_alt)):
        summary_data["TimeHist_Resultant_alt_total_m"] = time_hist_resultant_alt
        summary_data["Delta_TimeHist_Resultant_alt_minus_primary_m"] = (
            time_hist_resultant_alt - time_hist_resultant
        )

    summary_df = pd.DataFrame(summary_data)

    return {
        "layer_names": layer_names,
        "depths": depths,
        "integration_meta": integration_meta,
        "summary_df": summary_df,
    }


def compute_legacy_methods(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    options: Mapping[str, Any] | None = None,
) -> pd.DataFrame:
    bundle = _compute_legacy_bundle(x_xl, y_xl, options)
    return bundle["summary_df"].copy()


def _build_comparison_df(strain_df: pd.DataFrame, legacy_df: pd.DataFrame) -> pd.DataFrame:
    strain_cols = [
        "Layer_Index",
        "Depth_m",
        "X_base_rel_max_m",
        "Y_base_rel_max_m",
        "X_tbdy_total_max_m",
        "Y_tbdy_total_max_m",
        "Total_base_rel_max_m",
        "Total_tbdy_total_max_m",
        "Total_input_proxy_rel_max_m",
    ]
    for optional_col in (
        "X_base_rel_pos_max_m",
        "X_base_rel_neg_min_m",
        "Y_base_rel_pos_max_m",
        "Y_base_rel_neg_min_m",
        "X_tbdy_total_pos_max_m",
        "X_tbdy_total_neg_min_m",
        "Y_tbdy_total_pos_max_m",
        "Y_tbdy_total_neg_min_m",
        "X_profile_offset_total_est_m",
        "Y_profile_offset_total_est_m",
        "Total_profile_offset_total_est_m",
        "X_tbdy_total_alt_max_m",
        "Y_tbdy_total_alt_max_m",
        "X_tbdy_total_alt_pos_max_m",
        "X_tbdy_total_alt_neg_min_m",
        "Y_tbdy_total_alt_pos_max_m",
        "Y_tbdy_total_alt_neg_min_m",
        "Total_tbdy_total_alt_max_m",
        "Delta_Total_tbdy_alt_minus_primary_m",
        "Ratio_Total_tbdy_alt_to_primary",
    ):
        if optional_col in strain_df.columns:
            strain_cols.append(optional_col)

    legacy_cols = [
        "Layer_Index",
        "Depth_m",
        "Profile_X_max_m",
        "Profile_Y_max_m",
        "Profile_RSS_total_m",
        "TimeHist_Resultant_total_m",
    ]
    for optional_col in (
        "Direction_X_maxabs_m",
        "Direction_Y_maxabs_m",
        "Direction_X_pos_max_m",
        "Direction_X_neg_min_m",
        "Direction_Y_pos_max_m",
        "Direction_Y_neg_min_m",
        "TimeHist_X_maxabs_m",
        "TimeHist_Y_maxabs_m",
        "TimeHist_X_pos_max_m",
        "TimeHist_X_neg_min_m",
        "TimeHist_Y_pos_max_m",
        "TimeHist_Y_neg_min_m",
        "Direction_X_maxabs_alt_m",
        "Direction_Y_maxabs_alt_m",
        "Direction_X_alt_pos_max_m",
        "Direction_X_alt_neg_min_m",
        "Direction_Y_alt_pos_max_m",
        "Direction_Y_alt_neg_min_m",
        "TimeHist_X_maxabs_alt_m",
        "TimeHist_Y_maxabs_alt_m",
        "TimeHist_X_alt_pos_max_m",
        "TimeHist_X_alt_neg_min_m",
        "TimeHist_Y_alt_pos_max_m",
        "TimeHist_Y_alt_neg_min_m",
        "TimeHist_Resultant_alt_total_m",
        "Delta_TimeHist_Resultant_alt_minus_primary_m",
    ):
        if optional_col in legacy_df.columns:
            legacy_cols.append(optional_col)

    strain_view = strain_df[strain_cols].copy()
    legacy_view = legacy_df[legacy_cols].copy().rename(columns={"Depth_m": "Depth_m_legacy"})
    merged = strain_view.merge(
        legacy_view,
        on=["Layer_Index"],
        how="left",
    )
    if "Depth_m_legacy" in merged.columns:
        merged["Depth_m"] = (
            pd.to_numeric(merged.get("Depth_m"), errors="coerce")
            .fillna(pd.to_numeric(merged.get("Depth_m_legacy"), errors="coerce"))
        )
        merged = merged.drop(columns=["Depth_m_legacy"])

    merged = merged.sort_values(["Layer_Index", "Depth_m"], kind="stable").reset_index(drop=True)
    if merged.empty:
        return merged

    profile_x_bottom = float(merged["Profile_X_max_m"].iloc[-1])
    profile_y_bottom = float(merged["Profile_Y_max_m"].iloc[-1])
    profile_rss_bottom = float(merged["Profile_RSS_total_m"].iloc[-1])

    merged["Profile_X_minus_bottom_m"] = merged["Profile_X_max_m"] - profile_x_bottom
    merged["Profile_Y_minus_bottom_m"] = merged["Profile_Y_max_m"] - profile_y_bottom
    merged["Profile_RSS_minus_bottom_m"] = merged["Profile_RSS_total_m"] - profile_rss_bottom
    merged["Delta_Xbase_vs_ProfileXminusbottom_m"] = (
        merged["X_base_rel_max_m"] - merged["Profile_X_minus_bottom_m"]
    )
    merged["Delta_Ybase_vs_ProfileYminusbottom_m"] = (
        merged["Y_base_rel_max_m"] - merged["Profile_Y_minus_bottom_m"]
    )

    merged["Delta_base_vs_profile_m"] = merged["Total_base_rel_max_m"] - merged["Profile_RSS_total_m"]
    merged["Delta_tbdy_vs_profile_m"] = merged["Total_tbdy_total_max_m"] - merged["Profile_RSS_total_m"]
    merged["Delta_base_vs_timehist_m"] = merged["Total_base_rel_max_m"] - merged["TimeHist_Resultant_total_m"]
    merged["Delta_inputproxy_vs_profile_m"] = (
        merged["Total_input_proxy_rel_max_m"] - merged["Profile_RSS_total_m"]
    )
    if "Total_tbdy_total_alt_max_m" in merged.columns:
        merged["Delta_tbdy_alt_vs_profile_m"] = (
            merged["Total_tbdy_total_alt_max_m"] - merged["Profile_RSS_total_m"]
        )
        merged["Delta_tbdy_alt_vs_primary_m"] = (
            merged["Total_tbdy_total_alt_max_m"] - merged["Total_tbdy_total_max_m"]
        )
    if "Total_profile_offset_total_est_m" in merged.columns:
        merged["Delta_profileoffset_vs_profile_m"] = (
            merged["Total_profile_offset_total_est_m"] - merged["Profile_RSS_total_m"]
        )

    with np.errstate(divide="ignore", invalid="ignore"):
        merged["Ratio_base_to_profile"] = np.where(
            merged["Profile_RSS_total_m"] != 0,
            merged["Total_base_rel_max_m"] / merged["Profile_RSS_total_m"],
            np.nan,
        )
        merged["Ratio_tbdy_to_profile"] = np.where(
            merged["Profile_RSS_total_m"] != 0,
            merged["Total_tbdy_total_max_m"] / merged["Profile_RSS_total_m"],
            np.nan,
        )
        if "Total_tbdy_total_alt_max_m" in merged.columns:
            merged["Ratio_tbdy_alt_to_profile"] = np.where(
                merged["Profile_RSS_total_m"] != 0,
                merged["Total_tbdy_total_alt_max_m"] / merged["Profile_RSS_total_m"],
                np.nan,
            )
        merged["Ratio_base_to_timehist"] = np.where(
            merged["TimeHist_Resultant_total_m"] != 0,
            merged["Total_base_rel_max_m"] / merged["TimeHist_Resultant_total_m"],
            np.nan,
        )
        if "Total_profile_offset_total_est_m" in merged.columns:
            merged["Ratio_profileoffset_to_profile"] = np.where(
                merged["Profile_RSS_total_m"] != 0,
                merged["Total_profile_offset_total_est_m"] / merged["Profile_RSS_total_m"],
                np.nan,
            )

    return merged


def _build_depth_profiles_df(
    comparison_df: pd.DataFrame,
    *,
    include_resultants: bool = True,
) -> pd.DataFrame:
    cols = [
        "Layer_Index",
        "Depth_m",
    ]

    directional_cols = [
        "Direction_X_pos_max_m",
        "Direction_X_neg_min_m",
        "Direction_Y_pos_max_m",
        "Direction_Y_neg_min_m",
        "X_base_rel_pos_max_m",
        "X_base_rel_neg_min_m",
        "Y_base_rel_pos_max_m",
        "Y_base_rel_neg_min_m",
        "Direction_X_alt_pos_max_m",
        "Direction_X_alt_neg_min_m",
        "Direction_Y_alt_pos_max_m",
        "Direction_Y_alt_neg_min_m",
        "TimeHist_X_pos_max_m",
        "TimeHist_X_neg_min_m",
        "TimeHist_Y_pos_max_m",
        "TimeHist_Y_neg_min_m",
        "X_tbdy_total_pos_max_m",
        "X_tbdy_total_neg_min_m",
        "Y_tbdy_total_pos_max_m",
        "Y_tbdy_total_neg_min_m",
        "TimeHist_X_alt_pos_max_m",
        "TimeHist_X_alt_neg_min_m",
        "TimeHist_Y_alt_pos_max_m",
        "TimeHist_Y_alt_neg_min_m",
        "X_tbdy_total_alt_pos_max_m",
        "X_tbdy_total_alt_neg_min_m",
        "Y_tbdy_total_alt_pos_max_m",
        "Y_tbdy_total_alt_neg_min_m",
    ]
    for col in directional_cols:
        if col in comparison_df.columns:
            cols.append(col)

    # Backward-compatible directional maxabs series.
    for col in (
        "Direction_X_maxabs_m",
        "Direction_Y_maxabs_m",
        "X_base_rel_max_m",
        "Y_base_rel_max_m",
        "TimeHist_X_maxabs_m",
        "TimeHist_Y_maxabs_m",
        "X_tbdy_total_max_m",
        "Y_tbdy_total_max_m",
        "Direction_X_maxabs_alt_m",
        "Direction_Y_maxabs_alt_m",
        "TimeHist_X_maxabs_alt_m",
        "TimeHist_Y_maxabs_alt_m",
        "X_tbdy_total_alt_max_m",
        "Y_tbdy_total_alt_max_m",
    ):
        if col in comparison_df.columns and col not in cols:
            cols.append(col)

    if include_resultants:
        for col in (
            "Total_base_rel_max_m",
            "Total_tbdy_total_max_m",
            "Total_input_proxy_rel_max_m",
            "Profile_RSS_total_m",
            "TimeHist_Resultant_total_m",
            "Total_tbdy_total_alt_max_m",
            "TimeHist_Resultant_alt_total_m",
            "Total_profile_offset_total_est_m",
        ):
            if col in comparison_df.columns and col not in cols:
                cols.append(col)

    return comparison_df[cols].copy()


def _build_base_corrected_profiles_df(comparison_df: pd.DataFrame) -> pd.DataFrame:
    return comparison_df[
        [
            "Layer_Index",
            "Depth_m",
            "X_base_rel_max_m",
            "Profile_X_minus_bottom_m",
            "Delta_Xbase_vs_ProfileXminusbottom_m",
            "Y_base_rel_max_m",
            "Profile_Y_minus_bottom_m",
            "Delta_Ybase_vs_ProfileYminusbottom_m",
            "Total_base_rel_max_m",
            "Profile_RSS_minus_bottom_m",
        ]
    ].copy()


def _build_pair_profile_sheet_df(x_xl: pd.ExcelFile, y_xl: pd.ExcelFile) -> pd.DataFrame:
    depth_x, profile_x = _parse_profile_displacement_max(x_xl)
    depth_y, profile_y = _parse_profile_displacement_max(y_xl)
    n = min(depth_x.size, depth_y.size, profile_x.size, profile_y.size)
    if n == 0:
        return pd.DataFrame(
            columns=[
                "Layer_Index",
                "Depth_m",
                "Profile_X_raw_max_m",
                "Profile_Y_raw_max_m",
                "Profile_RSS_raw_max_m",
                "Profile_X_relative_m",
                "Profile_Y_relative_m",
                "Profile_RSS_relative_m",
            ]
        )

    depths = depth_x[:n]
    x_raw = np.abs(profile_x[:n])
    y_raw = np.abs(profile_y[:n])
    rss_raw = np.sqrt(x_raw**2 + y_raw**2)
    x_rel = x_raw - x_raw[-1]
    y_rel = y_raw - y_raw[-1]
    rss_rel = rss_raw - rss_raw[-1]
    return pd.DataFrame(
        {
            "Layer_Index": np.arange(1, n + 1, dtype=int),
            "Depth_m": depths,
            "Profile_X_raw_max_m": x_raw,
            "Profile_Y_raw_max_m": y_raw,
            "Profile_RSS_raw_max_m": rss_raw,
            "Profile_X_relative_m": x_rel,
            "Profile_Y_relative_m": y_rel,
            "Profile_RSS_relative_m": rss_rel,
        }
    )


def _build_single_profile_sheet_df(xl: pd.ExcelFile) -> pd.DataFrame:
    profile_depths, profile_max = _parse_profile_displacement_max(xl)
    n = min(profile_depths.size, profile_max.size)
    if n == 0:
        return pd.DataFrame(columns=["Layer_Index", "Depth_m", "Profile_raw_max_m", "Profile_relative_m"])

    depths = profile_depths[:n]
    raw = np.abs(profile_max[:n])
    rel = raw - raw[-1]
    return pd.DataFrame(
        {
            "Layer_Index": np.arange(1, n + 1, dtype=int),
            "Depth_m": depths,
            "Profile_raw_max_m": raw,
            "Profile_relative_m": rel,
        }
    )


def _build_input_motion_added_profile_df(
    profile_sheet_df: pd.DataFrame,
    input_motion_max_abs: float,
    stem: str,
) -> pd.DataFrame:
    if (
        profile_sheet_df is None
        or profile_sheet_df.empty
        or "Depth_m" not in profile_sheet_df.columns
        or "Profile_relative_m" not in profile_sheet_df.columns
        or not np.isfinite(input_motion_max_abs)
    ):
        return pd.DataFrame(columns=["Depth_m"])

    depths = pd.to_numeric(profile_sheet_df["Depth_m"], errors="coerce").to_numpy(dtype=float)
    rel = pd.to_numeric(profile_sheet_df["Profile_relative_m"], errors="coerce").to_numpy(dtype=float)
    mask = np.isfinite(depths) & np.isfinite(rel)
    if not np.any(mask):
        return pd.DataFrame(columns=["Depth_m"])

    return pd.DataFrame(
        {
            "Depth_m": depths[mask],
            f"{stem}_input_added_total_m": rel[mask] + float(input_motion_max_abs),
        }
    )


def _get_common_time_for_layer(
    time_a: np.ndarray,
    time_b: np.ndarray,
) -> np.ndarray:
    start = max(float(time_a[0]), float(time_b[0]))
    end = min(float(time_a[-1]), float(time_b[-1]))
    if end <= start:
        return time_a.copy()

    dta = np.median(np.diff(time_a)) if time_a.size > 1 else 0.01
    dtb = np.median(np.diff(time_b)) if time_b.size > 1 else 0.01
    dt = min(dta, dtb)
    if not np.isfinite(dt) or dt <= 0:
        dt = 0.01

    count = int(math.floor((end - start) / dt)) + 1
    return start + np.arange(count, dtype=float) * dt


def _configure_chart_axes(chart) -> None:
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"


def _add_depth_profile_chart(
    worksheet,
    n_rows: int,
    *,
    depth_col: int = 2,
    series_start_col: int = 3,
    series_cols: Sequence[int] | None = None,
    title: str = "Depth-Dependent Total Displacement Profiles",
    anchor: str = "H2",
) -> None:
    if n_rows < 2:
        return

    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Displacement (m)"
    chart.y_axis.title = "Depth (m)"
    chart.scatterStyle = "lineMarker"
    chart.legend.position = "r"
    chart.y_axis.scaling.orientation = "maxMin"
    chart.height = 9.5
    chart.width = 15.0
    _configure_chart_axes(chart)

    if worksheet.max_column < series_start_col:
        return

    y_values = Reference(worksheet, min_col=depth_col, min_row=2, max_row=n_rows + 1)
    if series_cols is None:
        plot_cols = range(series_start_col, worksheet.max_column + 1)
    else:
        plot_cols = [col for col in series_cols if 1 <= col <= worksheet.max_column]
    for col in plot_cols:
        x_values = Reference(worksheet, min_col=col, min_row=2, max_row=n_rows + 1)
        # ScatterChart for openpyxl expects Series(y_values, x_values).
        series = Series(y_values, x_values, title=worksheet.cell(row=1, column=col).value)
        chart.series.append(series)

    if not chart.series:
        return

    worksheet.add_chart(chart, anchor)


def _add_base_corrected_chart(worksheet, n_rows: int) -> None:
    if n_rows < 2:
        return

    chart_x = ScatterChart()
    chart_x.title = "X Profile: Strain Base-Relative vs Deepsoil(Base-Corrected)"
    chart_x.x_axis.title = "Displacement (m)"
    chart_x.y_axis.title = "Depth (m)"
    chart_x.scatterStyle = "lineMarker"
    chart_x.legend.position = "r"
    chart_x.y_axis.scaling.orientation = "maxMin"
    chart_x.height = 8.5
    chart_x.width = 14.0
    _configure_chart_axes(chart_x)

    y_values = Reference(worksheet, min_col=2, min_row=2, max_row=n_rows + 1)
    for col in (3, 4):
        x_values = Reference(worksheet, min_col=col, min_row=2, max_row=n_rows + 1)
        series = Series(y_values, x_values, title=worksheet.cell(row=1, column=col).value)
        chart_x.series.append(series)

    chart_y = ScatterChart()
    chart_y.title = "Y Profile: Strain Base-Relative vs Deepsoil(Base-Corrected)"
    chart_y.x_axis.title = "Displacement (m)"
    chart_y.y_axis.title = "Depth (m)"
    chart_y.scatterStyle = "lineMarker"
    chart_y.legend.position = "r"
    chart_y.y_axis.scaling.orientation = "maxMin"
    chart_y.height = 8.5
    chart_y.width = 14.0
    _configure_chart_axes(chart_y)

    for col in (6, 7):
        x_values = Reference(worksheet, min_col=col, min_row=2, max_row=n_rows + 1)
        series = Series(y_values, x_values, title=worksheet.cell(row=1, column=col).value)
        chart_y.series.append(series)

    worksheet.add_chart(chart_x, "L2")
    worksheet.add_chart(chart_y, "L22")


def _add_all_layers_chart(
    worksheet,
    n_rows: int,
    n_series: int,
    title: str,
) -> None:
    if n_rows < 2 or n_series < 1:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Displacement (m)"
    chart.x_axis.title = "Time (s)"
    chart.height = 10.0
    chart.width = 18.0
    chart.legend.position = "r"
    _configure_chart_axes(chart)

    categories = Reference(worksheet, min_col=1, min_row=2, max_row=n_rows + 1)
    values = Reference(worksheet, min_col=2, max_col=n_series + 1, min_row=1, max_row=n_rows + 1)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, "B2")


def build_output_workbook(
    strain_df: pd.DataFrame,
    legacy_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    *,
    include_resultant_profiles: bool = True,
    profile_sheet_df: pd.DataFrame | None = None,
    x_time_df: pd.DataFrame | None = None,
    y_time_df: pd.DataFrame | None = None,
    resultant_time_df: pd.DataFrame | None = None,
    tbdy_total_x_time_df: pd.DataFrame | None = None,
    tbdy_total_y_time_df: pd.DataFrame | None = None,
    tbdy_total_resultant_time_df: pd.DataFrame | None = None,
    x_time_alt_df: pd.DataFrame | None = None,
    y_time_alt_df: pd.DataFrame | None = None,
    resultant_time_alt_df: pd.DataFrame | None = None,
    tbdy_total_x_time_alt_df: pd.DataFrame | None = None,
    tbdy_total_y_time_alt_df: pd.DataFrame | None = None,
    tbdy_total_resultant_time_alt_df: pd.DataFrame | None = None,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        strain_df.to_excel(writer, sheet_name="Strain_Relative", index=False)
        legacy_df.to_excel(writer, sheet_name="Legacy_Methods", index=False)
        comparison_df.to_excel(writer, sheet_name="Comparison", index=False)

        depth_profiles_df = _build_depth_profiles_df(
            comparison_df,
            include_resultants=include_resultant_profiles,
        )
        depth_profiles_df.to_excel(writer, sheet_name="Depth_Profiles", index=False)

        base_corrected_df = _build_base_corrected_profiles_df(comparison_df)
        base_corrected_df.to_excel(writer, sheet_name="Profile_BaseCorrected", index=False)

        if x_time_df is not None and not x_time_df.empty:
            x_time_df.to_excel(writer, sheet_name="Direction_X_Time", index=False)
        if y_time_df is not None and not y_time_df.empty:
            y_time_df.to_excel(writer, sheet_name="Direction_Y_Time", index=False)
        if resultant_time_df is not None and not resultant_time_df.empty:
            resultant_time_df.to_excel(writer, sheet_name="Resultant_Time", index=False)
        if tbdy_total_x_time_df is not None and not tbdy_total_x_time_df.empty:
            tbdy_total_x_time_df.to_excel(writer, sheet_name="TBDY_Total_X_Time", index=False)
        if tbdy_total_y_time_df is not None and not tbdy_total_y_time_df.empty:
            tbdy_total_y_time_df.to_excel(writer, sheet_name="TBDY_Total_Y_Time", index=False)
        if tbdy_total_resultant_time_df is not None and not tbdy_total_resultant_time_df.empty:
            tbdy_total_resultant_time_df.to_excel(writer, sheet_name="TBDY_Total_Resultant_Time", index=False)
        if x_time_alt_df is not None and not x_time_alt_df.empty:
            x_time_alt_df.to_excel(writer, sheet_name="Direction_X_Time_ALT", index=False)
        if y_time_alt_df is not None and not y_time_alt_df.empty:
            y_time_alt_df.to_excel(writer, sheet_name="Direction_Y_Time_ALT", index=False)
        if resultant_time_alt_df is not None and not resultant_time_alt_df.empty:
            resultant_time_alt_df.to_excel(writer, sheet_name="Resultant_Time_ALT", index=False)
        if tbdy_total_x_time_alt_df is not None and not tbdy_total_x_time_alt_df.empty:
            tbdy_total_x_time_alt_df.to_excel(writer, sheet_name="TBDY_Total_X_Time_ALT", index=False)
        if tbdy_total_y_time_alt_df is not None and not tbdy_total_y_time_alt_df.empty:
            tbdy_total_y_time_alt_df.to_excel(writer, sheet_name="TBDY_Total_Y_Time_ALT", index=False)
        if tbdy_total_resultant_time_alt_df is not None and not tbdy_total_resultant_time_alt_df.empty:
            tbdy_total_resultant_time_alt_df.to_excel(writer, sheet_name="TBDY_Total_Resultant_Time_ALT", index=False)

        workbook = writer.book

        if "Depth_Profiles" in writer.sheets:
            ws_depth = writer.sheets["Depth_Profiles"]
            _add_depth_profile_chart(ws_depth, len(depth_profiles_df), depth_col=2, series_start_col=3)

        if "Profile_BaseCorrected" in writer.sheets:
            ws_bc = writer.sheets["Profile_BaseCorrected"]
            _add_base_corrected_chart(ws_bc, len(base_corrected_df))

        if "Direction_X_Time" in writer.sheets:
            ws_x = writer.sheets["Direction_X_Time"]
            _add_all_layers_chart(
                ws_x,
                ws_x.max_row - 1,
                ws_x.max_column - 1,
                "Direction X: All Layers Signed Displacement-Time",
            )

        if "Direction_Y_Time" in writer.sheets:
            ws_y = writer.sheets["Direction_Y_Time"]
            _add_all_layers_chart(
                ws_y,
                ws_y.max_row - 1,
                ws_y.max_column - 1,
                "Direction Y: All Layers Signed Displacement-Time",
            )

        if "Resultant_Time" in writer.sheets:
            ws_r = writer.sheets["Resultant_Time"]
            _add_all_layers_chart(
                ws_r,
                ws_r.max_row - 1,
                ws_r.max_column - 1,
                "Resultant: All Layers Displacement-Time",
            )

        if "TBDY_Total_X_Time" in writer.sheets:
            ws_tx = writer.sheets["TBDY_Total_X_Time"]
            _add_all_layers_chart(
                ws_tx,
                ws_tx.max_row - 1,
                ws_tx.max_column - 1,
                "TBDY Total X: u(base)+u(rel)",
            )

        if "TBDY_Total_Y_Time" in writer.sheets:
            ws_ty = writer.sheets["TBDY_Total_Y_Time"]
            _add_all_layers_chart(
                ws_ty,
                ws_ty.max_row - 1,
                ws_ty.max_column - 1,
                "TBDY Total Y: u(base)+u(rel)",
            )

        if "TBDY_Total_Resultant_Time" in writer.sheets:
            ws_tr = writer.sheets["TBDY_Total_Resultant_Time"]
            _add_all_layers_chart(
                ws_tr,
                ws_tr.max_row - 1,
                ws_tr.max_column - 1,
                "TBDY Total Resultant: All Layers",
            )

        if "Direction_X_Time_ALT" in writer.sheets:
            ws_x_alt = writer.sheets["Direction_X_Time_ALT"]
            _add_all_layers_chart(
                ws_x_alt,
                ws_x_alt.max_row - 1,
                ws_x_alt.max_column - 1,
                "Direction X ALT: FFT-Regularized",
            )

        if "Direction_Y_Time_ALT" in writer.sheets:
            ws_y_alt = writer.sheets["Direction_Y_Time_ALT"]
            _add_all_layers_chart(
                ws_y_alt,
                ws_y_alt.max_row - 1,
                ws_y_alt.max_column - 1,
                "Direction Y ALT: FFT-Regularized",
            )

        if "Resultant_Time_ALT" in writer.sheets:
            ws_r_alt = writer.sheets["Resultant_Time_ALT"]
            _add_all_layers_chart(
                ws_r_alt,
                ws_r_alt.max_row - 1,
                ws_r_alt.max_column - 1,
                "Resultant ALT: FFT-Regularized",
            )

        if "TBDY_Total_X_Time_ALT" in writer.sheets:
            ws_tx_alt = writer.sheets["TBDY_Total_X_Time_ALT"]
            _add_all_layers_chart(
                ws_tx_alt,
                ws_tx_alt.max_row - 1,
                ws_tx_alt.max_column - 1,
                "TBDY Total X ALT: FFT-Regularized",
            )

        if "TBDY_Total_Y_Time_ALT" in writer.sheets:
            ws_ty_alt = writer.sheets["TBDY_Total_Y_Time_ALT"]
            _add_all_layers_chart(
                ws_ty_alt,
                ws_ty_alt.max_row - 1,
                ws_ty_alt.max_column - 1,
                "TBDY Total Y ALT: FFT-Regularized",
            )

        if "TBDY_Total_Resultant_Time_ALT" in writer.sheets:
            ws_tr_alt = writer.sheets["TBDY_Total_Resultant_Time_ALT"]
            _add_all_layers_chart(
                ws_tr_alt,
                ws_tr_alt.max_row - 1,
                ws_tr_alt.max_column - 1,
                "TBDY Total Resultant ALT: FFT-Regularized",
            )

        _ = workbook

    return buffer.getvalue()


def build_single_output_workbook(
    summary_df: pd.DataFrame,
    direction_time_df: pd.DataFrame,
    profile_sheet_df: pd.DataFrame | None = None,
    strain_rel_time_df: pd.DataFrame | None = None,
    tbdy_total_time_df: pd.DataFrame | None = None,
    input_proxy_rel_time_df: pd.DataFrame | None = None,
    direction_time_alt_df: pd.DataFrame | None = None,
    strain_rel_time_alt_df: pd.DataFrame | None = None,
    tbdy_total_time_alt_df: pd.DataFrame | None = None,
    input_proxy_rel_time_alt_df: pd.DataFrame | None = None,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Single_Direction_Summary", index=False)
        direction_time_df.to_excel(writer, sheet_name="Direction_Time", index=False)

        if strain_rel_time_df is not None and not strain_rel_time_df.empty:
            strain_rel_time_df.to_excel(writer, sheet_name="Strain_Relative_Time", index=False)
        if tbdy_total_time_df is not None and not tbdy_total_time_df.empty:
            tbdy_total_time_df.to_excel(writer, sheet_name="TBDY_Total_Time", index=False)
        if input_proxy_rel_time_df is not None and not input_proxy_rel_time_df.empty:
            input_proxy_rel_time_df.to_excel(writer, sheet_name="InputProxy_Relative_Time", index=False)
        if direction_time_alt_df is not None and not direction_time_alt_df.empty:
            direction_time_alt_df.to_excel(writer, sheet_name="Direction_Time_ALT", index=False)
        if strain_rel_time_alt_df is not None and not strain_rel_time_alt_df.empty:
            strain_rel_time_alt_df.to_excel(writer, sheet_name="Strain_Relative_Time_ALT", index=False)
        if tbdy_total_time_alt_df is not None and not tbdy_total_time_alt_df.empty:
            tbdy_total_time_alt_df.to_excel(writer, sheet_name="TBDY_Total_Time_ALT", index=False)
        if input_proxy_rel_time_alt_df is not None and not input_proxy_rel_time_alt_df.empty:
            input_proxy_rel_time_alt_df.to_excel(writer, sheet_name="InputProxy_Relative_Time_ALT", index=False)

        if "Direction_Time" in writer.sheets:
            ws_dir = writer.sheets["Direction_Time"]
            _add_all_layers_chart(
                ws_dir,
                ws_dir.max_row - 1,
                ws_dir.max_column - 1,
                "Single Direction: All Layers Displacement-Time",
            )

        if "Strain_Relative_Time" in writer.sheets:
            ws_sr = writer.sheets["Strain_Relative_Time"]
            _add_all_layers_chart(
                ws_sr,
                ws_sr.max_row - 1,
                ws_sr.max_column - 1,
                "Single Direction: Strain Base-Relative Time",
            )

        if "TBDY_Total_Time" in writer.sheets:
            ws_tb = writer.sheets["TBDY_Total_Time"]
            _add_all_layers_chart(
                ws_tb,
                ws_tb.max_row - 1,
                ws_tb.max_column - 1,
                "Single Direction: TBDY Total Time (u_base + u_rel)",
            )

        if "InputProxy_Relative_Time" in writer.sheets:
            ws_ip = writer.sheets["InputProxy_Relative_Time"]
            _add_all_layers_chart(
                ws_ip,
                ws_ip.max_row - 1,
                ws_ip.max_column - 1,
                "Single Direction: Input-Proxy Relative Time",
            )
        if "Direction_Time_ALT" in writer.sheets:
            ws_dir_alt = writer.sheets["Direction_Time_ALT"]
            _add_all_layers_chart(
                ws_dir_alt,
                ws_dir_alt.max_row - 1,
                ws_dir_alt.max_column - 1,
                "Single Direction ALT: FFT-Regularized",
            )
        if "Strain_Relative_Time_ALT" in writer.sheets:
            ws_sr_alt = writer.sheets["Strain_Relative_Time_ALT"]
            _add_all_layers_chart(
                ws_sr_alt,
                ws_sr_alt.max_row - 1,
                ws_sr_alt.max_column - 1,
                "Single Direction ALT: Strain Base-Relative",
            )
        if "TBDY_Total_Time_ALT" in writer.sheets:
            ws_tb_alt = writer.sheets["TBDY_Total_Time_ALT"]
            _add_all_layers_chart(
                ws_tb_alt,
                ws_tb_alt.max_row - 1,
                ws_tb_alt.max_column - 1,
                "Single Direction ALT: TBDY Total",
            )
        if "InputProxy_Relative_Time_ALT" in writer.sheets:
            ws_ip_alt = writer.sheets["InputProxy_Relative_Time_ALT"]
            _add_all_layers_chart(
                ws_ip_alt,
                ws_ip_alt.max_row - 1,
                ws_ip_alt.max_column - 1,
                "Single Direction ALT: Input-Proxy Relative",
            )

    return buffer.getvalue()


def _build_db_depth_profiles_df(summary_df: pd.DataFrame, include_resultants: bool = True) -> pd.DataFrame:
    cols = [
        "Layer_Index",
        "Depth_m",
        "X_total_pos_max_m",
        "X_total_neg_min_m",
        "Y_total_pos_max_m",
        "Y_total_neg_min_m",
        "X_relative_pos_max_m",
        "X_relative_neg_min_m",
        "Y_relative_pos_max_m",
        "Y_relative_neg_min_m",
        "X_total_maxabs_m",
        "Y_total_maxabs_m",
        "X_relative_maxabs_m",
        "Y_relative_maxabs_m",
    ]
    if include_resultants:
        cols.extend(
            [
                "Total_resultant_maxabs_m",
                "Relative_resultant_maxabs_m",
            ]
        )
    present = [col for col in cols if col in summary_df.columns]
    return summary_df[present].copy()


def build_db_pair_output_workbook(
    summary_df: pd.DataFrame,
    x_total_time_df: pd.DataFrame,
    y_total_time_df: pd.DataFrame,
    x_relative_time_df: pd.DataFrame,
    y_relative_time_df: pd.DataFrame,
    *,
    include_resultant_profiles: bool = True,
    total_resultant_time_df: pd.DataFrame | None = None,
    relative_resultant_time_df: pd.DataFrame | None = None,
) -> bytes:
    depth_profiles_df = _build_db_depth_profiles_df(summary_df, include_resultants=include_resultant_profiles)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="DB_Profile_Summary", index=False)
        depth_profiles_df.to_excel(writer, sheet_name="DB_Depth_Profiles", index=False)
        x_total_time_df.to_excel(writer, sheet_name="DB_Total_X_Time", index=False)
        y_total_time_df.to_excel(writer, sheet_name="DB_Total_Y_Time", index=False)
        x_relative_time_df.to_excel(writer, sheet_name="DB_Relative_X_Time", index=False)
        y_relative_time_df.to_excel(writer, sheet_name="DB_Relative_Y_Time", index=False)
        if include_resultant_profiles and total_resultant_time_df is not None and not total_resultant_time_df.empty:
            total_resultant_time_df.to_excel(writer, sheet_name="DB_Total_Resultant_Time", index=False)
        if include_resultant_profiles and relative_resultant_time_df is not None and not relative_resultant_time_df.empty:
            relative_resultant_time_df.to_excel(writer, sheet_name="DB_Relative_Resultant_Time", index=False)

        if "DB_Depth_Profiles" in writer.sheets:
            ws_depth = writer.sheets["DB_Depth_Profiles"]
            _add_depth_profile_chart(ws_depth, len(depth_profiles_df), depth_col=2, series_start_col=3)
        if "DB_Total_X_Time" in writer.sheets:
            ws_x = writer.sheets["DB_Total_X_Time"]
            _add_all_layers_chart(ws_x, ws_x.max_row - 1, ws_x.max_column - 1, "DB Total X: All Layers")
        if "DB_Total_Y_Time" in writer.sheets:
            ws_y = writer.sheets["DB_Total_Y_Time"]
            _add_all_layers_chart(ws_y, ws_y.max_row - 1, ws_y.max_column - 1, "DB Total Y: All Layers")
        if "DB_Relative_X_Time" in writer.sheets:
            ws_rx = writer.sheets["DB_Relative_X_Time"]
            _add_all_layers_chart(ws_rx, ws_rx.max_row - 1, ws_rx.max_column - 1, "DB Relative X: All Layers")
        if "DB_Relative_Y_Time" in writer.sheets:
            ws_ry = writer.sheets["DB_Relative_Y_Time"]
            _add_all_layers_chart(ws_ry, ws_ry.max_row - 1, ws_ry.max_column - 1, "DB Relative Y: All Layers")
        if "DB_Total_Resultant_Time" in writer.sheets:
            ws_tr = writer.sheets["DB_Total_Resultant_Time"]
            _add_all_layers_chart(ws_tr, ws_tr.max_row - 1, ws_tr.max_column - 1, "DB Total Resultant")
        if "DB_Relative_Resultant_Time" in writer.sheets:
            ws_rr = writer.sheets["DB_Relative_Resultant_Time"]
            _add_all_layers_chart(ws_rr, ws_rr.max_row - 1, ws_rr.max_column - 1, "DB Relative Resultant")

    return buffer.getvalue()


def build_db_single_output_workbook(
    summary_df: pd.DataFrame,
    total_time_df: pd.DataFrame,
    relative_time_df: pd.DataFrame,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="DB_Summary", index=False)
        total_time_df.to_excel(writer, sheet_name="DB_Total_Time", index=False)
        relative_time_df.to_excel(writer, sheet_name="DB_Relative_Time", index=False)

        if "DB_Total_Time" in writer.sheets:
            ws_total = writer.sheets["DB_Total_Time"]
            _add_all_layers_chart(ws_total, ws_total.max_row - 1, ws_total.max_column - 1, "DB Total: All Layers")
        if "DB_Relative_Time" in writer.sheets:
            ws_rel = writer.sheets["DB_Relative_Time"]
            _add_all_layers_chart(ws_rel, ws_rel.max_row - 1, ws_rel.max_column - 1, "DB Relative: All Layers")

    return buffer.getvalue()


def _db_method2_summary_sheet_name(axis_label: str) -> str:
    axis = axis_label.upper()
    if axis == "X":
        return "Method2_DB_Summary_X"
    if axis == "Y":
        return "Method2_DB_Summary_Y"
    return "Method2_DB_Summary"


def _db_method2_total_sheet_name(axis_label: str) -> str:
    axis = axis_label.upper()
    if axis == "X":
        return "Method2_DB_Total_X_Time"
    if axis == "Y":
        return "Method2_DB_Total_Y_Time"
    return "Method2_DB_Total_Time"


def _db_method2_relative_sheet_name(axis_label: str) -> str:
    axis = axis_label.upper()
    if axis == "X":
        return "Method2_DB_Relative_X_Time"
    if axis == "Y":
        return "Method2_DB_Relative_Y_Time"
    return "Method2_DB_Relative_Time"


def _build_db_method2_workbook(
    summary_df: pd.DataFrame,
    total_time_df: pd.DataFrame,
    relative_time_df: pd.DataFrame,
    axis_label: str,
) -> bytes:
    summary_sheet = _db_method2_summary_sheet_name(axis_label)
    total_sheet = _db_method2_total_sheet_name(axis_label)
    relative_sheet = _db_method2_relative_sheet_name(axis_label)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=summary_sheet, index=False)
        total_time_df.to_excel(writer, sheet_name=total_sheet, index=False)
        relative_time_df.to_excel(writer, sheet_name=relative_sheet, index=False)

        if total_sheet in writer.sheets:
            ws_total = writer.sheets[total_sheet]
            _add_all_layers_chart(
                ws_total,
                ws_total.max_row - 1,
                ws_total.max_column - 1,
                f"Method-2 DB Total {axis_label}: All Layers",
            )
        if relative_sheet in writer.sheets:
            ws_relative = writer.sheets[relative_sheet]
            _add_all_layers_chart(
                ws_relative,
                ws_relative.max_row - 1,
                ws_relative.max_column - 1,
                f"Method-2 DB Relative {axis_label}: All Layers",
            )

    return buffer.getvalue()


def _build_db_method3_aggregate_workbook(
    profile_x_df: pd.DataFrame,
    profile_y_df: pd.DataFrame,
    profile_single_df: pd.DataFrame,
    relative_x_df: pd.DataFrame | None = None,
    relative_y_df: pd.DataFrame | None = None,
    relative_single_df: pd.DataFrame | None = None,
) -> bytes:
    x_df = profile_x_df if profile_x_df is not None and not profile_x_df.empty else pd.DataFrame(columns=["Depth_m"])
    y_df = profile_y_df if profile_y_df is not None and not profile_y_df.empty else pd.DataFrame(columns=["Depth_m"])
    single_df = profile_single_df if profile_single_df is not None and not profile_single_df.empty else pd.DataFrame(columns=["Depth_m"])
    rel_x_df = relative_x_df if relative_x_df is not None and not relative_x_df.empty else pd.DataFrame(columns=["Depth_m"])
    rel_y_df = relative_y_df if relative_y_df is not None and not relative_y_df.empty else pd.DataFrame(columns=["Depth_m"])
    rel_single_df = (
        relative_single_df
        if relative_single_df is not None and not relative_single_df.empty
        else pd.DataFrame(columns=["Depth_m"])
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        x_df.to_excel(writer, sheet_name="Method3_DB_Profile_X", index=False)
        y_df.to_excel(writer, sheet_name="Method3_DB_Profile_Y", index=False)
        single_df.to_excel(writer, sheet_name="Method3_DB_Profile_Single", index=False)
        rel_x_df.to_excel(writer, sheet_name="Method3_DB_Relative_X", index=False)
        rel_y_df.to_excel(writer, sheet_name="Method3_DB_Relative_Y", index=False)
        rel_single_df.to_excel(writer, sheet_name="Method3_DB_Relative_Single", index=False)

        if "Method3_DB_Profile_X" in writer.sheets:
            ws_x = writer.sheets["Method3_DB_Profile_X"]
            _add_depth_profile_chart(ws_x, len(x_df), depth_col=1, series_start_col=2)
        if "Method3_DB_Profile_Y" in writer.sheets:
            ws_y = writer.sheets["Method3_DB_Profile_Y"]
            _add_depth_profile_chart(ws_y, len(y_df), depth_col=1, series_start_col=2)
        if "Method3_DB_Profile_Single" in writer.sheets:
            ws_single = writer.sheets["Method3_DB_Profile_Single"]
            _add_depth_profile_chart(ws_single, len(single_df), depth_col=1, series_start_col=2)
        if "Method3_DB_Relative_X" in writer.sheets:
            ws_rel_x = writer.sheets["Method3_DB_Relative_X"]
            _add_depth_profile_chart(ws_rel_x, len(rel_x_df), depth_col=1, series_start_col=2)
        if "Method3_DB_Relative_Y" in writer.sheets:
            ws_rel_y = writer.sheets["Method3_DB_Relative_Y"]
            _add_depth_profile_chart(ws_rel_y, len(rel_y_df), depth_col=1, series_start_col=2)
        if "Method3_DB_Relative_Single" in writer.sheets:
            ws_rel_single = writer.sheets["Method3_DB_Relative_Single"]
            _add_depth_profile_chart(ws_rel_single, len(rel_single_df), depth_col=1, series_start_col=2)

    return buffer.getvalue()


def _import_sqlite3():
    try:
        import sqlite3  # type: ignore

        return sqlite3
    except ModuleNotFoundError as exc:  # pragma: no cover - environment-specific
        raise ModuleNotFoundError(
            "sqlite3 is required for DB direct mode. In Pyodide, load the 'sqlite3' package before importing disp_core."
        ) from exc


def _extract_db_method2_single(
    db_bytes: bytes,
    file_name: str,
    options: Mapping[str, Any] | None = None,
    axis_override: str | None = None,
) -> Dict[str, Any]:
    _ = options
    bundle = _read_db_disp_bundle(db_bytes, file_name, axis_override=axis_override)
    axis_label = str(bundle.get("axis", "SINGLE")).upper()
    summary_df = bundle["summary_df"].copy()
    output_bytes = _build_db_method2_workbook(
        summary_df,
        bundle["total_time_df"],
        bundle["relative_time_df"],
        axis_label,
    )
    record_label = bundle["recordLabel"]
    profile_df = pd.DataFrame(
        {
            "Depth_m": bundle["depths"],
            f"{record_label}_maxabs_m": np.max(np.abs(bundle["disp_matrix"]), axis=1),
        }
    )
    relative_profile_df = pd.DataFrame(
        {
            "Depth_m": bundle["depths"],
            f"{record_label}_relative_maxabs_m": np.max(np.abs(bundle["relative_matrix"]), axis=1),
        }
    )
    return {
        "skipped": False,
        "axis": axis_label,
        "profile_df": profile_df,
        "relative_profile_df": relative_profile_df,
        "result": {
            "pairKey": f"DB_METHOD2|{record_label}",
            "xFileName": file_name,
            "yFileName": "",
            "outputFileName": f"output_method2_db_{record_label}.xlsx",
            "outputBytes": output_bytes,
            "metrics": {
                "mode": "db_method2_single",
                "axis": axis_label,
                "layerCount": int(len(summary_df)),
                "timeSeriesSheets": 2,
                "timeSheets": [
                    _db_method2_total_sheet_name(axis_label),
                    _db_method2_relative_sheet_name(axis_label),
                ],
                "surfaceBaseTotal_m": float(summary_df["DB_Total_maxabs_m"].iloc[0]),
                "surfaceProfileRSS_m": float(summary_df["DB_Relative_maxabs_m"].iloc[0]),
                "baseReference": "db_direct",
                "integrationPrimary": "deepsoil_db",
                "useDb3Directly": True,
            },
        },
    }


def _process_db_batch_files(file_map: Mapping[str, bytes], options: Mapping[str, Any] | None = None) -> Dict[str, Any]:
    normalized_options = _normalize_options(options)
    fail_fast = bool(normalized_options.get("failFast", False))
    manual_pairing_enabled = _to_bool(normalized_options.get("manualPairingEnabled", False), False)
    method2_enabled = _to_bool(
        normalized_options.get("method2Enabled", normalized_options.get("method23Enabled", True)),
        True,
    )
    method3_enabled = _to_bool(
        normalized_options.get("method3Enabled", normalized_options.get("method23Enabled", True)),
        True,
    )

    logs: List[Dict[str, str]] = []
    errors: List[Dict[str, Any]] = []
    results: List[Dict[str, Any]] = []
    source_catalog: List[Dict[str, Any]] = []
    summary_catalog: List[Dict[str, Any]] = []
    return_web_results = _to_bool(normalized_options.get("_returnWebResults", False), False)

    def _store_result(result: Dict[str, Any]) -> None:
        if return_web_results:
            results.append(_build_web_result_payload(result))
        else:
            results.append(result)

    def _store_source_entry(entry: Dict[str, Any] | None) -> None:
        if isinstance(entry, dict) and entry.get("families"):
            source_catalog.append(entry)

    def _store_summary_entry(entry: Dict[str, Any] | None) -> None:
        if isinstance(entry, dict) and entry.get("variants"):
            summary_catalog.append(entry)

    file_names = sorted(file_map.keys())
    db_candidates = sorted([name for name in file_names if _candidate_kind(name) == "db" and _is_candidate_file(name, False)])
    pairs, missing, pair_warnings = _resolve_xy_pairs(
        db_candidates,
        include_manip=False,
        manual_pairing_enabled=manual_pairing_enabled,
        manual_pairs=normalized_options.get("manualPairs"),
    )
    used_in_pairs = {name for pair in pairs for name in pair}
    singles = sorted([name for name in db_candidates if name not in used_in_pairs])
    manual_axis_map: Dict[str, str] = {}
    for x_name, y_name in pairs:
        manual_axis_map[x_name] = "X"
        manual_axis_map[y_name] = "Y"

    _log(logs, "info", f"DB3 direct mode: on")
    _log(logs, "info", f"DB candidates: {len(db_candidates)}")
    _log(logs, "info", f"Detected DB X/Y pairs: {len(pairs)}")
    _log(logs, "info", f"Detected DB single files: {len(singles)}")
    _log(logs, "info", f"Manual pairing: {'on' if manual_pairing_enabled else 'off'}")
    _log(logs, "info", f"Method-2 output: {'on' if method2_enabled else 'off'}")
    _log(logs, "info", f"Method-3 output: {'on' if method3_enabled else 'off'}")

    for warning in pair_warnings:
        _log(logs, "warning", warning)
    for missing_x in missing:
        _log(logs, "warning", f"No Y match for DB X file: {missing_x}")

    for x_name, y_name in pairs:
        try:
            x_bundle = _read_db_disp_bundle(file_map[x_name], x_name, axis_override="X")
            y_bundle = _read_db_disp_bundle(file_map[y_name], y_name, axis_override="Y")
            n_layers = min(
                int(x_bundle["disp_matrix"].shape[0]),
                int(y_bundle["disp_matrix"].shape[0]),
                int(x_bundle["depths"].size),
                int(y_bundle["depths"].size),
            )
            if n_layers > 0:
                summary_df = pd.DataFrame(
                    {
                        "Depth_m": x_bundle["depths"][:n_layers],
                        "X_total_maxabs_m": np.max(np.abs(x_bundle["disp_matrix"][:n_layers, :]), axis=1),
                        "Y_total_maxabs_m": np.max(np.abs(y_bundle["disp_matrix"][:n_layers, :]), axis=1),
                        "X_relative_maxabs_m": np.max(np.abs(x_bundle["relative_matrix"][:n_layers, :]), axis=1),
                        "Y_relative_maxabs_m": np.max(np.abs(y_bundle["relative_matrix"][:n_layers, :]), axis=1),
                        "Total_resultant_maxabs_m": np.max(
                            np.sqrt(
                                x_bundle["disp_matrix"][:n_layers, :] ** 2 + y_bundle["disp_matrix"][:n_layers, :] ** 2
                            ),
                            axis=1,
                        ),
                        "Relative_resultant_maxabs_m": np.max(
                            np.sqrt(
                                x_bundle["relative_matrix"][:n_layers, :] ** 2
                                + y_bundle["relative_matrix"][:n_layers, :] ** 2
                            ),
                            axis=1,
                        ),
                    }
                )
                _store_source_entry(_build_db_pair_source_catalog_entry(x_name, y_name, summary_df, x_bundle, y_bundle))
                _store_summary_entry(_build_db_pair_summary_entry(x_name, y_name, summary_df))
        except Exception as exc:  # noqa: BLE001
            _log(logs, "warning", f"DB source catalog skipped for pair {x_name} + {y_name}: {exc}")

    for name in singles:
        try:
            bundle = _read_db_disp_bundle(file_map[name], name, axis_override=manual_axis_map.get(name))
            _store_source_entry(_build_db_single_source_catalog_entry(name, bundle))
            _store_summary_entry(_build_db_single_summary_entry(name, bundle))
        except Exception as exc:  # noqa: BLE001
            _log(logs, "warning", f"DB source catalog skipped for single {name}: {exc}")

    method2_detected = len(db_candidates) if (method2_enabled or method3_enabled) else 0
    method2_processed = 0
    method2_failed = 0
    method3_produced = 0
    progress_total = method2_detected + (1 if method3_enabled and method2_detected > 0 else 0)
    progress_completed = 0
    profile_x_frames: List[pd.DataFrame] = []
    profile_y_frames: List[pd.DataFrame] = []
    profile_single_frames: List[pd.DataFrame] = []
    relative_x_frames: List[pd.DataFrame] = []
    relative_y_frames: List[pd.DataFrame] = []
    relative_single_frames: List[pd.DataFrame] = []

    def _advance_progress(message: str) -> None:
        nonlocal progress_completed
        if progress_total <= 0:
            return
        progress_completed += 1
        _report_batch_progress(normalized_options, progress_completed, progress_total, message)

    if progress_total > 0:
        _report_batch_progress(
            normalized_options,
            0,
            progress_total,
            f"DB batch calculation started (0/{progress_total})",
        )

    if method2_enabled or method3_enabled:
        for name in db_candidates:
            try:
                extracted = _extract_db_method2_single(
                    file_map[name],
                    name,
                    normalized_options,
                    axis_override=manual_axis_map.get(name),
                )
                axis = str(extracted.get("axis", "")).upper()
                profile_df = extracted.get("profile_df")
                relative_profile_df = extracted.get("relative_profile_df")
                if method2_enabled:
                    _store_result(extracted["result"])
                    method2_processed += 1
                if method3_enabled and isinstance(profile_df, pd.DataFrame) and not profile_df.empty:
                    if axis == "X":
                        profile_x_frames.append(profile_df)
                    elif axis == "Y":
                        profile_y_frames.append(profile_df)
                    else:
                        profile_single_frames.append(profile_df)
                if method3_enabled and isinstance(relative_profile_df, pd.DataFrame) and not relative_profile_df.empty:
                    if axis == "X":
                        relative_x_frames.append(relative_profile_df)
                    elif axis == "Y":
                        relative_y_frames.append(relative_profile_df)
                    else:
                        relative_single_frames.append(relative_profile_df)
                _log(logs, "info", f"Processed DB method basis file: {name}")
                _advance_progress(f"Processed DB method basis file ({progress_completed + 1}/{progress_total}): {name}")
            except Exception as exc:  # noqa: BLE001
                method2_failed += 1
                errors.append({"pairKey": f"DB_METHOD2|{name}", "reason": str(exc)})
                _log(logs, "error", f"Failed DB file {name}: {exc}")
                _advance_progress(f"Failed DB method basis file ({progress_completed + 1}/{progress_total}): {name}")
                if fail_fast:
                    break

    if method3_enabled and (not fail_fast or not errors):
        profile_x_df = _merge_profile_frames(profile_x_frames)
        profile_y_df = _merge_profile_frames(profile_y_frames)
        profile_single_df = _merge_profile_frames(profile_single_frames)
        relative_x_df = _merge_profile_frames(relative_x_frames)
        relative_y_df = _merge_profile_frames(relative_y_frames)
        relative_single_df = _merge_profile_frames(relative_single_frames)
        if not profile_x_df.empty or not profile_y_df.empty or not profile_single_df.empty:
            try:
                method3_bytes = _build_db_method3_aggregate_workbook(
                    profile_x_df,
                    profile_y_df,
                    profile_single_df,
                    relative_x_df=relative_x_df,
                    relative_y_df=relative_y_df,
                    relative_single_df=relative_single_df,
                )
                _store_result(
                    {
                        "pairKey": "DB_METHOD3|ALL",
                        "xFileName": "",
                        "yFileName": "",
                        "outputFileName": "output_method3_db_profiles_all.xlsx",
                        "outputBytes": method3_bytes,
                        "previewCharts": [
                            chart
                            for chart in (
                                _aggregate_depth_preview_chart("DB Method-3 X Profiles", "Method3_DB_Profile_X", profile_x_df),
                                _aggregate_depth_preview_chart("DB Method-3 Y Profiles", "Method3_DB_Profile_Y", profile_y_df),
                                _aggregate_depth_preview_chart(
                                    "DB Method-3 Single Profiles",
                                    "Method3_DB_Profile_Single",
                                    profile_single_df,
                                ),
                            )
                            if chart is not None
                        ],
                        "metrics": {
                            "mode": "db_method3_aggregate",
                            "baseReference": "db_direct",
                            "integrationPrimary": "deepsoil_db",
                            "layerCount": max(int(len(profile_x_df)), int(len(profile_y_df)), int(len(profile_single_df))),
                            "xDepthRows": int(len(profile_x_df)),
                            "yDepthRows": int(len(profile_y_df)),
                            "singleDepthRows": int(len(profile_single_df)),
                            "xProfileColumns": max(0, int(profile_x_df.shape[1]) - 1),
                            "yProfileColumns": max(0, int(profile_y_df.shape[1]) - 1),
                            "singleProfileColumns": max(0, int(profile_single_df.shape[1]) - 1),
                            "xRelativeColumns": max(0, int(relative_x_df.shape[1]) - 1),
                            "yRelativeColumns": max(0, int(relative_y_df.shape[1]) - 1),
                            "singleRelativeColumns": max(0, int(relative_single_df.shape[1]) - 1),
                            "useDb3Directly": True,
                        },
                    }
                )
                method3_produced = 1
                _log(logs, "info", "Produced DB Method-3 aggregate workbook: output_method3_db_profiles_all.xlsx")
                _advance_progress(
                    f"Produced DB Method-3 aggregate workbook ({progress_completed + 1}/{progress_total})"
                )
            except Exception as exc:  # noqa: BLE001
                errors.append({"pairKey": "DB_METHOD3|ALL", "reason": str(exc)})
                _log(logs, "error", f"Failed DB Method-3 aggregate workbook: {exc}")
                _advance_progress(
                    f"Failed DB Method-3 aggregate workbook ({progress_completed + 1}/{progress_total})"
                )
        else:
            _log(logs, "warning", "DB Method-3 aggregate workbook skipped: no valid X/Y/Single DB profiles found.")
            _advance_progress(
                f"Skipped DB Method-3 aggregate workbook ({progress_completed + 1}/{progress_total})"
            )

    processed_total = method2_processed + method3_produced
    failed_total = method2_failed

    return {
        "results": results,
        "sourceCatalog": source_catalog,
        "summaryCatalog": summary_catalog,
        "logs": logs,
        "errors": errors,
        "metrics": {
            "pairsDetected": len(pairs),
            "pairsProcessed": 0,
            "pairsFailed": 0,
            "pairsMissing": len(missing),
            "dbCandidates": len(db_candidates),
            "xlsxCandidates": 0,
            "singlesDetected": len(singles),
            "singlesProcessed": 0,
            "singlesFailed": 0,
            "method2Enabled": bool(method2_enabled),
            "method3Enabled": bool(method3_enabled),
            "includeResultantProfiles": False,
            "baseReference": "db_direct",
            "integrationPrimary": "deepsoil_db",
            "integrationCompareEnabled": False,
            "altIntegrationMethod": None,
            "altLowCutPolicy": None,
            "useDb3Directly": True,
            "manualPairingEnabled": bool(manual_pairing_enabled),
            "manualPairsApplied": len(pairs),
            "method2Detected": method2_detected,
            "method2Processed": method2_processed,
            "method2Failed": method2_failed,
            "method3Produced": method3_produced,
            "processedTotal": processed_total,
            "failedTotal": failed_total,
        },
    }


def _read_db_disp_bundle(
    db_bytes: bytes,
    file_name: str,
    axis_override: str | None = None,
) -> Dict[str, Any]:
    sqlite3 = _import_sqlite3()
    suffix = _candidate_suffix(file_name)
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix or ".db3") as temp_file:
            temp_file.write(db_bytes)
            temp_path = temp_file.name

        conn = sqlite3.connect(temp_path)
        try:
            profile_df = pd.read_sql_query(
                (
                    "SELECT LAYER_NUMBER, DEPTH_LAYER_TOP, DEPTH_LAYER_MID, "
                    "MIN_DISP_RELATIVE, MAX_DISP_RELATIVE "
                    "FROM PROFILES ORDER BY LAYER_NUMBER"
                ),
                conn,
            )
            if profile_df.empty:
                raise ValueError(f"PROFILES table is empty in DB file: {file_name}")

            vel_columns = [row[1] for row in conn.execute("PRAGMA table_info(VEL_DISP)").fetchall()]
            if not vel_columns:
                raise ValueError(f"VEL_DISP table not found in DB file: {file_name}")

            layers: List[int] = []
            depths: List[float] = []
            total_cols: List[str] = []
            relative_cols: List[str] = []
            for _, row in profile_df.iterrows():
                layer_no = int(row["LAYER_NUMBER"])
                total_col = f"LAYER{layer_no}_DISP_TOTAL"
                relative_col = f"LAYER{layer_no}_DISP_RELATIVE"
                if total_col not in vel_columns or relative_col not in vel_columns:
                    continue
                depth = float(row["DEPTH_LAYER_TOP"])
                if not np.isfinite(depth):
                    continue
                layers.append(layer_no)
                depths.append(depth)
                total_cols.append(total_col)
                relative_cols.append(relative_col)

            if not layers:
                raise ValueError(f"No displacement columns found in VEL_DISP for DB file: {file_name}")

            query_cols = ["TIME", *total_cols, *relative_cols]
            vel_df = pd.read_sql_query(f"SELECT {', '.join(query_cols)} FROM VEL_DISP", conn)
        finally:
            conn.close()

        vel_df["TIME"] = pd.to_numeric(vel_df["TIME"], errors="coerce")
        vel_df = vel_df.dropna(subset=["TIME"]).sort_values("TIME")
        if vel_df.empty:
            raise ValueError(f"VEL_DISP contains no valid time rows for DB file: {file_name}")

        time = vel_df["TIME"].to_numpy(dtype=float)
        total_matrix = np.vstack([pd.to_numeric(vel_df[col], errors="coerce").fillna(0.0).to_numpy(dtype=float) for col in total_cols])
        relative_matrix = np.vstack(
            [pd.to_numeric(vel_df[col], errors="coerce").fillna(0.0).to_numpy(dtype=float) for col in relative_cols]
        )
        depth_array = np.asarray(depths, dtype=float)

        total_maxabs = np.max(np.abs(total_matrix), axis=1)
        total_pos = np.max(total_matrix, axis=1)
        total_neg = np.min(total_matrix, axis=1)
        relative_maxabs = np.max(np.abs(relative_matrix), axis=1)
        relative_pos = np.max(relative_matrix, axis=1)
        relative_neg = np.min(relative_matrix, axis=1)

        axis_label = str(axis_override or _infer_axis_label(file_name)).upper()
        if axis_label not in {"X", "Y", "SINGLE"}:
            axis_label = _infer_axis_label(file_name)
        summary_df = pd.DataFrame(
            {
                "Layer_Index": np.arange(1, len(layers) + 1, dtype=int),
                "Depth_m": depth_array,
                "Axis": [axis_label] * len(layers),
                "DB_Total_maxabs_m": total_maxabs,
                "DB_Total_pos_max_m": total_pos,
                "DB_Total_neg_min_m": total_neg,
                "DB_Relative_maxabs_m": relative_maxabs,
                "DB_Relative_pos_max_m": relative_pos,
                "DB_Relative_neg_min_m": relative_neg,
            }
        )

        return {
            "axis": axis_label,
            "recordLabel": _record_label_from_name(file_name) or Path(file_name).stem,
            "layer_numbers": np.asarray(layers, dtype=int),
            "depths": depth_array,
            "time": time,
            "disp_matrix": total_matrix,
            "relative_matrix": relative_matrix,
            "summary_df": summary_df,
            "total_time_df": _build_layer_time_df(time, depth_array, total_matrix, "db_total_m"),
            "relative_time_df": _build_layer_time_df(time, depth_array, relative_matrix, "db_relative_m"),
        }
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except OSError:
                pass


def process_db_single(
    file_bytes: bytes,
    file_name: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    _ = options
    bundle = _read_db_disp_bundle(file_bytes, file_name)
    summary_df = bundle["summary_df"].copy()
    output_bytes = build_db_single_output_workbook(
        summary_df,
        bundle["total_time_df"],
        bundle["relative_time_df"],
    )
    record_label = bundle["recordLabel"]
    output_file_name = f"output_db_single_{record_label}.xlsx"
    return {
        "pairKey": f"DB_SINGLE|{record_label}",
        "xFileName": file_name,
        "yFileName": "",
        "outputFileName": output_file_name,
        "outputBytes": output_bytes,
        "previewCharts": _db_single_preview_charts(bundle),
        "summaryCatalogEntry": _build_db_single_summary_entry(file_name, bundle),
        "metrics": {
            "mode": "db_single",
            "axis": bundle["axis"],
            "layerCount": int(len(summary_df)),
            "timeSeriesSheets": 2,
            "timeSheets": ["DB_Total_Time", "DB_Relative_Time"],
            "surfaceBaseTotal_m": float(summary_df["DB_Total_maxabs_m"].iloc[0]),
            "surfaceProfileRSS_m": float(summary_df["DB_Relative_maxabs_m"].iloc[0]),
            "baseReference": "db_direct",
            "integrationPrimary": "deepsoil_db",
        },
    }


def process_db_pair(
    x_bytes: bytes,
    y_bytes: bytes,
    x_name: str,
    y_name: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    normalized_options = _normalize_options(options)
    include_resultant_profiles = _include_resultant_profiles(normalized_options)

    x_bundle = _read_db_disp_bundle(x_bytes, x_name)
    y_bundle = _read_db_disp_bundle(y_bytes, y_name)
    n_layers = min(
        int(x_bundle["disp_matrix"].shape[0]),
        int(y_bundle["disp_matrix"].shape[0]),
        int(x_bundle["depths"].size),
        int(y_bundle["depths"].size),
    )
    if n_layers <= 0:
        raise ValueError(f"No common layers found in DB pair: {x_name}, {y_name}")

    depths = x_bundle["depths"][:n_layers]
    x_total = x_bundle["disp_matrix"][:n_layers, :]
    y_total = y_bundle["disp_matrix"][:n_layers, :]
    x_relative = x_bundle["relative_matrix"][:n_layers, :]
    y_relative = y_bundle["relative_matrix"][:n_layers, :]

    total_resultant_time_df = _build_resultant_time_df(
        {"time": x_bundle["time"], "depths": depths, "total_matrix": x_total},
        {"time": y_bundle["time"], "depths": depths, "total_matrix": y_total},
        matrix_key="total_matrix",
        value_suffix="db_total_resultant_m",
    )
    relative_resultant_time_df = _build_resultant_time_df(
        {"time": x_bundle["time"], "depths": depths, "relative_matrix": x_relative},
        {"time": y_bundle["time"], "depths": depths, "relative_matrix": y_relative},
        matrix_key="relative_matrix",
        value_suffix="db_relative_resultant_m",
    )

    total_resultant_max = np.max(np.abs(total_resultant_time_df.iloc[:, 1:].to_numpy(dtype=float)), axis=0)
    relative_resultant_max = np.max(np.abs(relative_resultant_time_df.iloc[:, 1:].to_numpy(dtype=float)), axis=0)

    summary_df = pd.DataFrame(
        {
            "Layer_Index": np.arange(1, n_layers + 1, dtype=int),
            "Depth_m": depths,
            "X_total_maxabs_m": np.max(np.abs(x_total), axis=1),
            "X_total_pos_max_m": np.max(x_total, axis=1),
            "X_total_neg_min_m": np.min(x_total, axis=1),
            "Y_total_maxabs_m": np.max(np.abs(y_total), axis=1),
            "Y_total_pos_max_m": np.max(y_total, axis=1),
            "Y_total_neg_min_m": np.min(y_total, axis=1),
            "X_relative_maxabs_m": np.max(np.abs(x_relative), axis=1),
            "X_relative_pos_max_m": np.max(x_relative, axis=1),
            "X_relative_neg_min_m": np.min(x_relative, axis=1),
            "Y_relative_maxabs_m": np.max(np.abs(y_relative), axis=1),
            "Y_relative_pos_max_m": np.max(y_relative, axis=1),
            "Y_relative_neg_min_m": np.min(y_relative, axis=1),
            "Total_resultant_maxabs_m": total_resultant_max,
            "Relative_resultant_maxabs_m": relative_resultant_max,
        }
    )

    output_bytes = build_db_pair_output_workbook(
        summary_df,
        _build_layer_time_df(x_bundle["time"], depths, x_total, "db_total_m"),
        _build_layer_time_df(y_bundle["time"], depths, y_total, "db_total_m"),
        _build_layer_time_df(x_bundle["time"], depths, x_relative, "db_relative_m"),
        _build_layer_time_df(y_bundle["time"], depths, y_relative, "db_relative_m"),
        include_resultant_profiles=include_resultant_profiles,
        total_resultant_time_df=total_resultant_time_df,
        relative_resultant_time_df=relative_resultant_time_df,
    )
    record_label = x_bundle["recordLabel"]
    output_file_name = f"output_db_pair_{record_label}.xlsx"
    return {
        "pairKey": f"DB|{_build_pair_key(x_name, y_name)}",
        "xFileName": x_name,
        "yFileName": y_name,
        "outputFileName": output_file_name,
        "outputBytes": output_bytes,
        "previewCharts": _db_pair_preview_charts(summary_df, x_bundle, y_bundle),
        "summaryCatalogEntry": _build_db_pair_summary_entry(x_name, y_name, summary_df),
        "metrics": {
            "mode": "db_pair",
            "layerCount": int(len(summary_df)),
            "timeSeriesSheets": 4 + (2 if include_resultant_profiles else 0),
            "timeSheets": (
                [
                    "DB_Total_X_Time",
                    "DB_Total_Y_Time",
                    "DB_Relative_X_Time",
                    "DB_Relative_Y_Time",
                ]
                + (
                    ["DB_Total_Resultant_Time", "DB_Relative_Resultant_Time"]
                    if include_resultant_profiles
                    else []
                )
            ),
            "surfaceBaseTotal_m": float(summary_df["Total_resultant_maxabs_m"].iloc[0]),
            "surfaceProfileRSS_m": float(summary_df["Relative_resultant_maxabs_m"].iloc[0]),
            "includeResultantProfiles": bool(include_resultant_profiles),
            "baseReference": "db_direct",
            "integrationPrimary": "deepsoil_db",
        },
    }


def _infer_axis_label(file_name: str) -> str:
    stem = Path(str(file_name).replace("\\", "/")).stem
    upper_name = stem.upper()
    if "_X_" in upper_name:
        return "X"
    if "_Y_" in upper_name:
        return "Y"
    if re.search(r"[_.-]X([_.-]|$)", upper_name) or upper_name.endswith("_X") or upper_name.startswith("X_"):
        return "X"
    if re.search(r"[_.-]Y([_.-]|$)", upper_name) or upper_name.endswith("_Y") or upper_name.startswith("Y_"):
        return "Y"
    if re.search(r"(HN1|H1|HNE|EW|000|180|270|360|225|210)$", upper_name) or re.search(
        r"[_.-](HN1|H1|HNE|E|W|EW|000|180|270|360|225|210)(?=[_.-]|$)",
        upper_name,
    ):
        return "X"
    if re.search(r"(HN2|H2|HNN|NS|090|045|135|315|300)$", upper_name) or re.search(
        r"[_.-](HN2|H2|HNN|N|S|NS|090|045|135|315|300)(?=[_.-]|$)",
        upper_name,
    ):
        return "Y"
    if re.search(r"\d(E|EW|W|X)$", upper_name):
        return "X"
    if re.search(r"\d(N|NS|S|Y)$", upper_name):
        return "Y"
    if re.search(r"HORIZ?ONTAL[_.\-\s]*1(?=[_.\-\s]|$)", upper_name):
        return "X"
    if re.search(r"HORIZ?ONTAL[_.\-\s]*2(?=[_.\-\s]|$)", upper_name):
        return "Y"
    if re.search(r"[A-Za-z]X\d+$", stem):
        return "X"
    if re.search(r"[A-Za-z]Y\d+$", stem):
        return "Y"
    return "SINGLE"


def _candidate_suffix(name: str) -> str:
    return Path(str(name)).suffix.lower()


def _candidate_kind(name: str) -> str:
    suffix = _candidate_suffix(name)
    if suffix == XLSX_SUFFIX:
        return "xlsx"
    if suffix in DB_SUFFIXES:
        return "db"
    return "unknown"


def _record_label_from_name(name: str) -> str:
    raw = str(name).replace("\\", "/").strip("/")
    if not raw:
        return ""
    path = Path(raw)
    stem = path.stem
    if stem.lower() in {"deepsoilout", "deepsoil"} and path.parent and str(path.parent) not in {"", "."}:
        return Path(str(path.parent)).name
    return stem


def _build_pair_key(x_name: str, y_name: str) -> str:
    x_stem = _record_label_from_name(x_name) or Path(x_name).stem
    y_stem = _record_label_from_name(y_name) or Path(y_name).stem
    base = x_stem.replace("_X_", "_").replace("_H1", "")
    return f"{base}|{y_stem}"


def _preview_select_indices(size: int, max_items: int) -> np.ndarray:
    if size <= 0:
        return np.array([], dtype=int)
    if size <= max_items:
        return np.arange(size, dtype=int)
    return np.unique(np.linspace(0, size - 1, num=max_items, dtype=int))


def _preview_points(
    x: np.ndarray | Sequence[float],
    y: np.ndarray | Sequence[float],
    *,
    max_points: int = PREVIEW_MAX_POINTS,
) -> List[Dict[str, float]]:
    x_arr = np.asarray(x, dtype=float)
    y_arr = np.asarray(y, dtype=float)
    mask = np.isfinite(x_arr) & np.isfinite(y_arr)
    x_arr = x_arr[mask]
    y_arr = y_arr[mask]
    if x_arr.size == 0 or y_arr.size == 0:
        return []

    indices = _preview_select_indices(int(x_arr.size), max_points)
    return [
        {"x": round(float(x_arr[idx]), 6), "y": round(float(y_arr[idx]), 6)}
        for idx in indices
    ]


def _preview_series(
    name: str,
    x: np.ndarray | Sequence[float],
    y: np.ndarray | Sequence[float],
    *,
    max_points: int = PREVIEW_MAX_POINTS,
) -> Dict[str, Any] | None:
    points = _preview_points(x, y, max_points=max_points)
    if not points:
        return None
    return {
        "name": str(name),
        "points": points,
    }


def _preview_chart(
    title: str,
    sheet_name: str,
    kind: str,
    series: Sequence[Dict[str, Any] | None],
    *,
    x_label: str,
    y_label: str,
    invert_y: bool = False,
) -> Dict[str, Any] | None:
    clean_series = [item for item in series if isinstance(item, dict) and item.get("points")]
    if not clean_series:
        return None
    return {
        "title": str(title),
        "sheetName": str(sheet_name),
        "kind": str(kind),
        "xLabel": str(x_label),
        "yLabel": str(y_label),
        "invertY": bool(invert_y),
        "series": clean_series,
    }


def _preview_layer_indices(depths: np.ndarray | Sequence[float]) -> List[int]:
    depth_arr = np.asarray(depths, dtype=float)
    size = int(depth_arr.size)
    if size <= 0:
        return []
    if size <= 3:
        return list(range(size))
    return sorted({0, size // 2, size - 1})


def _preview_layer_label(index: int, total: int, depth: float, axis_label: str | None = None) -> str:
    if index == 0:
        role = "Surface"
    elif index == total - 1:
        role = "Bottom"
    elif total > 2 and index == total // 2:
        role = "Mid"
    else:
        role = f"L{index + 1:02d}"
    axis_suffix = f" {axis_label}" if axis_label else ""
    return f"{role}{axis_suffix} @ {depth:.2f}m"


def _matrix_preview_chart(
    title: str,
    sheet_name: str,
    time: np.ndarray | Sequence[float],
    depths: np.ndarray | Sequence[float],
    matrix: np.ndarray,
    *,
    axis_label: str | None = None,
    y_label: str = "Displacement (m)",
) -> Dict[str, Any] | None:
    time_arr = np.asarray(time, dtype=float)
    depth_arr = np.asarray(depths, dtype=float)
    matrix_arr = np.asarray(matrix, dtype=float)
    if time_arr.size == 0 or matrix_arr.ndim != 2:
        return None

    n_layers = min(int(matrix_arr.shape[0]), int(depth_arr.size))
    if n_layers <= 0:
        return None

    series: List[Dict[str, Any] | None] = []
    for idx in _preview_layer_indices(depth_arr[:n_layers]):
        series.append(
            _preview_series(
                _preview_layer_label(idx, n_layers, float(depth_arr[idx]), axis_label),
                time_arr,
                matrix_arr[idx, :],
            )
        )

    return _preview_chart(
        title,
        sheet_name,
        "line",
        series,
        x_label="Time (s)",
        y_label=y_label,
    )


def _df_depth_preview_chart(
    title: str,
    sheet_name: str,
    frame: pd.DataFrame,
    series_specs: Sequence[Tuple[str, str]],
) -> Dict[str, Any] | None:
    if frame is None or frame.empty or "Depth_m" not in frame.columns:
        return None

    depths = pd.to_numeric(frame["Depth_m"], errors="coerce").to_numpy(dtype=float)
    series: List[Dict[str, Any] | None] = []
    for column_name, label in series_specs:
        if column_name not in frame.columns:
            continue
        values = pd.to_numeric(frame[column_name], errors="coerce").to_numpy(dtype=float)
        series.append(_preview_series(label, values, depths))

    return _preview_chart(
        title,
        sheet_name,
        "depth",
        series,
        x_label="Displacement (m)",
        y_label="Depth (m)",
        invert_y=True,
    )


def _preview_label_from_column(name: str) -> str:
    label = str(name)
    label = re.sub(r"_(maxabs|profile_rel|relative_maxabs|input_added_total)(_alt)?_m$", "", label, flags=re.IGNORECASE)
    label = re.sub(r"_m$", "", label, flags=re.IGNORECASE)
    return label


def _aggregate_depth_preview_chart(title: str, sheet_name: str, frame: pd.DataFrame) -> Dict[str, Any] | None:
    if frame is None or frame.empty or "Depth_m" not in frame.columns:
        return None

    data_columns = [column for column in frame.columns if column != "Depth_m"]
    if not data_columns:
        return None

    selected_columns = [data_columns[idx] for idx in _preview_select_indices(len(data_columns), PREVIEW_MAX_SERIES)]
    depths = pd.to_numeric(frame["Depth_m"], errors="coerce").to_numpy(dtype=float)
    series: List[Dict[str, Any] | None] = []
    for column_name in selected_columns:
        values = pd.to_numeric(frame[column_name], errors="coerce").to_numpy(dtype=float)
        series.append(_preview_series(_preview_label_from_column(column_name), values, depths))

    return _preview_chart(
        title,
        sheet_name,
        "depth",
        series,
        x_label="Displacement (m)",
        y_label="Depth (m)",
        invert_y=True,
    )


def _pair_preview_charts(
    comparison_df: pd.DataFrame,
    include_resultant_profiles: bool,
    x_direction_bundle: Mapping[str, Any],
    y_direction_bundle: Mapping[str, Any],
    strain_bundle: Mapping[str, Any],
) -> List[Dict[str, Any]]:
    charts: List[Dict[str, Any]] = []

    depth_specs = (
        [
            ("Total_base_rel_max_m", "Base RSS"),
            ("Total_tbdy_total_max_m", "TBDY RSS"),
            ("Profile_RSS_total_m", "Profile RSS"),
            ("TimeHist_Resultant_total_m", "TimeHist RSS"),
        ]
        if include_resultant_profiles
        else [
            ("X_base_rel_max_m", "Base X"),
            ("Y_base_rel_max_m", "Base Y"),
            ("X_tbdy_total_max_m", "TBDY X"),
            ("Y_tbdy_total_max_m", "TBDY Y"),
        ]
    )
    depth_chart = _df_depth_preview_chart("Depth Profiles", "Depth_Profiles", comparison_df, depth_specs)
    if depth_chart is not None:
        charts.append(depth_chart)

    base_corrected_df = _build_base_corrected_profiles_df(comparison_df)
    base_corrected_chart = _df_depth_preview_chart(
        "Base-Corrected Profiles",
        "Profile_BaseCorrected",
        base_corrected_df,
        [
            ("X_base_rel_max_m", "Base X"),
            ("Profile_X_minus_bottom_m", "Profile X adj"),
            ("Y_base_rel_max_m", "Base Y"),
            ("Profile_Y_minus_bottom_m", "Profile Y adj"),
        ],
    )
    if base_corrected_chart is not None:
        charts.append(base_corrected_chart)

    x_chart = _matrix_preview_chart(
        "Direction X Time",
        "Direction_X_Time",
        x_direction_bundle["time"],
        x_direction_bundle["depths"],
        x_direction_bundle["disp_matrix"],
        axis_label="X",
    )
    if x_chart is not None:
        charts.append(x_chart)

    y_chart = _matrix_preview_chart(
        "Direction Y Time",
        "Direction_Y_Time",
        y_direction_bundle["time"],
        y_direction_bundle["depths"],
        y_direction_bundle["disp_matrix"],
        axis_label="Y",
    )
    if y_chart is not None:
        charts.append(y_chart)

    resultant_matrix = np.sqrt(x_direction_bundle["disp_matrix"] ** 2 + y_direction_bundle["disp_matrix"] ** 2)
    resultant_chart = _matrix_preview_chart(
        "Resultant Time",
        "Resultant_Time",
        x_direction_bundle["time"],
        x_direction_bundle["depths"],
        resultant_matrix,
        y_label="Resultant (m)",
    )
    if resultant_chart is not None:
        charts.append(resultant_chart)

    tbdy_total_resultant = np.sqrt(strain_bundle["u_tbdy_total_x"] ** 2 + strain_bundle["u_tbdy_total_y"] ** 2)
    tbdy_chart = _matrix_preview_chart(
        "TBDY Total Resultant",
        "TBDY_Total_Resultant_Time",
        strain_bundle["time"],
        strain_bundle["depths"],
        tbdy_total_resultant,
        y_label="TBDY Total (m)",
    )
    if tbdy_chart is not None:
        charts.append(tbdy_chart)

    return charts


def _single_preview_charts(
    summary_df: pd.DataFrame,
    direction_bundle: Mapping[str, Any],
    strain_bundle: Mapping[str, Any],
) -> List[Dict[str, Any]]:
    charts: List[Dict[str, Any]] = []

    summary_chart = _df_depth_preview_chart(
        "Single Direction Summary",
        "Single_Direction_Summary",
        summary_df,
        [
            ("Base_rel_max_m", "Base Rel"),
            ("TBDY_total_max_m", "TBDY Total"),
            ("Profile_max_m", "Profile"),
            ("TimeHist_maxabs_m", "TimeHist"),
        ],
    )
    if summary_chart is not None:
        charts.append(summary_chart)

    direction_chart = _matrix_preview_chart(
        "Direction Time",
        "Direction_Time",
        direction_bundle["time"],
        direction_bundle["depths"],
        direction_bundle["disp_matrix"],
        axis_label=str(direction_bundle.get("axis", "")),
    )
    if direction_chart is not None:
        charts.append(direction_chart)

    strain_chart = _matrix_preview_chart(
        "Strain Relative Time",
        "Strain_Relative_Time",
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_rel_base"],
        y_label="Base Relative (m)",
    )
    if strain_chart is not None:
        charts.append(strain_chart)

    tbdy_chart = _matrix_preview_chart(
        "TBDY Total Time",
        "TBDY_Total_Time",
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_tbdy_total"],
        y_label="TBDY Total (m)",
    )
    if tbdy_chart is not None:
        charts.append(tbdy_chart)

    input_proxy_chart = _matrix_preview_chart(
        "Input Proxy Relative",
        "InputProxy_Relative_Time",
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_rel_input"],
        y_label="Input Proxy Relative (m)",
    )
    if input_proxy_chart is not None:
        charts.append(input_proxy_chart)

    return charts


def _method2_preview_charts(
    axis_label: str,
    strain_bundle: Mapping[str, Any],
) -> List[Dict[str, Any]]:
    chart = _matrix_preview_chart(
        f"Method-2 TBDY {axis_label}",
        _method2_sheet_name(axis_label),
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_tbdy_total"],
        axis_label=axis_label,
        y_label="TBDY Total (m)",
    )
    return [chart] if chart is not None else []


def _db_single_preview_charts(bundle: Mapping[str, Any]) -> List[Dict[str, Any]]:
    charts: List[Dict[str, Any]] = []
    summary_chart = _df_depth_preview_chart(
        "DB Summary",
        "DB_Summary",
        bundle["summary_df"],
        [
            ("DB_Total_maxabs_m", "DB Total"),
            ("DB_Relative_maxabs_m", "DB Relative"),
        ],
    )
    if summary_chart is not None:
        charts.append(summary_chart)

    total_chart = _matrix_preview_chart(
        "DB Total Time",
        "DB_Total_Time",
        bundle["time"],
        bundle["depths"],
        bundle["disp_matrix"],
        axis_label=str(bundle.get("axis", "")),
    )
    if total_chart is not None:
        charts.append(total_chart)

    relative_chart = _matrix_preview_chart(
        "DB Relative Time",
        "DB_Relative_Time",
        bundle["time"],
        bundle["depths"],
        bundle["relative_matrix"],
        axis_label=str(bundle.get("axis", "")),
        y_label="Relative (m)",
    )
    if relative_chart is not None:
        charts.append(relative_chart)

    return charts


def _db_pair_preview_charts(
    summary_df: pd.DataFrame,
    x_bundle: Mapping[str, Any],
    y_bundle: Mapping[str, Any],
) -> List[Dict[str, Any]]:
    charts: List[Dict[str, Any]] = []
    depth_chart = _df_depth_preview_chart(
        "DB Depth Profiles",
        "DB_Depth_Profiles",
        summary_df,
        [
            ("Total_resultant_maxabs_m", "DB Total RSS"),
            ("Relative_resultant_maxabs_m", "DB Relative RSS"),
            ("X_total_maxabs_m", "DB X Total"),
            ("Y_total_maxabs_m", "DB Y Total"),
        ],
    )
    if depth_chart is not None:
        charts.append(depth_chart)

    total_resultant = np.sqrt(x_bundle["disp_matrix"] ** 2 + y_bundle["disp_matrix"] ** 2)
    total_chart = _matrix_preview_chart(
        "DB Total Resultant",
        "DB_Total_Resultant_Time",
        x_bundle["time"],
        x_bundle["depths"],
        total_resultant,
        y_label="DB Total (m)",
    )
    if total_chart is not None:
        charts.append(total_chart)

    relative_resultant = np.sqrt(x_bundle["relative_matrix"] ** 2 + y_bundle["relative_matrix"] ** 2)
    relative_chart = _matrix_preview_chart(
        "DB Relative Resultant",
        "DB_Relative_Resultant_Time",
        x_bundle["time"],
        x_bundle["depths"],
        relative_resultant,
        y_label="DB Relative (m)",
    )
    if relative_chart is not None:
        charts.append(relative_chart)

    return charts


def _extract_method2_single(
    xlsx_bytes: bytes,
    file_name: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    normalized_options = _normalize_options(options)
    axis_label = _infer_axis_label(file_name)
    if axis_label not in {"X", "Y"}:
        return {
            "skipped": True,
            "reason": f"Axis could not be inferred from file name: {file_name}",
            "axis": axis_label,
        }

    with pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl") as xl:
        strain_bundle = _compute_single_strain_bundle(xl, axis_label, normalized_options)
        profile_sheet_view_df = _build_single_profile_sheet_df(xl)
        input_proxy = np.asarray(strain_bundle.get("u_input_proxy", np.array([])), dtype=float)
        input_motion_max_abs = float(np.max(np.abs(input_proxy))) if input_proxy.size else float("nan")
    return _build_method2_extract_from_bundle(
        file_name,
        axis_label,
        strain_bundle,
        profile_sheet_view_df,
        input_motion_max_abs,
        normalized_options,
    )


def _process_single_file_xlsx(
    xl: pd.ExcelFile,
    file_name: str,
    options: Mapping[str, Any] | None = None,
    *,
    include_method2_extract: bool = False,
) -> Dict[str, Any]:
    normalized_options = _normalize_options(options)
    axis_label = _infer_axis_label(file_name)

    strain_bundle = _compute_single_strain_bundle(xl, axis_label, normalized_options)
    direction_bundle = _compute_single_direction_disp_bundle(xl, axis_label, normalized_options)
    profile_depths, profile_max = _parse_profile_displacement_max(xl)
    profile_sheet_df = _build_single_profile_sheet_df(xl)

    summary_df = strain_bundle["summary_df"].copy()
    n_layers = min(
        len(summary_df),
        int(direction_bundle["disp_matrix"].shape[0]),
        int(profile_max.size),
        int(profile_depths.size),
    )
    summary_df = summary_df.iloc[:n_layers].copy()
    summary_df["Profile_max_m"] = np.abs(profile_max[:n_layers])
    summary_df["TimeHist_maxabs_m"] = np.max(
        np.abs(direction_bundle["disp_matrix"][:n_layers, :]),
        axis=1,
    )
    if direction_bundle.get("disp_matrix_alt") is not None:
        timehist_alt = np.max(np.abs(direction_bundle["disp_matrix_alt"][:n_layers, :]), axis=1)
        summary_df["TimeHist_maxabs_alt_m"] = timehist_alt
        summary_df["Delta_TimeHist_alt_minus_primary_m"] = timehist_alt - summary_df["TimeHist_maxabs_m"].to_numpy(
            dtype=float
        )
        with np.errstate(divide="ignore", invalid="ignore"):
            summary_df["Ratio_TimeHist_alt_to_primary"] = np.where(
                summary_df["TimeHist_maxabs_m"].to_numpy(dtype=float) != 0,
                timehist_alt / summary_df["TimeHist_maxabs_m"].to_numpy(dtype=float),
                np.nan,
            )

    strain_rel_time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_rel_base"],
        "base_rel_m",
    )
    tbdy_total_time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_tbdy_total"],
        "tbdy_total_m",
    )
    input_proxy_rel_time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_rel_input"],
        "input_proxy_rel_m",
    )
    direction_time_alt_df = direction_bundle.get("table_alt_df")
    strain_rel_time_alt_df = strain_rel_time_df.copy() if direction_time_alt_df is not None else None
    tbdy_total_time_alt_df = None
    input_proxy_rel_time_alt_df = None
    if strain_bundle.get("u_tbdy_total_alt") is not None:
        tbdy_total_time_alt_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            strain_bundle["u_tbdy_total_alt"],
            "tbdy_total_alt_m",
        )
    if strain_bundle.get("u_rel_input_alt") is not None:
        input_proxy_rel_time_alt_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            strain_bundle["u_rel_input_alt"],
            "input_proxy_rel_alt_m",
        )

    output_bytes = build_single_output_workbook(
        summary_df=summary_df,
        direction_time_df=direction_bundle["table_df"],
        profile_sheet_df=profile_sheet_df,
        strain_rel_time_df=strain_rel_time_df,
        tbdy_total_time_df=tbdy_total_time_df,
        input_proxy_rel_time_df=input_proxy_rel_time_df,
        direction_time_alt_df=direction_time_alt_df,
        strain_rel_time_alt_df=strain_rel_time_alt_df,
        tbdy_total_time_alt_df=tbdy_total_time_alt_df,
        input_proxy_rel_time_alt_df=input_proxy_rel_time_alt_df,
    )

    output_file_name = f"output_single_{Path(file_name).stem}.xlsx"
    integration_meta = strain_bundle.get("integration_meta", {})
    time_sheets = [
        "Direction_Time",
        "Strain_Relative_Time",
        "TBDY_Total_Time",
        "InputProxy_Relative_Time",
    ]
    if direction_bundle.get("table_alt_df") is not None:
        time_sheets.extend(
            [
                "Direction_Time_ALT",
                "Strain_Relative_Time_ALT",
                "TBDY_Total_Time_ALT",
                "InputProxy_Relative_Time_ALT",
            ]
        )

    method2_extracted = None
    if include_method2_extract:
        if axis_label in {"X", "Y"}:
            input_proxy = np.asarray(strain_bundle.get("u_input_proxy", np.array([])), dtype=float)
            input_motion_max_abs = float(np.max(np.abs(input_proxy))) if input_proxy.size else float("nan")
            method2_extracted = _build_method2_extract_from_bundle(
                file_name,
                axis_label,
                strain_bundle,
                profile_sheet_df,
                input_motion_max_abs,
                normalized_options,
            )
        else:
            method2_extracted = {
                "skipped": True,
                "reason": f"Axis could not be inferred from file name: {file_name}",
                "axis": axis_label,
            }

    return {
        "result": {
            "pairKey": f"SINGLE|{Path(file_name).stem}",
            "xFileName": file_name,
            "yFileName": "",
            "outputFileName": output_file_name,
            "outputBytes": output_bytes,
            "previewCharts": _single_preview_charts(summary_df, direction_bundle, strain_bundle),
            "metrics": {
                "mode": "single",
                "axis": axis_label,
                "baseReference": str(strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE)),
                "integrationPrimary": str(integration_meta.get("integrationPrimary", "cumtrapz")),
                "integrationCompareEnabled": bool(integration_meta.get("integrationCompareEnabled", False)),
                "altIntegrationMethod": integration_meta.get("altIntegrationMethod"),
                "altLowCutHz": integration_meta.get("altLowCutHz"),
                "layerCount": int(len(summary_df)),
                "timeSeriesSheets": len(time_sheets),
                "timeSheets": time_sheets,
                "surfaceBaseTotal_m": float(summary_df["Base_rel_max_m"].iloc[0]),
                "surfaceTBDYTotal_m": float(summary_df["TBDY_total_max_m"].iloc[0]),
                "surfaceTBDYTotalAlt_m": (
                    float(summary_df["TBDY_total_alt_max_m"].iloc[0])
                    if "TBDY_total_alt_max_m" in summary_df.columns
                    else float("nan")
                ),
                "surfaceProfileRSS_m": float(summary_df["Profile_max_m"].iloc[0]),
            },
        },
        "method2Extracted": method2_extracted,
        "sourceCatalogEntry": _build_single_source_catalog_entry(
            xl,
            file_name,
            axis_label,
            summary_df,
            direction_bundle,
            strain_bundle,
            profile_sheet_df,
        ),
        "summaryCatalogEntry": _build_single_summary_entry(
            xl,
            file_name,
            axis_label,
            summary_df,
            profile_sheet_df,
            strain_bundle,
            direction_bundle,
            normalized_options,
        ),
    }


def process_single_file(
    file_bytes: bytes,
    file_name: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    with pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl") as xl:
        payload = _process_single_file_xlsx(xl, file_name, options)
    return payload["result"]


def _process_xy_pair_xlsx(
    x_xl: pd.ExcelFile,
    y_xl: pd.ExcelFile,
    x_name: str,
    y_name: str,
    options: Mapping[str, Any] | None = None,
    *,
    include_method2_extract: bool = False,
) -> Dict[str, Any]:
    normalized_options = _normalize_options(options)
    include_resultant_profiles = _include_resultant_profiles(normalized_options)

    strain_bundle = _compute_strain_bundle(x_xl, y_xl, normalized_options)
    legacy_bundle = _compute_legacy_bundle(x_xl, y_xl, normalized_options)
    profile_sheet_df = _build_pair_profile_sheet_df(x_xl, y_xl)
    x_direction_bundle = _compute_single_direction_disp_bundle(x_xl, "X", normalized_options)
    y_direction_bundle = _compute_single_direction_disp_bundle(y_xl, "Y", normalized_options)

    strain_df = strain_bundle["summary_df"].copy()
    legacy_df = legacy_bundle["summary_df"].copy()
    comparison_df = _build_comparison_df(strain_df, legacy_df)
    resultant_time_df = _build_resultant_time_df(x_direction_bundle, y_direction_bundle)
    resultant_time_alt_df = None
    if x_direction_bundle.get("disp_matrix_alt") is not None and y_direction_bundle.get("disp_matrix_alt") is not None:
        resultant_time_alt_df = _build_resultant_time_df(
            x_direction_bundle,
            y_direction_bundle,
            matrix_key="disp_matrix_alt",
            value_suffix="resultant_alt_m",
        )
    tbdy_total_x_time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_tbdy_total_x"],
        "tbdy_total_x_m",
    )
    tbdy_total_y_time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        strain_bundle["u_tbdy_total_y"],
        "tbdy_total_y_m",
    )
    tbdy_total_resultant_matrix = np.sqrt(
        strain_bundle["u_tbdy_total_x"] ** 2 + strain_bundle["u_tbdy_total_y"] ** 2
    )
    tbdy_total_resultant_time_df = _build_layer_time_df(
        strain_bundle["time"],
        strain_bundle["depths"],
        tbdy_total_resultant_matrix,
        "tbdy_total_resultant_m",
    )
    tbdy_total_x_time_alt_df = None
    tbdy_total_y_time_alt_df = None
    tbdy_total_resultant_time_alt_df = None
    if strain_bundle.get("u_tbdy_total_x_alt") is not None and strain_bundle.get("u_tbdy_total_y_alt") is not None:
        tbdy_total_x_time_alt_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            strain_bundle["u_tbdy_total_x_alt"],
            "tbdy_total_x_alt_m",
        )
        tbdy_total_y_time_alt_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            strain_bundle["u_tbdy_total_y_alt"],
            "tbdy_total_y_alt_m",
        )
        tbdy_total_resultant_matrix_alt = np.sqrt(
            strain_bundle["u_tbdy_total_x_alt"] ** 2 + strain_bundle["u_tbdy_total_y_alt"] ** 2
        )
        tbdy_total_resultant_time_alt_df = _build_layer_time_df(
            strain_bundle["time"],
            strain_bundle["depths"],
            tbdy_total_resultant_matrix_alt,
            "tbdy_total_resultant_alt_m",
        )

    output_bytes = build_output_workbook(
        strain_df,
        legacy_df,
        comparison_df,
        include_resultant_profiles=include_resultant_profiles,
        profile_sheet_df=profile_sheet_df,
        x_time_df=x_direction_bundle["table_df"],
        y_time_df=y_direction_bundle["table_df"],
        resultant_time_df=resultant_time_df,
        tbdy_total_x_time_df=tbdy_total_x_time_df,
        tbdy_total_y_time_df=tbdy_total_y_time_df,
        tbdy_total_resultant_time_df=tbdy_total_resultant_time_df,
        x_time_alt_df=x_direction_bundle.get("table_alt_df"),
        y_time_alt_df=y_direction_bundle.get("table_alt_df"),
        resultant_time_alt_df=resultant_time_alt_df,
        tbdy_total_x_time_alt_df=tbdy_total_x_time_alt_df,
        tbdy_total_y_time_alt_df=tbdy_total_y_time_alt_df,
        tbdy_total_resultant_time_alt_df=tbdy_total_resultant_time_alt_df,
    )
    output_file_name = f"output_total_{Path(x_name).stem}.xlsx"
    integration_meta: Dict[str, Any] = {
        "integrationPrimary": "cumtrapz",
        "integrationCompareEnabled": False,
    }
    for source in (strain_bundle, legacy_bundle, x_direction_bundle, y_direction_bundle):
        candidate = source.get("integration_meta")
        if isinstance(candidate, dict):
            integration_meta.update(candidate)
    time_sheets = [
        "Direction_X_Time",
        "Direction_Y_Time",
        "Resultant_Time",
        "TBDY_Total_X_Time",
        "TBDY_Total_Y_Time",
        "TBDY_Total_Resultant_Time",
    ]
    if x_direction_bundle.get("table_alt_df") is not None and y_direction_bundle.get("table_alt_df") is not None:
        time_sheets.extend(
            [
                "Direction_X_Time_ALT",
                "Direction_Y_Time_ALT",
                "Resultant_Time_ALT",
                "TBDY_Total_X_Time_ALT",
                "TBDY_Total_Y_Time_ALT",
                "TBDY_Total_Resultant_Time_ALT",
            ]
        )

    method2_extracted: Dict[str, Dict[str, Any]] = {}
    if include_method2_extract:
        x_profile_sheet_view_df = _build_single_profile_sheet_df(x_xl)
        y_profile_sheet_view_df = _build_single_profile_sheet_df(y_xl)
        x_input_proxy = np.asarray(strain_bundle.get("u_input_proxy_x", np.array([])), dtype=float)
        y_input_proxy = np.asarray(strain_bundle.get("u_input_proxy_y", np.array([])), dtype=float)
        x_input_motion_max_abs = float(np.max(np.abs(x_input_proxy))) if x_input_proxy.size else float("nan")
        y_input_motion_max_abs = float(np.max(np.abs(y_input_proxy))) if y_input_proxy.size else float("nan")
        method2_extracted[x_name] = _build_method2_extract_from_bundle(
            x_name,
            "X",
            {
                "time": strain_bundle["time"],
                "depths": strain_bundle["depths"],
                "u_tbdy_total": strain_bundle["u_tbdy_total_x"],
                "u_tbdy_total_alt": strain_bundle.get("u_tbdy_total_x_alt"),
                "base_reference": strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE),
                "integration_meta": strain_bundle.get("integration_meta", {}),
            },
            x_profile_sheet_view_df,
            x_input_motion_max_abs,
            normalized_options,
        )
        method2_extracted[y_name] = _build_method2_extract_from_bundle(
            y_name,
            "Y",
            {
                "time": strain_bundle["time"],
                "depths": strain_bundle["depths"],
                "u_tbdy_total": strain_bundle["u_tbdy_total_y"],
                "u_tbdy_total_alt": strain_bundle.get("u_tbdy_total_y_alt"),
                "base_reference": strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE),
                "integration_meta": strain_bundle.get("integration_meta", {}),
            },
            y_profile_sheet_view_df,
            y_input_motion_max_abs,
            normalized_options,
        )

    return {
        "result": {
            "pairKey": _build_pair_key(x_name, y_name),
            "xFileName": x_name,
            "yFileName": y_name,
            "outputFileName": output_file_name,
            "outputBytes": output_bytes,
            "previewCharts": _pair_preview_charts(
                comparison_df,
                include_resultant_profiles,
                x_direction_bundle,
                y_direction_bundle,
                strain_bundle,
            ),
            "metrics": {
                "mode": "pair",
                "baseReference": str(strain_bundle.get("base_reference", DEFAULT_BASE_REFERENCE)),
                "includeResultantProfiles": bool(include_resultant_profiles),
                "integrationPrimary": str(integration_meta.get("integrationPrimary", "cumtrapz")),
                "integrationCompareEnabled": bool(integration_meta.get("integrationCompareEnabled", False)),
                "altIntegrationMethod": integration_meta.get("altIntegrationMethod"),
                "altLowCutHz": integration_meta.get("altLowCutHz"),
                "layerCount": int(len(strain_df)),
                "timeSeriesSheets": len(time_sheets),
                "timeSheets": time_sheets,
                "surfaceBaseTotal_m": float(strain_df["Total_base_rel_max_m"].iloc[0]),
                "surfaceTBDYTotal_m": float(strain_df["Total_tbdy_total_max_m"].iloc[0]),
                "surfaceTBDYTotalAlt_m": (
                    float(strain_df["Total_tbdy_total_alt_max_m"].iloc[0])
                    if "Total_tbdy_total_alt_max_m" in strain_df.columns
                    else float("nan")
                ),
                "surfaceProfileRSS_m": float(legacy_df["Profile_RSS_total_m"].iloc[0]),
            },
        },
        "method2Extracted": method2_extracted,
        "sourceCatalogEntry": None,
        "sourceCatalogEntries": _build_pair_source_catalog_entries(
            x_xl,
            y_xl,
            x_name,
            y_name,
            comparison_df,
            profile_sheet_df,
            x_direction_bundle,
            y_direction_bundle,
            strain_bundle,
        ),
        "summaryCatalogEntry": _build_pair_summary_entry(
            x_xl,
            y_xl,
            x_name,
            y_name,
            comparison_df,
            profile_sheet_df,
            strain_bundle,
            x_direction_bundle,
            y_direction_bundle,
            normalized_options,
        ),
    }


def process_xy_pair(
    x_bytes: bytes,
    y_bytes: bytes,
    x_name: str,
    y_name: str,
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    with pd.ExcelFile(io.BytesIO(x_bytes), engine="openpyxl") as x_xl, pd.ExcelFile(
        io.BytesIO(y_bytes), engine="openpyxl"
    ) as y_xl:
        payload = _process_xy_pair_xlsx(x_xl, y_xl, x_name, y_name, options)
    return payload["result"]


def _is_candidate_file(name: str, include_manip: bool) -> bool:
    lower_name = name.lower()
    suffix = _candidate_suffix(name)
    if suffix not in (XLSX_SUFFIX, *DB_SUFFIXES):
        return False
    if any(lower_name.startswith(prefix) for prefix in EXCLUDE_PREFIXES):
        return False
    if suffix == XLSX_SUFFIX and not include_manip and any(lower_name.endswith(item) for item in EXCLUDE_SUFFIXES):
        return False
    return True


def _derive_y_name(x_name: str) -> str:
    replaced = re.sub(r"_X_", "_Y_", x_name, count=1, flags=re.IGNORECASE)
    replaced = re.sub(r"_H1", "_H2", replaced, count=1, flags=re.IGNORECASE)
    if replaced != x_name:
        return replaced
    replaced = re.sub(r"(?i)(HORIZ?ONTAL[_.\-\s]*?)1(?=[_.\-\s]|$)", r"\g<1>2", x_name, count=1)
    if replaced != x_name:
        return replaced
    replaced = re.sub(r"X(\d+)(\.[^.]+)?$", r"Y\1\2", x_name, count=1, flags=re.IGNORECASE)
    return replaced


def find_xy_pairs(file_names: Sequence[str], include_manip: bool = False) -> Tuple[List[Tuple[str, str]], List[str]]:
    candidates = {name for name in file_names if _is_candidate_file(name, include_manip)}

    x_files = sorted([name for name in candidates if _infer_axis_label(name) == "X"])

    pairs: List[Tuple[str, str]] = []
    missing: List[str] = []

    for x_name in x_files:
        y_name = _derive_y_name(x_name)
        if y_name in candidates:
            pairs.append((x_name, y_name))
        else:
            missing.append(x_name)

    return pairs, missing


def _coerce_manual_pairs(raw_pairs: Any) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    if not isinstance(raw_pairs, (list, tuple)):
        return out

    for item in raw_pairs:
        x_name = ""
        y_name = ""
        if isinstance(item, Mapping):
            x_name = str(item.get("xName") or item.get("x") or "").strip()
            y_name = str(item.get("yName") or item.get("y") or "").strip()
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            x_name = str(item[0] or "").strip()
            y_name = str(item[1] or "").strip()

        if x_name and y_name:
            out.append((x_name, y_name))

    return out


def _resolve_xy_pairs(
    file_names: Sequence[str],
    *,
    include_manip: bool = False,
    manual_pairing_enabled: bool = False,
    manual_pairs: Any = None,
) -> Tuple[List[Tuple[str, str]], List[str], List[str]]:
    candidates = {name for name in file_names if _is_candidate_file(name, include_manip)}

    if not manual_pairing_enabled:
        pairs, missing = find_xy_pairs(file_names, include_manip=include_manip)
        return pairs, missing, []

    warnings: List[str] = []
    resolved_pairs: List[Tuple[str, str]] = []
    used_names: set[str] = set()
    raw_manual_pairs = _coerce_manual_pairs(manual_pairs)

    if not raw_manual_pairs:
        warnings.append("Manual pairing enabled but no valid X/Y selections were provided.")
        return [], [], warnings

    for x_name, y_name in raw_manual_pairs:
        if x_name not in candidates or y_name not in candidates:
            warnings.append(f"Manual pair ignored (missing candidate): {x_name} + {y_name}")
            continue
        x_key = x_name.lower()
        y_key = y_name.lower()
        if x_key in used_names or y_key in used_names:
            warnings.append(f"Manual pair ignored (candidate already used): {x_name} + {y_name}")
            continue
        used_names.add(x_key)
        used_names.add(y_key)
        resolved_pairs.append((x_name, y_name))

    return resolved_pairs, [], warnings


def process_batch_files(file_map: Mapping[str, bytes], options: Mapping[str, Any] | None = None) -> Dict[str, Any]:
    normalized_options = _normalize_options(options)
    if _use_db3_directly(normalized_options):
        return _process_db_batch_files(file_map, normalized_options)

    base_reference = _normalize_base_reference(normalized_options.get("baseReference", DEFAULT_BASE_REFERENCE))
    integration_cfg = _integration_compare_config(normalized_options)
    manual_pairing_enabled = _to_bool(normalized_options.get("manualPairingEnabled", False), False)
    fallback_options: Dict[str, Any] | None = None
    if base_reference == "deepest_layer":
        fallback_options = dict(normalized_options)
        fallback_options["baseReference"] = "input"
    include_manip = bool(normalized_options.get("includeManip", False))
    fail_fast = bool(normalized_options.get("failFast", False))
    primary_outputs_enabled = _primary_outputs_enabled(normalized_options)
    method23_outputs_enabled = _method23_outputs_enabled(normalized_options)
    method2_enabled = _to_bool(
        normalized_options.get("method2Enabled", normalized_options.get("method23Enabled", True)),
        True,
    ) and method23_outputs_enabled
    method3_enabled = _to_bool(
        normalized_options.get("method3Enabled", normalized_options.get("method23Enabled", True)),
        True,
    ) and method23_outputs_enabled
    include_resultant_profiles = _include_resultant_profiles(normalized_options)

    logs: List[Dict[str, str]] = []
    errors: List[Dict[str, str]] = []
    results: List[Dict[str, Any]] = []
    source_catalog: List[Dict[str, Any]] = []
    summary_catalog: List[Dict[str, Any]] = []
    return_web_results = _to_bool(normalized_options.get("_returnWebResults", False), False)

    def _store_result(result: Dict[str, Any]) -> None:
        if return_web_results:
            results.append(_build_web_result_payload(result))
        else:
            results.append(result)

    def _store_source_entry(entry: Dict[str, Any] | None) -> None:
        if isinstance(entry, dict) and entry.get("families"):
            source_catalog.append(entry)

    def _store_source_entries(entries: Any) -> None:
        if isinstance(entries, dict):
            _store_source_entry(entries)
            return
        if not isinstance(entries, (list, tuple)):
            return
        for item in entries:
            _store_source_entry(item)

    def _store_summary_entry(entry: Dict[str, Any] | None) -> None:
        if isinstance(entry, dict) and entry.get("variants"):
            summary_catalog.append(entry)

    file_names = sorted(file_map.keys())
    xlsx_candidates = sorted(
        [name for name in file_names if _candidate_kind(name) == "xlsx" and _is_candidate_file(name, include_manip)]
    )
    db_candidates = sorted(
        [name for name in file_names if _candidate_kind(name) == "db" and _is_candidate_file(name, include_manip)]
    )
    candidates = xlsx_candidates
    pairs, missing, pair_warnings = _resolve_xy_pairs(
        xlsx_candidates,
        include_manip=include_manip,
        manual_pairing_enabled=manual_pairing_enabled,
        manual_pairs=normalized_options.get("manualPairs"),
    )

    used_in_pairs = set()
    for x_name, y_name in pairs:
        used_in_pairs.add(x_name)
        used_in_pairs.add(y_name)

    singles = sorted([name for name in candidates if name not in used_in_pairs])

    _log(logs, "info", f"Candidate files: {len(candidates)} (xlsx={len(xlsx_candidates)}, db={len(db_candidates)})")
    _log(logs, "info", f"Detected X/Y pairs: {len(pairs)}")
    _log(logs, "info", f"Detected single files: {len(singles)}")
    _log(logs, "info", f"Manual pairing: {'on' if manual_pairing_enabled else 'off'}")
    _log(logs, "info", f"Method-2 output: {'on' if method2_enabled else 'off'}")
    _log(logs, "info", f"Method-3 output: {'on' if method3_enabled else 'off'}")
    _log(logs, "info", f"Depth profile resultants: {'on' if include_resultant_profiles else 'off'}")
    _log(logs, "info", f"Integration compare: {'on' if integration_cfg['enabled'] else 'off'}")
    if integration_cfg["enabled"]:
        _log(logs, "info", f"Alt integration method: {integration_cfg['method']} ({ALT_LOWCUT_POLICY})")
    _log(logs, "info", f"Base reference: {base_reference}")
    _log(logs, "info", f"Processing config: {_processing_summary_text(normalized_options)}")

    for warning in pair_warnings:
        _log(logs, "warning", warning)
    for missing_x in missing:
        if missing_x in singles:
            _log(logs, "warning", f"No Y match for X file; processing single: {missing_x}")
        else:
            _log(logs, "warning", f"No Y match for X file: {missing_x}")

    pair_processed = 0
    pair_failed = 0
    single_processed = 0
    single_failed = 0
    method2_detected = len(xlsx_candidates) if (method2_enabled or method3_enabled) else 0
    method2_processed = 0
    method2_failed = 0
    method3_produced = 0
    method3_failed = 0
    primary_step_count = (len(pairs) + len(singles)) if primary_outputs_enabled else 0
    progress_total = primary_step_count + method2_detected + (1 if method3_enabled and method2_detected > 0 else 0)
    progress_completed = 0
    method2_profile_x_frames: List[pd.DataFrame] = []
    method2_profile_y_frames: List[pd.DataFrame] = []
    method2_profile_x_alt_frames: List[pd.DataFrame] = []
    method2_profile_y_alt_frames: List[pd.DataFrame] = []
    method2_relative_profile_x_frames: List[pd.DataFrame] = []
    method2_relative_profile_y_frames: List[pd.DataFrame] = []
    method2_input_added_profile_x_frames: List[pd.DataFrame] = []
    method2_input_added_profile_y_frames: List[pd.DataFrame] = []
    prefetched_method2: Dict[str, Dict[str, Any]] = {}

    def _advance_progress(message: str) -> None:
        nonlocal progress_completed
        if progress_total <= 0:
            return
        progress_completed += 1
        _report_batch_progress(normalized_options, progress_completed, progress_total, message)

    if progress_total > 0:
        _report_batch_progress(
            normalized_options,
            0,
            progress_total,
            f"Batch calculation started (0/{progress_total})",
        )

    def _is_deepest_table_error(exc: Exception) -> bool:
        text = str(exc).lower()
        return "table index is out of bounds" in text or "table index out of bounds" in text

    def _consume_method2_extracted(extracted: Dict[str, Any]) -> str:
        nonlocal method2_processed

        if extracted.get("skipped", False):
            return "skipped"

        if method2_enabled:
            _store_result(extracted["result"])
            method2_processed += 1

        axis = str(extracted.get("axis", "")).upper()
        profile_df = extracted.get("profile_df")
        relative_profile_df = extracted.get("relative_profile_df")
        input_added_profile_df = extracted.get("input_added_profile_df")
        profile_alt_df = extracted.get("profile_alt_df")
        if method3_enabled and isinstance(profile_df, pd.DataFrame) and not profile_df.empty:
            if axis == "X":
                method2_profile_x_frames.append(profile_df)
            elif axis == "Y":
                method2_profile_y_frames.append(profile_df)
        if method3_enabled and isinstance(relative_profile_df, pd.DataFrame) and not relative_profile_df.empty:
            if axis == "X":
                method2_relative_profile_x_frames.append(relative_profile_df)
            elif axis == "Y":
                method2_relative_profile_y_frames.append(relative_profile_df)
        if method3_enabled and isinstance(input_added_profile_df, pd.DataFrame) and not input_added_profile_df.empty:
            if axis == "X":
                method2_input_added_profile_x_frames.append(input_added_profile_df)
            elif axis == "Y":
                method2_input_added_profile_y_frames.append(input_added_profile_df)
        if method3_enabled and isinstance(profile_alt_df, pd.DataFrame) and not profile_alt_df.empty:
            if axis == "X":
                method2_profile_x_alt_frames.append(profile_alt_df)
            elif axis == "Y":
                method2_profile_y_alt_frames.append(profile_alt_df)

        return "processed"

    if primary_outputs_enabled:
        for x_name, y_name in pairs:
            try:
                kind_x = _candidate_kind(x_name)
                kind_y = _candidate_kind(y_name)
                if kind_x == "xlsx" and kind_y == "xlsx":
                    with pd.ExcelFile(io.BytesIO(file_map[x_name]), engine="openpyxl") as x_xl, pd.ExcelFile(
                        io.BytesIO(file_map[y_name]), engine="openpyxl"
                    ) as y_xl:
                        pair_payload = _process_xy_pair_xlsx(
                            x_xl,
                            y_xl,
                            x_name,
                            y_name,
                            normalized_options,
                            include_method2_extract=bool(method2_enabled or method3_enabled),
                        )
                    result = pair_payload["result"]
                    prefetched_method2.update(pair_payload.get("method2Extracted", {}))
                    _store_source_entries(pair_payload.get("sourceCatalogEntries"))
                    _store_source_entry(pair_payload.get("sourceCatalogEntry"))
                    _store_summary_entry(pair_payload.get("summaryCatalogEntry"))
                elif kind_x == "db" and kind_y == "db":
                    result = process_db_pair(
                        file_map[x_name],
                        file_map[y_name],
                        x_name,
                        y_name,
                        normalized_options,
                    )
                    _store_summary_entry(result.get("summaryCatalogEntry"))
                else:
                    raise ValueError(f"Mismatched pair types are not supported: {x_name}, {y_name}")
                _store_result(result)
                pair_processed += 1
                _log(logs, "info", f"Processed pair: {x_name} + {y_name}")
                _advance_progress(f"Processed pair ({progress_completed + 1}/{progress_total}): {x_name} + {y_name}")
            except Exception as exc:  # noqa: BLE001
                if _candidate_kind(x_name) == "xlsx" and fallback_options is not None and _is_deepest_table_error(exc):
                    try:
                        with pd.ExcelFile(io.BytesIO(file_map[x_name]), engine="openpyxl") as x_xl, pd.ExcelFile(
                            io.BytesIO(file_map[y_name]), engine="openpyxl"
                        ) as y_xl:
                            pair_payload = _process_xy_pair_xlsx(
                                x_xl,
                                y_xl,
                                x_name,
                                y_name,
                                fallback_options,
                                include_method2_extract=bool(method2_enabled or method3_enabled),
                            )
                        result = pair_payload["result"]
                        prefetched_method2.update(pair_payload.get("method2Extracted", {}))
                        _store_source_entries(pair_payload.get("sourceCatalogEntries"))
                        _store_source_entry(pair_payload.get("sourceCatalogEntry"))
                        _store_summary_entry(pair_payload.get("summaryCatalogEntry"))
                        _store_result(result)
                        pair_processed += 1
                        _log(
                            logs,
                            "warning",
                            f"Deepest-layer base failed for pair ({x_name}, {y_name}); fallback to input base reference.",
                        )
                        _advance_progress(
                            f"Processed pair fallback ({progress_completed + 1}/{progress_total}): {x_name} + {y_name}"
                        )
                        continue
                    except Exception as fallback_exc:  # noqa: BLE001
                        exc = fallback_exc

                pair_failed += 1
                errors.append({"pairKey": f"{x_name}|{y_name}", "reason": str(exc)})
                _log(logs, "error", f"Failed pair {x_name} + {y_name}: {exc}")
                _advance_progress(f"Failed pair ({progress_completed + 1}/{progress_total}): {x_name} + {y_name}")
                if fail_fast:
                    break

    if primary_outputs_enabled and (not fail_fast or not errors):
        for name in singles:
            try:
                kind = _candidate_kind(name)
                if kind == "xlsx":
                    with pd.ExcelFile(io.BytesIO(file_map[name]), engine="openpyxl") as xl:
                        single_payload = _process_single_file_xlsx(
                            xl,
                            name,
                            normalized_options,
                            include_method2_extract=bool(method2_enabled or method3_enabled),
                        )
                    result = single_payload["result"]
                    extracted = single_payload.get("method2Extracted")
                    if isinstance(extracted, dict):
                        prefetched_method2[name] = extracted
                    _store_source_entries(single_payload.get("sourceCatalogEntries"))
                    _store_source_entry(single_payload.get("sourceCatalogEntry"))
                    _store_summary_entry(single_payload.get("summaryCatalogEntry"))
                elif kind == "db":
                    result = process_db_single(
                        file_map[name],
                        name,
                        normalized_options,
                    )
                    _store_summary_entry(result.get("summaryCatalogEntry"))
                else:
                    raise ValueError(f"Unsupported input file type: {name}")
                _store_result(result)
                single_processed += 1
                _log(logs, "info", f"Processed single: {name}")
                _advance_progress(f"Processed single ({progress_completed + 1}/{progress_total}): {name}")
            except Exception as exc:  # noqa: BLE001
                if _candidate_kind(name) == "xlsx" and fallback_options is not None and _is_deepest_table_error(exc):
                    try:
                        with pd.ExcelFile(io.BytesIO(file_map[name]), engine="openpyxl") as xl:
                            single_payload = _process_single_file_xlsx(
                                xl,
                                name,
                                fallback_options,
                                include_method2_extract=bool(method2_enabled or method3_enabled),
                            )
                        result = single_payload["result"]
                        extracted = single_payload.get("method2Extracted")
                        if isinstance(extracted, dict):
                            prefetched_method2[name] = extracted
                        _store_source_entries(single_payload.get("sourceCatalogEntries"))
                        _store_source_entry(single_payload.get("sourceCatalogEntry"))
                        _store_summary_entry(single_payload.get("summaryCatalogEntry"))
                        _store_result(result)
                        single_processed += 1
                        _log(
                            logs,
                            "warning",
                            f"Deepest-layer base failed for single ({name}); fallback to input base reference.",
                        )
                        _advance_progress(f"Processed single fallback ({progress_completed + 1}/{progress_total}): {name}")
                        continue
                    except Exception as fallback_exc:  # noqa: BLE001
                        exc = fallback_exc

                single_failed += 1
                errors.append({"pairKey": f"SINGLE|{name}", "reason": str(exc)})
                _log(logs, "error", f"Failed single {name}: {exc}")
                _advance_progress(f"Failed single ({progress_completed + 1}/{progress_total}): {name}")
                if fail_fast:
                    break

    if (method2_enabled or method3_enabled) and (not fail_fast or not errors):
        for name in xlsx_candidates:
            try:
                extracted = prefetched_method2.pop(name, None)
                if extracted is None:
                    extracted = _extract_method2_single(
                        file_map[name],
                        name,
                        normalized_options,
                    )
                consume_status = _consume_method2_extracted(extracted)
                if consume_status == "skipped":
                    _log(logs, "warning", str(extracted.get("reason", f"Skipped Method-2 file: {name}")))
                    _advance_progress(f"Skipped Method-2 basis file ({progress_completed + 1}/{progress_total}): {name}")
                    continue
                _log(logs, "info", f"Processed Method-2 basis file: {name}")
                _advance_progress(f"Processed Method-2 basis file ({progress_completed + 1}/{progress_total}): {name}")
            except Exception as exc:  # noqa: BLE001
                if fallback_options is not None and _is_deepest_table_error(exc):
                    try:
                        extracted = prefetched_method2.pop(name, None)
                        if extracted is None:
                            extracted = _extract_method2_single(
                                file_map[name],
                                name,
                                fallback_options,
                            )
                        consume_status = _consume_method2_extracted(extracted)
                        if consume_status == "skipped":
                            _log(logs, "warning", str(extracted.get("reason", f"Skipped Method-2 file: {name}")))
                            _advance_progress(
                                f"Skipped Method-2 basis file ({progress_completed + 1}/{progress_total}): {name}"
                            )
                            continue

                        _log(
                            logs,
                            "warning",
                            f"Deepest-layer base failed for Method-2 file ({name}); fallback to input base reference.",
                        )
                        _advance_progress(
                            f"Processed Method-2 fallback ({progress_completed + 1}/{progress_total}): {name}"
                        )
                        continue
                    except Exception as fallback_exc:  # noqa: BLE001
                        exc = fallback_exc

                method2_failed += 1
                errors.append({"pairKey": f"METHOD2|{name}", "reason": str(exc)})
                _log(logs, "error", f"Failed Method-2 file {name}: {exc}")
                _advance_progress(f"Failed Method-2 basis file ({progress_completed + 1}/{progress_total}): {name}")
                if fail_fast:
                    break

    if method3_enabled and (not fail_fast or not errors):
        profile_x_df = _merge_profile_frames(method2_profile_x_frames)
        profile_y_df = _merge_profile_frames(method2_profile_y_frames)
        profile_x_alt_df = _merge_profile_frames(method2_profile_x_alt_frames)
        profile_y_alt_df = _merge_profile_frames(method2_profile_y_alt_frames)
        relative_profile_x_df = _merge_profile_frames(method2_relative_profile_x_frames)
        relative_profile_y_df = _merge_profile_frames(method2_relative_profile_y_frames)
        input_added_profile_x_df = _merge_profile_frames(method2_input_added_profile_x_frames)
        input_added_profile_y_df = _merge_profile_frames(method2_input_added_profile_y_frames)
        profile_x_delta_df = _build_method3_delta_df(profile_x_df, profile_x_alt_df)
        profile_y_delta_df = _build_method3_delta_df(profile_y_df, profile_y_alt_df)

        if (
            not profile_x_df.empty
            or not profile_y_df.empty
            or not relative_profile_x_df.empty
            or not relative_profile_y_df.empty
            or not input_added_profile_x_df.empty
            or not input_added_profile_y_df.empty
        ):
            try:
                method3_bytes = _build_method3_aggregate_workbook(
                    profile_x_df,
                    profile_y_df,
                    profile_x_alt_df=profile_x_alt_df,
                    profile_y_alt_df=profile_y_alt_df,
                    profile_x_delta_df=profile_x_delta_df,
                    profile_y_delta_df=profile_y_delta_df,
                    relative_profile_x_df=relative_profile_x_df,
                    relative_profile_y_df=relative_profile_y_df,
                    input_added_profile_x_df=input_added_profile_x_df,
                    input_added_profile_y_df=input_added_profile_y_df,
                )
                _store_result(
                    {
                        "pairKey": "METHOD3|ALL",
                        "xFileName": "",
                        "yFileName": "",
                        "outputFileName": "output_method3_profiles_all.xlsx",
                        "outputBytes": method3_bytes,
                        "previewCharts": [
                            chart
                            for chart in (
                                _aggregate_depth_preview_chart("Method-3 X Profiles", "Method3_Profile_X", profile_x_df),
                                _aggregate_depth_preview_chart("Method-3 Y Profiles", "Method3_Profile_Y", profile_y_df),
                                _aggregate_depth_preview_chart(
                                    "Method-3 X Approx Total (Ubase + Urel)",
                                    "Method3_ApproxTotal_X",
                                    input_added_profile_x_df,
                                ),
                                _aggregate_depth_preview_chart(
                                    "Method-3 Y Approx Total (Ubase + Urel)",
                                    "Method3_ApproxTotal_Y",
                                    input_added_profile_y_df,
                                ),
                            )
                            if chart is not None
                        ],
                        "metrics": {
                            "mode": "method3_aggregate",
                            "baseReference": base_reference,
                            "integrationPrimary": "cumtrapz",
                            "integrationCompareEnabled": bool(integration_cfg["enabled"]),
                            "altIntegrationMethod": (
                                integration_cfg["method"] if integration_cfg["enabled"] else None
                            ),
                            "xDepthRows": int(len(profile_x_df)),
                            "yDepthRows": int(len(profile_y_df)),
                            "xProfileColumns": max(0, int(profile_x_df.shape[1]) - 1),
                            "yProfileColumns": max(0, int(profile_y_df.shape[1]) - 1),
                            "xProfileAltColumns": max(0, int(profile_x_alt_df.shape[1]) - 1),
                            "yProfileAltColumns": max(0, int(profile_y_alt_df.shape[1]) - 1),
                            "xRelativeProfileColumns": max(0, int(relative_profile_x_df.shape[1]) - 1),
                            "yRelativeProfileColumns": max(0, int(relative_profile_y_df.shape[1]) - 1),
                            "xInputAddedProfileColumns": max(0, int(input_added_profile_x_df.shape[1]) - 1),
                            "yInputAddedProfileColumns": max(0, int(input_added_profile_y_df.shape[1]) - 1),
                        },
                    }
                )
                _store_source_entry(
                    _build_method3_source_catalog_entry(
                        profile_x_df,
                        profile_y_df,
                        input_added_profile_x_df,
                        input_added_profile_y_df,
                    )
                )
                method3_produced = 1
                _log(logs, "info", "Produced Method-3 aggregate workbook: output_method3_profiles_all.xlsx")
                _advance_progress(
                    f"Produced Method-3 aggregate workbook ({progress_completed + 1}/{progress_total})"
                )
            except Exception as exc:  # noqa: BLE001
                method3_failed += 1
                errors.append({"pairKey": "METHOD3|ALL", "reason": str(exc)})
                _log(logs, "error", f"Failed Method-3 aggregate workbook: {exc}")
                _advance_progress(
                    f"Failed Method-3 aggregate workbook ({progress_completed + 1}/{progress_total})"
                )
        else:
            _log(logs, "warning", "Method-3 aggregate workbook skipped: no valid Method-2 profiles found.")
            _advance_progress(
                f"Skipped Method-3 aggregate workbook ({progress_completed + 1}/{progress_total})"
            )

    processed_total = pair_processed + single_processed + method2_processed + method3_produced
    failed_total = pair_failed + single_failed + method2_failed + method3_failed

    return {
        "results": results,
        "sourceCatalog": source_catalog,
        "summaryCatalog": summary_catalog,
        "logs": logs,
        "errors": errors,
        "metrics": {
            "pairsDetected": len(pairs),
            "pairsProcessed": pair_processed,
            "pairsFailed": pair_failed,
            "pairsMissing": len(missing),
            "dbCandidates": len(db_candidates),
            "xlsxCandidates": len(xlsx_candidates),
            "singlesDetected": len(singles),
            "singlesProcessed": single_processed,
            "singlesFailed": single_failed,
            "method2Enabled": bool(method2_enabled),
            "method3Enabled": bool(method3_enabled),
            "includeResultantProfiles": bool(include_resultant_profiles),
            "baseReference": base_reference,
            "integrationPrimary": "cumtrapz",
            "integrationCompareEnabled": bool(integration_cfg["enabled"]),
            "altIntegrationMethod": integration_cfg["method"] if integration_cfg["enabled"] else None,
            "altLowCutPolicy": ALT_LOWCUT_POLICY if integration_cfg["enabled"] else None,
            "manualPairingEnabled": bool(manual_pairing_enabled),
            "manualPairsApplied": len(pairs) if manual_pairing_enabled else 0,
            "method2Detected": method2_detected,
            "method2Processed": method2_processed,
            "method2Failed": method2_failed,
            "method3Produced": method3_produced,
            "processedTotal": processed_total,
            "failedTotal": failed_total,
        },
    }


def process_batch_directory(
    input_dir: str | os.PathLike[str],
    output_dir: str | os.PathLike[str],
    options: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    in_path = Path(input_dir)
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    file_map: Dict[str, bytes] = {}
    for item in in_path.rglob("*"):
        if not item.is_file():
            continue
        if item.name.startswith("~$"):
            continue
        if not _is_candidate_file(item.name, include_manip=bool((options or {}).get("includeManip", False))):
            continue
        relative_name = item.relative_to(in_path).as_posix()
        try:
            file_map[relative_name] = item.read_bytes()
        except PermissionError:
            continue

    summary = process_batch_files(file_map, options)

    for result in summary["results"]:
        target = out_path / result["outputFileName"]
        target.write_bytes(result["outputBytes"])
        result["writtenPath"] = str(target)

    return summary


__all__ = [
    "parse_profile_thickness",
    "compute_strain_relative",
    "compute_legacy_methods",
    "build_output_workbook",
    "build_single_output_workbook",
    "process_xy_pair",
    "process_single_file",
    "find_xy_pairs",
    "process_batch_files",
    "process_batch_directory",
]
